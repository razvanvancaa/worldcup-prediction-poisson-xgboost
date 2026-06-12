"""Build Excel v5 with Elo ratings, WC2026 predictions and match results."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Styles ──────────────────────────────────────────────────────────────────
HDR_FILL_ELO   = PatternFill('solid', fgColor='1F4E79')   # dark blue (Elo sheets)
HDR_FILL_PRED  = PatternFill('solid', fgColor='375623')   # dark green (predictions)
HDR_FILL_RES   = PatternFill('solid', fgColor='7B2C2C')   # dark red (results)
HDR_FONT       = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT     = Font(bold=True, size=11)
BORDER_THIN    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

WIN_FILL  = PatternFill('solid', fgColor='C6EFCE')  # green
DRAW_FILL = PatternFill('solid', fgColor='FFEB9C')  # yellow
LOSS_FILL = PatternFill('solid', fgColor='FFC7CE')  # red
BLUE_FILL = PatternFill('solid', fgColor='BDD7EE')  # light blue

def header_style(cell, fill):
    cell.font = HDR_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER_THIN

def cell_style(cell, bold=False, align='center'):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_THIN

def pct_color(cell, pct):
    """Color cell by win% value."""
    if pct is None:
        return
    if pct >= 70:
        cell.fill = PatternFill('solid', fgColor='00B050')  # strong green
    elif pct >= 55:
        cell.fill = WIN_FILL
    elif pct >= 45:
        cell.fill = DRAW_FILL
    elif pct >= 30:
        cell.fill = PatternFill('solid', fgColor='FF9B9B')
    else:
        cell.fill = LOSS_FILL

def add_elo_echipe(wb, wc_ratings):
    """Sheet 1: WC2026 team Elo ratings."""
    ws = wb.create_sheet('Elo_Echipe')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = [
        'Rang Global', 'Echipa', 'Elo Actual', 'Elo Max', 'Elo Mediu', 'Elo Min',
        'Schimb 1Y Elo', 'Schimb 2Y Elo', 'Schimb 5Y Elo',
        'Total Meciuri', 'Victorii', 'Egaluri', 'Infrangeri',
        'Goluri Marcate', 'Goluri Primite', 'Dif. Goluri',
        'Rata Victorii %', 'Med. Goluri/Meci',
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        header_style(cell, HDR_FILL_ELO)

    col_widths = [12, 24, 12, 10, 10, 10, 14, 14, 14, 14, 11, 11, 13, 16, 15, 14, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    for r_idx, row in enumerate(wc_ratings, 2):
        total = row.get('total_matches') or 0
        wins  = row.get('wins') or 0
        draws = row.get('draws') or 0
        losses= row.get('losses') or 0
        gf    = row.get('goals_for') or 0
        ga    = row.get('goals_against') or 0
        win_rate = round(wins / total * 100, 1) if total else None
        gpg = round(gf / total, 2) if total else None

        values = [
            row.get('global_rank'),
            row.get('team'),
            row.get('rating'),
            row.get('max_rating'),
            row.get('avg_rating'),
            row.get('min_rating'),
            row.get('rating_1y_chg'),
            row.get('rating_2y_chg'),
            row.get('rating_5y_chg'),
            total, wins, draws, losses,
            gf, ga, (gf - ga),
            win_rate, gpg,
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell_style(cell, align='center' if c_idx != 2 else 'left')

        # Color team cell by Elo tier
        elo = row.get('rating') or 0
        if elo >= 2000:
            ws.cell(r_idx, 3).fill = PatternFill('solid', fgColor='FFD700')  # gold
        elif elo >= 1900:
            ws.cell(r_idx, 3).fill = WIN_FILL
        elif elo >= 1800:
            ws.cell(r_idx, 3).fill = BLUE_FILL

        # Color Elo changes
        for col in [7, 8, 9]:
            chg = ws.cell(r_idx, col).value
            if chg is not None:
                ws.cell(r_idx, col).fill = WIN_FILL if chg > 0 else (LOSS_FILL if chg < 0 else PatternFill())

        # Color win rate
        if win_rate is not None:
            pct_color(ws.cell(r_idx, 17), win_rate)


def add_wc_predictii(wb, fixtures):
    """Sheet 2: WC2026 fixture predictions."""
    ws = wb.create_sheet('WC_Predictii')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = [
        'Data', 'Acasa', 'Deplasare', 'Competitie', 'Locatie',
        'Rang Acasa', 'Rang Deplasare',
        'Elo Acasa', 'Elo Deplasare', 'Dif. Elo',
        'Sansa Acasa %', 'Sansa Deplasare %',
        'Schimb Elo (Egal)',
        'Elo schimb (Acasa +1)', 'Elo schimb (Deplasare +1)',
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        header_style(cell, HDR_FILL_PRED)

    col_widths = [12, 24, 24, 20, 18, 13, 15, 11, 12, 11, 14, 16, 16, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35

    for r_idx, row in enumerate(fixtures, 2):
        h_elo = row.get('home_elo')
        a_elo = row.get('away_elo')
        dif = (h_elo - a_elo) if (h_elo and a_elo) else None
        values = [
            row.get('date'),
            row.get('home'),
            row.get('away'),
            row.get('tournament'),
            row.get('venue'),
            row.get('home_rank'),
            row.get('away_rank'),
            h_elo, a_elo, dif,
            row.get('win_exp_home_pct'),
            row.get('win_exp_away_pct'),
            row.get('draw_elo_change'),
            row.get('home_win1_elo'),
            row.get('away_win1_elo'),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            align = 'left' if c_idx in (2, 3, 4, 5) else 'center'
            cell_style(cell, align=align)

        # Color win expectancy cells
        wp_h = row.get('win_exp_home_pct')
        wp_a = row.get('win_exp_away_pct')
        pct_color(ws.cell(r_idx, 11), wp_h)
        pct_color(ws.cell(r_idx, 12), wp_a)

        # Highlight big Elo differences
        if dif is not None:
            if abs(dif) >= 200:
                ws.cell(r_idx, 10).fill = PatternFill('solid', fgColor='FF9800')


def add_wc_rezultate(wb, results):
    """Sheet 3: WC2026 match results."""
    ws = wb.create_sheet('WC_Calificari_Rezultate')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = [
        'Data', 'Acasa', 'Deplasare', 'Goluri Acasa', 'Goluri Deplasare',
        'Competitie', 'Locatie',
        'Elo Acasa', 'Elo Deplasare', 'Dif. Elo',
        'Schimb Elo (Acasa)', 'Rang Acasa', 'Rang Deplasare',
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        header_style(cell, HDR_FILL_RES)

    col_widths = [12, 24, 24, 14, 16, 22, 18, 11, 12, 11, 16, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for r_idx, row in enumerate(results, 2):
        h_s = row.get('home_score')
        a_s = row.get('away_score')
        h_elo = row.get('home_elo')
        a_elo = row.get('away_elo')
        dif = (h_elo - a_elo) if (h_elo and a_elo) else None

        values = [
            row.get('date'), row.get('home'), row.get('away'),
            h_s, a_s,
            row.get('tournament'), row.get('venue'),
            h_elo, a_elo, dif,
            row.get('elo_change'),
            row.get('home_rank'), row.get('away_rank'),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            align = 'left' if c_idx in (2, 3, 6, 7) else 'center'
            cell_style(cell, align=align)

        # Color result
        if h_s is not None and a_s is not None:
            if h_s > a_s:
                fill = WIN_FILL
            elif h_s < a_s:
                fill = LOSS_FILL
            else:
                fill = DRAW_FILL
            ws.cell(r_idx, 4).fill = fill
            ws.cell(r_idx, 5).fill = fill


def add_rezultate_recente(wb, matches):
    """Sheet 4: Recent results (2022-2025) for WC2026 teams."""
    ws = wb.create_sheet('Rezultate_Recente_2022-2025')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = [
        'Data', 'Acasa', 'Deplasare', 'Goluri Acasa', 'Goluri Deplasare',
        'Competitie', 'Locatie',
        'Elo Acasa', 'Elo Deplasare', 'Dif. Elo',
        'Schimb Elo (Acasa)', 'Rang Acasa', 'Rang Deplasare',
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        header_style(cell, HDR_FILL_RES)

    col_widths = [12, 24, 24, 14, 16, 22, 18, 11, 12, 11, 16, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for r_idx, row in enumerate(matches, 2):
        h_s = row.get('home_score')
        a_s = row.get('away_score')
        h_elo = row.get('home_elo')
        a_elo = row.get('away_elo')
        dif = (h_elo - a_elo) if (h_elo and a_elo) else None

        values = [
            row.get('date'), row.get('home'), row.get('away'),
            h_s, a_s,
            row.get('tournament'), row.get('venue'),
            h_elo, a_elo, dif,
            row.get('elo_change'),
            row.get('home_rank'), row.get('away_rank'),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            align = 'left' if c_idx in (2, 3, 6, 7) else 'center'
            cell_style(cell, align=align)

        # Light result coloring (no heavy fills for large sheets)
        if h_s is not None and a_s is not None:
            if h_s > a_s:
                ws.cell(r_idx, 4).fill = PatternFill('solid', fgColor='E2EFDA')
                ws.cell(r_idx, 5).fill = PatternFill('solid', fgColor='E2EFDA')
            elif h_s < a_s:
                ws.cell(r_idx, 4).fill = PatternFill('solid', fgColor='FCE4E4')
                ws.cell(r_idx, 5).fill = PatternFill('solid', fgColor='FCE4E4')


def build_v5():
    print('Loading Elo data...')
    with open('elo_data.json', encoding='utf-8') as f:
        data = json.load(f)

    print('Loading existing v4 workbook...')
    wb = openpyxl.load_workbook('statistici_calificari_cm2026_v4.xlsx')

    # Remove sheets if they already exist (re-run safety)
    for name in ['Elo_Echipe', 'WC_Predictii', 'WC_Calificari_Rezultate', 'Rezultate_Recente_2022-2025']:
        if name in wb.sheetnames:
            del wb[name]

    print(f'Adding Elo_Echipe ({len(data["wc_ratings"])} teams)...')
    add_elo_echipe(wb, data['wc_ratings'])

    print(f'Adding WC_Predictii ({len(data["wc_fixtures"])} fixtures)...')
    add_wc_predictii(wb, data['wc_fixtures'])

    print(f'Adding WC_Calificari_Rezultate ({len(data["wc_results"])} matches)...')
    add_wc_rezultate(wb, data['wc_results'])

    print(f'Adding Rezultate_Recente_2022-2025 ({len(data["recent_matches"])} matches)...')
    add_rezultate_recente(wb, data['recent_matches'])

    out = 'statistici_calificari_cm2026_v5.xlsx'
    wb.save(out)
    print(f'\nSaved: {out}')
    print(f'Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    build_v5()
