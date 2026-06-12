"""WC2026 Model V3 — XGBoost + Poisson Dixon-Coles Optimizat + Matrice 8x8

Imbunatatiri fata de v2:
  1. Implementare Ajustare Dixon-Coles (parametrul RHO optimizat) pt scoruri mici.
  2. Motor bazat pe Matrice de Probabilitati Joint (8x8) -> Precizie 100% pt Over/GG.
  3. Eliminare COMPLETA Data Leakage (scoase variabilele ai_gpg statice).
  4. Curatarea dependentelor statice pentru generarea pariurilor (Cornere/Cartonase).
  5. Include Sheet-ul "Cel Mai Bun Pariu" pentru vizualizare rapida.
"""
import json, math, re, warnings
from collections import defaultdict
import numpy as np
from scipy.optimize import minimize

try:
    from sklearn.ensemble import GradientBoostingClassifier
    from sklearn.calibration import CalibratedClassifierCV, calibration_curve
    from sklearn.metrics import log_loss, brier_score_loss, accuracy_score
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False
    print("ATENTIE: sklearn nu e instalat — ruleaza: pip install scikit-learn scipy")

try:
    import xgboost as xgb
    HAS_XGB = True
except ImportError:
    HAS_XGB = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Date ────────────────────────────────────────────────────────────────────
with open('aiscore_team_stats.json', encoding='utf-8') as f:
    AISCORE_RAW = json.load(f)
with open('elo_data.json', encoding='utf-8') as f:
    ELO_RAW = json.load(f)

ALIAS = {
    "Turkiye": "Turkey", "Türkiye": "Turkey",
    "Cote d'Ivoire": "Ivory Coast", "Côte d'Ivoire": "Ivory Coast",
    "IR Iran": "Iran", "Democratic Republic of the Congo": "DR Congo",
    "USA": "United States", "Curacao": "Curaçao", "Cabo Verde": "Cape Verde",
    "Bosnia and Herzegovina": "Bosnia and Herzegovina",
    "Bosnia & Herzegovina": "Bosnia and Herzegovina",
    "Korea Republic": "South Korea", "Republic of Korea": "South Korea",
}
def canon(n): return ALIAS.get(n, n)

ELO = {canon(r['team']): r for r in ELO_RAW['wc_ratings']}

# Flatten AiScore la dict canonical_name -> stats
AISCORE = {}
for _conf, _teams in AISCORE_RAW.items():
    for _name, _stats in _teams.items():
        _c = canon(_name)
        if _c not in AISCORE:
            AISCORE[_c] = {'_conf': _conf}
        for _k, _v in _stats.items():
            if _k not in AISCORE[_c]:
                AISCORE[_c][_k] = _v

def _num(s):
    if s is None: return None
    m = re.match(r'^(-?\d+(?:\.\d+)?)', str(s).strip().replace('−', '-'))
    return float(m.group(1)) if m else None

def get_conf(team):
    return AISCORE.get(team, {}).get('_conf', 'UEFA')

CONF_GAMES = {'CONMEBOL': 18, 'UEFA': 10, 'CAF': 10, 'AFC': 18, 'CONCACAF': 18, 'OFC': 4}

# Factori per-confederatie: calificarile CAF/OFC vs adversari slabi =
# statistici inflate fata de UEFA/CONMEBOL cu adversari puternici
CONF_QUAL_TO_WC = {
    'CONMEBOL': 0.90, 'UEFA': 0.85, 'CONCACAF': 0.78,
    'CAF': 0.70, 'AFC': 0.73, 'OFC': 0.62,
}

def per_game_stat(team, stat):
    val = _num(AISCORE.get(team, {}).get(stat))
    if val is None: return None
    conf = get_conf(team)
    base = CONF_GAMES.get(conf, 10)
    shots = _num(AISCORE.get(team, {}).get('total_shots'))
    g = max(4, min(base, round(shots / 13.0))) if shots and shots > 0 else base
    return val / g if g > 0 else None

def _qual_gpg(team):
    gpg = per_game_stat(team, 'goals') or 1.4
    return gpg * CONF_QUAL_TO_WC.get(get_conf(team), 0.85)

# Pre-calculat o singura data pentru viteza
QUAL_GOALS = {team: _qual_gpg(team) for team in AISCORE}

# ─── Conditii teren / stadion ────────────────────────────────────────────────
# climate: 'controlled'=indoor cu AC | 'extreme_heat'=>35C open | 'hot_humid' |
#          'hot' | 'mild' | 'cool'
STADIUM_DB = {
    'Estadio Azteca':           {'alt': 2240, 'climate': 'mild',         'controlled': False},
    'Estadio Akron':            {'alt': 1566, 'climate': 'warm',         'controlled': False},
    'Estadio BBVA':             {'alt':  540, 'climate': 'extreme_heat', 'controlled': False},
    'AT&T Stadium':             {'alt':   80, 'climate': 'hot',          'controlled': True},
    'NRG Stadium':              {'alt':   15, 'climate': 'hot_humid',    'controlled': True},
    'Hard Rock Stadium':        {'alt':    5, 'climate': 'hot_humid',    'controlled': False},
    'MetLife Stadium':          {'alt':   10, 'climate': 'mild',         'controlled': False},
    'Lincoln Financial Field':  {'alt':    6, 'climate': 'mild',         'controlled': False},
    'Gillette Stadium':         {'alt':   30, 'climate': 'mild',         'controlled': False},
    'Arrowhead Stadium':        {'alt':  230, 'climate': 'hot_humid',    'controlled': False},
    'Allegiant Stadium':        {'alt':  610, 'climate': 'extreme_heat', 'controlled': True},
    "Levi's Stadium":           {'alt':   17, 'climate': 'mild',         'controlled': False},
    'SoFi Stadium':             {'alt':    5, 'climate': 'mild',         'controlled': False},
    'Lumen Field':              {'alt':    3, 'climate': 'cool',         'controlled': False},
    'BC Place':                 {'alt':    5, 'climate': 'mild',         'controlled': True},
    'BMO Field':                {'alt':   76, 'climate': 'mild',         'controlled': False},
    'Mercedes-Benz Stadium':    {'alt':  290, 'climate': 'hot_humid',    'controlled': True},
}

# Altitudinea de antrenament a echipelor (m) — 0 = nivel mare
ALTITUDE_ADAPTED = {
    'Mexico': 2240, 'Colombia': 2000, 'Ecuador': 2800, 'Peru': 1500,
    'Bolivia': 3600,  # nu e la CM2026, dar in DB pentru completitudine
}

# Echipe adaptate la caldura extrema (clima calda/umeda in tara de origine)
HEAT_ADAPTED = {
    'Saudi Arabia', 'Iran', 'Iraq', 'Egypt', 'Tunisia', 'Algeria', 'Morocco',
    'Cape Verde', 'Senegal', 'Ivory Coast', 'Ghana', 'DR Congo', 'South Africa',
    'Panama', 'Haiti', 'Honduras', 'El Salvador', 'Costa Rica', 'Curaçao',
    'Brazil', 'Mexico', 'Australia', 'New Zealand', 'Indonesia', 'Qatar',
    'United Arab Emirates', 'Jordan',
}

def venue_conditions(home, away, stadium_name):
    """Returneaza factori de ajustare xG si cartonase bazati pe conditiile stadionului."""
    s = STADIUM_DB.get(stadium_name, {'alt': 0, 'climate': 'mild', 'controlled': False})
    alt, climate, controlled = s['alt'], s['climate'], s['controlled']

    def alt_factor(team):
        if controlled: return 1.0
        team_alt = ALTITUDE_ADAPTED.get(team, 0)
        if team_alt >= alt * 0.7: return 1.0   # echipa adaptata la aceasta altitudine
        # Penalizare ~6% la 2240m, ~3% la 1566m, 0 sub 1000m
        penalty = max(0.0, (alt - 1000) / 3000) * 0.15
        return max(0.85, 1.0 - penalty)

    def heat_factor(team):
        if controlled or climate in ('mild', 'cool', 'warm'): return 1.0
        adapted = team in HEAT_ADAPTED
        if climate == 'extreme_heat': return 0.98 if adapted else 0.92
        if climate == 'hot_humid':    return 0.99 if adapted else 0.95
        if climate == 'hot':          return 1.00 if adapted else 0.97
        return 1.0

    lh_m = alt_factor(home) * heat_factor(home)
    la_m = alt_factor(away) * heat_factor(away)

    # Caldura si altitudinea cresc oboseala → mai multe cartonase
    yc_m = 1.0
    if not controlled:
        if climate == 'extreme_heat': yc_m *= 1.12
        elif climate in ('hot_humid', 'hot'): yc_m *= 1.06
        if alt >= 1500: yc_m *= 1.08
        elif alt >= 800: yc_m *= 1.04

    cond_parts = []
    if not controlled and alt >= 800:
        cond_parts.append(f"Alt {alt}m")
    if not controlled and climate in ('extreme_heat', 'hot_humid'):
        cond_parts.append(climate.replace('_', ' '))

    return {
        'lh_mult': round(lh_m, 3), 'la_mult': round(la_m, 3),
        'yc_mult': round(yc_m, 3),
        'stadium': stadium_name, 'alt': alt,
        'climate': climate, 'controlled': controlled,
        'note': ', '.join(cond_parts) if cond_parts else 'normal',
    }

# Venue per meci Etapa 1
MATCH_VENUES = {
    ('Mexico', 'South Africa'):            'Estadio Azteca',
    ('South Korea', 'Czechia'):            'Estadio Akron',
    ('Canada', 'Bosnia and Herzegovina'):  'BMO Field',
    ('United States', 'Paraguay'):         'SoFi Stadium',
    ('Qatar', 'Switzerland'):              "Levi's Stadium",
    ('Brazil', 'Morocco'):                 'MetLife Stadium',
    ('Haiti', 'Scotland'):                 'Gillette Stadium',
    ('Australia', 'Turkey'):               'BC Place',
    ('Germany', 'Curaçao'):               'NRG Stadium',
    ('Netherlands', 'Japan'):              'AT&T Stadium',
    ('Ivory Coast', 'Ecuador'):            'Lincoln Financial Field',
    ('Sweden', 'Tunisia'):                 'Estadio BBVA',
    ('Spain', 'Cape Verde'):               'Mercedes-Benz Stadium',
    ('Belgium', 'Egypt'):                  'Lumen Field',
    ('Saudi Arabia', 'Uruguay'):           'Hard Rock Stadium',
    ('Iran', 'New Zealand'):               'SoFi Stadium',
    ('France', 'Senegal'):                 'MetLife Stadium',
    ('Iraq', 'Norway'):                    'Gillette Stadium',
    ('Argentina', 'Algeria'):              'Arrowhead Stadium',
    ('Austria', 'Jordan'):                 "Levi's Stadium",
    ('Portugal', 'DR Congo'):              'NRG Stadium',
    ('England', 'Croatia'):                'AT&T Stadium',
    ('Ghana', 'Panama'):                   'BMO Field',
    ('Uzbekistan', 'Colombia'):            'Estadio Azteca',
}

# ─── Absente jucatori cheie (umplut manual de utilizator) ────────────────────
# Format: team -> {'attack': 0.90, 'defense': 1.05}
# attack < 1.0 = atacant cheie lipseste  |  defense > 1.0 = aparare slabita
ABSENCES = {
    # Exemplu: 'France': {'attack': 0.85}  # daca Mbappe lipseste
}

# ─── Cote bookmaker (colectate automat 12 Iunie 2026) ────────────────────────
# h=acasa  d=egal  a=deplasare  o25=over2.5  u25=under2.5  gg=btts da  ng=btts nu
# Surse: Oddschecker / BetOnline / bet365 / OddsShark (best-price agregate)
# Modelul normalizeaza automat marja prin market_implied() — cotele raman utile
# chiar daca provin din bookmaker-i diferiti.
# ATENTIE: Mexico vs SA (2-0) si South Korea vs Czechia au fost deja jucate.
MARKET_ODDS = {
    # ── June 12, 2026 ─────────────────────────────────────────────────────────
    ('South Korea', 'Czechia'):            {'h': 2.60, 'd': 3.00, 'a': 3.00,
                                            'o15': 1.30, 'u15': 3.40,
                                            'o25': 2.20, 'u25': 1.63,
                                            'o35': 3.75, 'u35': 1.29,
                                            'gg': 1.81, 'ng': 1.89,
                                            'dc_1x': 1.60, 'dc_x2': 1.60, 'dc_12': 1.25},
    ('Canada', 'Bosnia and Herzegovina'):  {'h': 1.80, 'd': 3.50, 'a': 4.40,
                                            'o15': 1.40, 'u15': 3.00,
                                            'o25': 2.30, 'u25': 1.62,
                                            'o35': 4.33, 'u35': 1.22,
                                            'gg': 2.05, 'ng': 1.70,
                                            'dc_1x': 1.18, 'dc_x2': 1.91, 'dc_12': 1.29,
                                            'dnb_h': 1.33, 'dnb_a': 3.25,
                                            'cs_1_0': 6.00, 'cs_1_1': 7.00, 'cs_2_0': 8.00,
                                            'cs_0_0': 7.50, 'cs_2_1': 9.50, 'cs_0_1': 11.00,
                                            'cs_3_0': 15.00, 'cs_1_2': 17.00, 'cs_3_1': 19.00,
                                            'cs_2_2': 19.00, 'cs_0_2': 23.00, 'cs_4_0': 34.00,
                                            'cs_3_2': 34.00, 'cs_4_1': 41.00, 'cs_1_3': 41.00},

    # ── June 13, 2026 ─────────────────────────────────────────────────────────
    ('United States', 'Paraguay'):         {'h': 1.95, 'd': 3.40, 'a': 4.10,
                                            'o15': 1.36, 'u15': 3.20,
                                            'o25': 2.30, 'u25': 1.62,
                                            'o35': 4.00, 'u35': 1.25,
                                            'gg': 2.00, 'ng': 1.75,
                                            'dc_1x': 1.22, 'dc_x2': 1.80, 'dc_12': 1.22,
                                            'dnb_h': 1.44, 'dnb_a': 2.88,
                                            'cs_1_0': 6.50, 'cs_1_1': 6.50, 'cs_2_0': 9.00,
                                            'cs_0_0': 7.50, 'cs_2_1': 9.50, 'cs_0_1': 10.00,
                                            'cs_1_2': 15.00, 'cs_3_0': 19.00, 'cs_3_1': 19.00,
                                            'cs_2_2': 19.00, 'cs_0_2': 21.00, 'cs_3_2': 34.00,
                                            'cs_4_0': 41.00, 'cs_1_3': 41.00, 'cs_4_1': 41.00},
    ('Qatar', 'Switzerland'):              {'h': 16.00, 'd': 7.00, 'a': 2.33,
                                            'o15': 1.22, 'u15': 4.33,
                                            'o25': 1.67, 'u25': 2.20,
                                            'o35': 2.75, 'u35': 1.44,
                                            'gg': 2.50, 'ng': 1.50,
                                            'dc_1x': 4.00, 'dc_x2': 1.05, 'dc_12': 1.11,
                                            'dnb_h': 11.00, 'dnb_a': 1.05,
                                            'cs_0_2': 5.50, 'cs_0_1': 7.00, 'cs_0_3': 7.00,
                                            'cs_1_2': 11.00, 'cs_0_4': 11.00, 'cs_1_1': 13.00,
                                            'cs_1_3': 13.00, 'cs_0_0': 13.00, 'cs_1_4': 21.00,
                                            'cs_0_5': 21.00, 'cs_1_0': 23.00, 'cs_2_2': 34.00,
                                            'cs_1_5': 34.00, 'cs_2_1': 34.00, 'cs_2_3': 41.00},

    # ── June 14, 2026 ─────────────────────────────────────────────────────────
    ('Brazil', 'Morocco'):                 {'h': 1.65, 'd': 3.75, 'a': 5.50,
                                            'o15': 1.33, 'u15': 3.40,
                                            'o25': 2.00, 'u25': 1.80,
                                            'o35': 3.75, 'u35': 1.29,
                                            'gg': 2.05, 'ng': 1.70,
                                            'dc_1x': 1.14, 'dc_x2': 2.25, 'dc_12': 1.25,
                                            'dnb_h': 1.22, 'dnb_a': 4.00,
                                            'cs_1_0': 6.50, 'cs_2_0': 7.50, 'cs_1_1': 7.50,
                                            'cs_2_1': 9.00, 'cs_0_0': 9.00, 'cs_3_0': 12.00,
                                            'cs_0_1': 13.00, 'cs_3_1': 15.00, 'cs_1_2': 19.00,
                                            'cs_2_2': 21.00, 'cs_4_0': 26.00, 'cs_0_2': 29.00,
                                            'cs_4_1': 29.00, 'cs_3_2': 34.00, 'cs_1_3': 51.00},
    ('Haiti', 'Scotland'):                 {'h': 6.50, 'd': 4.33, 'a': 1.60,
                                            'o15': 1.29, 'u15': 3.75,
                                            'o25': 1.91, 'u25': 1.91,
                                            'o35': 3.40, 'u35': 1.33,
                                            'gg': 1.95, 'ng': 1.80,
                                            'dc_1x': 2.30, 'dc_x2': 1.17, 'dc_12': 1.20,
                                            'dnb_h': 4.00, 'dnb_a': 1.22,
                                            'cs_0_1': 7.00, 'cs_0_2': 7.50, 'cs_1_1': 8.50,
                                            'cs_1_2': 9.00, 'cs_0_3': 12.00, 'cs_0_0': 11.00,
                                            'cs_1_3': 15.00, 'cs_1_0': 15.00, 'cs_2_1': 19.00,
                                            'cs_2_2': 21.00, 'cs_0_4': 23.00, 'cs_1_4': 26.00,
                                            'cs_2_3': 29.00, 'cs_2_0': 29.00, 'cs_3_1': 41.00},
    ('Australia', 'Turkey'):               {'h': 5.60, 'd': 3.80, 'a': 1.75,
                                            'o15': 1.33, 'u15': 3.40,
                                            'o25': 2.00, 'u25': 1.80,
                                            'o35': 3.50, 'u35': 1.30,
                                            'gg': 1.95, 'ng': 1.80,
                                            'dc_1x': 2.10, 'dc_x2': 1.17, 'dc_12': 1.25,
                                            'dnb_h': 3.75, 'dnb_a': 1.25,
                                            'cs_0_1': 7.00, 'cs_1_1': 7.00, 'cs_0_2': 8.00,
                                            'cs_1_2': 9.00, 'cs_0_0': 9.50, 'cs_1_0': 13.00,
                                            'cs_0_3': 13.00, 'cs_1_3': 15.00, 'cs_2_1': 17.00,
                                            'cs_2_2': 19.00, 'cs_2_0': 29.00, 'cs_0_4': 29.00,
                                            'cs_2_3': 29.00, 'cs_1_4': 34.00, 'cs_3_1': 41.00},
    ('Germany', 'Curaçao'):               {'h': 1.04, 'd': 19.00, 'a': 41.00,
                                            'o15': 1.05, 'u15': 11.00,
                                            'o25': 1.20, 'u25': 4.50,
                                            'o35': 1.50, 'u35': 2.63,
                                            'gg': 2.50, 'ng': 1.50,
                                            'dc_1x': 1.01, 'dc_x2': 13.00, 'dc_12': 1.03,
                                            'dnb_h': 1.01, 'dnb_a': 26.00,
                                            'cs_3_0': 7.00, 'cs_4_0': 7.00, 'cs_2_0': 9.00,
                                            'cs_5_0': 9.00, 'cs_6_0': 13.00, 'cs_3_1': 15.00,
                                            'cs_4_1': 15.00, 'cs_1_0': 17.00, 'cs_2_1': 19.00,
                                            'cs_5_1': 19.00, 'cs_7_0': 21.00, 'cs_6_1': 26.00,
                                            'cs_1_1': 34.00, 'cs_8_0': 41.00, 'cs_7_1': 41.00},
    ('Netherlands', 'Japan'):              {'h': 2.05, 'd': 3.50, 'a': 3.60,
                                            'o15': 1.29, 'u15': 3.75,
                                            'o25': 1.91, 'u25': 1.91,
                                            'o35': 3.40, 'u35': 1.33,
                                            'gg': 1.75, 'ng': 2.00,
                                            'dc_1x': 1.29, 'dc_x2': 1.80, 'dc_12': 1.29,
                                            'dnb_h': 1.44, 'dnb_a': 2.63,
                                            'cs_1_1': 7.00, 'cs_1_0': 8.00, 'cs_2_1': 9.00,
                                            'cs_2_0': 10.00, 'cs_0_1': 12.00, 'cs_0_0': 11.00,
                                            'cs_1_2': 13.00, 'cs_2_2': 15.00, 'cs_3_1': 17.00,
                                            'cs_3_0': 19.00, 'cs_0_2': 21.00, 'cs_3_2': 29.00,
                                            'cs_1_3': 34.00, 'cs_4_1': 34.00, 'cs_2_3': 41.00},

    # ── June 15, 2026 ─────────────────────────────────────────────────────────
    ('Ivory Coast', 'Ecuador'):            {'h': 3.60, 'd': 2.90, 'a': 2.50,
                                            'o15': 1.44, 'u15': 2.75,
                                            'o25': 3.50, 'u25': 1.25,
                                            'o35': 5.00, 'u35': 1.18,
                                            'gg': 2.50, 'ng': 1.45,
                                            'dc_1x': 1.70, 'dc_x2': 1.50, 'dc_12': 1.12,
                                            'dnb_h': 2.25, 'dnb_a': 1.57,
                                            'cs_0_1': 6.00, 'cs_1_1': 6.00, 'cs_0_0': 5.50,
                                            'cs_1_0': 8.00, 'cs_0_2': 10.00, 'cs_1_2': 11.00,
                                            'cs_2_1': 15.00, 'cs_2_0': 17.00, 'cs_0_3': 26.00,
                                            'cs_2_2': 21.00, 'cs_1_3': 26.00, 'cs_3_1': 41.00,
                                            'cs_3_0': 41.00, 'cs_2_3': 41.00, 'cs_0_4': 51.00},
    ('Sweden', 'Tunisia'):                 {'h': 1.95, 'd': 3.50, 'a': 4.50,
                                            'o15': 1.40, 'u15': 2.00,
                                            'o25': 2.20, 'u25': 1.67,
                                            'o35': 4.00, 'u35': 1.25,
                                            'gg': 2.00, 'ng': 1.75,
                                            'dc_1x': 1.20, 'dc_x2': 1.91, 'dc_12': 1.30,
                                            'dnb_h': 1.36, 'dnb_a': 2.25,
                                            'cs_1_0': 6.50, 'cs_1_1': 6.50, 'cs_2_0': 7.50,
                                            'cs_2_1': 8.50, 'cs_0_0': 8.00, 'cs_0_1': 10.00,
                                            'cs_1_2': 16.00, 'cs_3_0': 16.00, 'cs_3_1': 18.00,
                                            'cs_2_2': 18.00, 'cs_0_2': 22.00, 'cs_3_2': 33.00,
                                            'cs_4_0': 33.00, 'cs_4_1': 40.00, 'cs_1_3': 40.00},
    ('Spain', 'Cape Verde'):               {'h': 1.09, 'd': 12.00, 'a': 29.00,
                                            'o15': 1.12, 'u15': 7.00,
                                            'o25': 1.57, 'u25': 2.30,
                                            'o35': 2.00, 'u35': 1.80,
                                            'gg': 2.62, 'ng': 1.44,
                                            'dc_1x': 1.01, 'dc_x2': 3.50, 'dc_12': 1.01,
                                            'dnb_h': 1.02, 'dnb_a': 18.00,
                                            'cs_3_0': 2.75, 'cs_2_0': 7.00, 'cs_4_0': 8.00,
                                            'cs_1_0': 11.00, 'cs_5_0': 12.00, 'cs_3_1': 13.00,
                                            'cs_2_1': 13.00, 'cs_4_1': 15.00, 'cs_1_1': 21.00,
                                            'cs_6_0': 21.00, 'cs_5_1': 23.00, 'cs_0_0': 23.00,
                                            'cs_6_1': 41.00, 'cs_3_2': 41.00, 'cs_7_0': 41.00},
    ('Belgium', 'Egypt'):                  {'h': 1.61, 'd': 3.80, 'a': 4.75,
                                            'o15': 1.30, 'u15': 3.50,
                                            'o25': 1.72, 'u25': 2.10,
                                            'o35': 3.40, 'u35': 1.33,
                                            'gg': 2.20, 'ng': 1.75,
                                            'dc_1x': 1.14, 'dc_x2': 2.20, 'dc_12': 1.25,
                                            'dnb_h': 1.22, 'dnb_a': 4.00,
                                            'cs_1_0': 7.00, 'cs_2_0': 6.50, 'cs_1_1': 6.50,
                                            'cs_2_1': 8.00, 'cs_0_0': 9.00, 'cs_3_0': 12.00,
                                            'cs_0_1': 14.00, 'cs_3_1': 14.00, 'cs_1_2': 18.00,
                                            'cs_2_2': 18.00, 'cs_4_0': 25.00, 'cs_0_2': 28.00,
                                            'cs_4_1': 28.00, 'cs_3_2': 28.00, 'cs_1_3': 40.00},

    # ── June 16, 2026 ─────────────────────────────────────────────────────────
    ('Saudi Arabia', 'Uruguay'):           {'h': 6.25, 'd': 4.33, 'a': 1.50,
                                            'o15': 1.30, 'u15': 3.50,
                                            'o25': 2.00, 'u25': 1.80,
                                            'o35': 3.50, 'u35': 1.30,
                                            'gg': 2.20, 'ng': 1.62,
                                            'dc_1x': 2.63, 'dc_x2': 1.10, 'dc_12': 1.20,
                                            'dnb_h': 5.50, 'dnb_a': 1.14,
                                            'cs_0_1': 6.00, 'cs_0_2': 6.50, 'cs_1_1': 9.00,
                                            'cs_1_2': 9.50, 'cs_0_3': 9.50, 'cs_0_0': 10.00,
                                            'cs_1_3': 15.00, 'cs_1_0': 17.00, 'cs_0_4': 19.00,
                                            'cs_2_1': 23.00, 'cs_2_2': 23.00, 'cs_1_4': 26.00,
                                            'cs_2_3': 34.00, 'cs_2_0': 41.00, 'cs_0_5': 41.00},
    ('Iran', 'New Zealand'):               {'h': 1.85, 'd': 3.50, 'a': 4.20,
                                            'o15': 1.44, 'u15': 3.10,
                                            'o25': 2.30, 'u25': 1.62,
                                            'o35': 3.33, 'u35': 1.22,
                                            'gg': 2.05, 'ng': 1.70,
                                            'dc_1x': 1.18, 'dc_x2': 1.91, 'dc_12': 1.30,
                                            'dnb_h': 1.30, 'dnb_a': 2.40,
                                            'cs_1_0': 6.00, 'cs_1_1': 5.50, 'cs_2_0': 7.00,
                                            'cs_0_0': 6.50, 'cs_2_1': 8.50, 'cs_0_1': 10.00,
                                            'cs_3_0': 14.00, 'cs_1_2': 16.00, 'cs_3_1': 18.00,
                                            'cs_2_2': 20.00, 'cs_0_2': 22.00, 'cs_4_0': 33.00,
                                            'cs_3_2': 40.00, 'cs_4_1': 40.00, 'cs_1_3': 40.00},
    ('France', 'Senegal'):                 {'h': 1.44, 'd': 4.50, 'a': 6.50,
                                            'o15': 1.30, 'u15': 3.50,
                                            'o25': 1.91, 'u25': 1.91,
                                            'o35': 3.40, 'u35': 1.33,
                                            'gg': 2.10, 'ng': 1.67,
                                            'dc_1x': 1.10, 'dc_x2': 2.63, 'dc_12': 1.18,
                                            'dnb_h': 1.14, 'dnb_a': 4.50,
                                            'cs_1_0': 6.50, 'cs_2_0': 6.50, 'cs_1_1': 8.50,
                                            'cs_2_1': 9.00, 'cs_3_0': 10.00, 'cs_0_0': 10.00,
                                            'cs_3_1': 13.00, 'cs_0_1': 15.00, 'cs_1_2': 21.00,
                                            'cs_4_0': 19.00, 'cs_2_2': 23.00, 'cs_4_1': 26.00,
                                            'cs_3_2': 34.00, 'cs_0_2': 41.00, 'cs_5_0': 41.00},

    # ── June 17, 2026 ─────────────────────────────────────────────────────────
    ('Iraq', 'Norway'):                    {'h': 12.50, 'd': 7.00, 'a': 1.21,
                                            'o15': 1.17, 'u15': 4.00,
                                            'o25': 1.57, 'u25': 2.38,
                                            'o35': 2.75, 'u35': 1.44,
                                            'gg': 2.25, 'ng': 1.57,
                                            'dc_1x': 10.00, 'dc_x2': 1.04, 'dc_12': 1.10,
                                            'dnb_h': 9.00, 'dnb_a': 1.06,
                                            'cs_0_2': 6.00, 'cs_0_3': 7.00, 'cs_0_1': 8.00,
                                            'cs_1_2': 11.00, 'cs_0_4': 11.00, 'cs_1_3': 12.00,
                                            'cs_1_1': 13.00, 'cs_0_0': 15.00, 'cs_1_4': 19.00,
                                            'cs_0_5': 21.00, 'cs_1_0': 26.00, 'cs_2_2': 34.00,
                                            'cs_1_5': 34.00, 'cs_2_3': 34.00, 'cs_2_1': 34.00},
    ('Argentina', 'Algeria'):              {'h': 1.40, 'd': 4.50, 'a': 7.75,
                                            'o15': 1.30, 'u15': 3.50,
                                            'o25': 1.91, 'u25': 1.91,
                                            'o35': 3.40, 'u35': 1.33,
                                            'gg': 2.20, 'ng': 1.62,
                                            'dc_1x': 1.07, 'dc_x2': 3.00, 'dc_12': 1.18,
                                            'dnb_h': 1.11, 'dnb_a': 5.50,
                                            'cs_2_0': 5.50, 'cs_1_0': 6.00, 'cs_3_0': 9.00,
                                            'cs_2_1': 9.00, 'cs_1_1': 8.50, 'cs_0_0': 11.00,
                                            'cs_3_1': 13.00, 'cs_4_0': 17.00, 'cs_0_1': 19.00,
                                            'cs_4_1': 26.00, 'cs_2_2': 26.00, 'cs_1_2': 26.00,
                                            'cs_3_2': 34.00, 'cs_5_0': 34.00, 'cs_0_2': 41.00},
    ('Austria', 'Jordan'):                 {'h': 1.33, 'd': 5.50, 'a': 8.00,
                                            'o15': 1.20, 'u15': 3.50,
                                            'o25': 1.67, 'u25': 2.20,
                                            'o35': 3.00, 'u35': 1.40,
                                            'gg': 2.00, 'ng': 1.75,
                                            'dc_1x': 1.10, 'dc_x2': 2.25, 'dc_12': 1.14,
                                            'dnb_h': 1.11, 'dnb_a': 5.50,
                                            'cs_2_0': 6.00, 'cs_1_0': 6.50, 'cs_2_1': 8.00,
                                            'cs_3_0': 8.00, 'cs_1_1': 10.00, 'cs_3_1': 11.00,
                                            'cs_4_0': 14.00, 'cs_0_0': 14.00, 'cs_4_1': 20.00,
                                            'cs_0_1': 20.00, 'cs_2_2': 22.00, 'cs_1_2': 22.00,
                                            'cs_3_2': 28.00, 'cs_5_0': 28.00, 'cs_5_1': 40.00},
    ('Portugal', 'DR Congo'):              {'h': 1.27, 'd': 6.50, 'a': 5.75,
                                            'o15': 1.30, 'u15': 3.50,
                                            'o25': 2.00, 'u25': 1.80,
                                            'o35': 3.50, 'u35': 1.30,
                                            'gg': 2.20, 'ng': 1.62,
                                            'dc_1x': 1.08, 'dc_x2': 2.80, 'dc_12': 1.06,
                                            'dnb_h': 1.08, 'dnb_a': 7.00,
                                            'cs_2_0': 5.00, 'cs_1_0': 6.00, 'cs_3_0': 7.00,
                                            'cs_2_1': 9.00, 'cs_1_1': 10.00, 'cs_3_1': 12.00,
                                            'cs_4_0': 12.00, 'cs_0_0': 12.00, 'cs_4_1': 20.00,
                                            'cs_0_1': 20.00, 'cs_2_2': 28.00, 'cs_5_0': 25.00,
                                            'cs_1_2': 28.00, 'cs_3_2': 33.00, 'cs_5_1': 40.00},
    ('England', 'Croatia'):                {'h': 1.72, 'd': 3.70, 'a': 4.75,
                                            'o15': 1.36, 'u15': 3.20,
                                            'o25': 2.10, 'u25': 1.73,
                                            'o35': 4.00, 'u35': 1.25,
                                            'gg': 2.05, 'ng': 1.70,
                                            'dc_1x': 1.18, 'dc_x2': 2.05, 'dc_12': 1.25,
                                            'dnb_h': 1.29, 'dnb_a': 3.50,
                                            'cs_1_0': 6.50, 'cs_1_1': 7.50, 'cs_2_0': 7.50,
                                            'cs_2_1': 9.00, 'cs_0_0': 9.00, 'cs_0_1': 12.00,
                                            'cs_3_0': 13.00, 'cs_3_1': 17.00, 'cs_1_2': 17.00,
                                            'cs_2_2': 21.00, 'cs_0_2': 26.00, 'cs_4_0': 29.00,
                                            'cs_3_2': 34.00, 'cs_4_1': 34.00, 'cs_1_3': 41.00},

    # ── June 18, 2026 ─────────────────────────────────────────────────────────
    ('Ghana', 'Panama'):                   {'h': 2.05, 'd': 3.30, 'a': 3.60,
                                            'o15': 1.40, 'u15': 2.00,
                                            'o25': 2.20, 'u25': 1.67,
                                            'o35': 3.33, 'u35': 1.22,
                                            'gg': 1.95, 'ng': 1.80,
                                            'dc_1x': 1.30, 'dc_x2': 1.67, 'dc_12': 1.14,
                                            'dnb_h': 1.53, 'dnb_a': 1.38,
                                            'cs_1_1': 6.00, 'cs_1_0': 6.00, 'cs_0_1': 8.00,
                                            'cs_0_0': 7.50, 'cs_2_1': 9.00, 'cs_2_0': 9.00,
                                            'cs_1_2': 12.00, 'cs_0_2': 16.00, 'cs_2_2': 18.00,
                                            'cs_3_1': 20.00, 'cs_3_0': 20.00, 'cs_1_3': 33.00,
                                            'cs_3_2': 33.00, 'cs_0_3': 40.00, 'cs_2_3': 40.00},
    ('Uzbekistan', 'Colombia'):            {'h': 7.75, 'd': 4.60, 'a': 1.40,
                                            'o15': 1.30, 'u15': 2.50,
                                            'o25': 1.91, 'u25': 1.91,
                                            'o35': 4.00, 'u35': 1.33,
                                            'gg': 2.20, 'ng': 1.62,
                                            'dc_1x': 3.00, 'dc_x2': 1.08, 'dc_12': 1.18,
                                            'dnb_h': 7.00, 'dnb_a': 1.10,
                                            'cs_0_2': 5.00, 'cs_0_1': 5.00, 'cs_0_3': 8.00,
                                            'cs_1_2': 8.50, 'cs_1_1': 8.00, 'cs_0_0': 9.00,
                                            'cs_1_3': 12.00, 'cs_0_4': 16.00, 'cs_1_0': 18.00,
                                            'cs_1_4': 25.00, 'cs_2_2': 25.00, 'cs_2_1': 28.00,
                                            'cs_2_3': 33.00, 'cs_0_5': 33.00, 'cs_2_0': 40.00},
}

def market_implied(home, away):
    """Returneaza probabilitatile implicite din piata (fara marja bookmaker)."""
    o = MARKET_ODDS.get((home, away))
    if not o: return None
    raw = {'h': 1/o['h'], 'd': 1/o['d'], 'a': 1/o['a']}
    total = sum(raw.values())
    return {k: v/total for k, v in raw.items()}

def value_score(model_prob_pct, bookmaker_odds):
    """
    Calculeaza VALUE = (prob_model * cota) - 1
    > 0 = pariu cu valoare pozitiva (EV+)
    < 0 = pariu fara valoare (dai mai mult decat primesti)
    """
    if bookmaker_odds is None or bookmaker_odds <= 1.0:
        return None
    return round((model_prob_pct / 100) * bookmaker_odds - 1.0, 3)

def get_odds_for_bet(home, away, cat, label, pred):
    """Returns bookmaker odds for a specific bet if available in MARKET_ODDS."""
    o = MARKET_ODDS.get((home, away), {})
    if not o: return None
    if cat == 'result':
        if '1 (' in pred or pred.startswith('1 '): return o.get('h')
        if 'X' in pred or 'egal' in pred.lower(): return o.get('d')
        if '2 (' in pred or pred.startswith('2 '): return o.get('a')
    if cat == 'goals_ou':
        if 'Over 1.5'  in pred: return o.get('o15')
        if 'Under 1.5' in pred: return o.get('u15')
        if 'Over 2.5'  in pred: return o.get('o25')
        if 'Under 2.5' in pred: return o.get('u25')
        if 'Over 3.5'  in pred: return o.get('o35')
        if 'Under 3.5' in pred: return o.get('u35')
    if cat == 'gg':
        if pred == 'Da': return o.get('gg')
        if pred == 'Nu': return o.get('ng')
    if cat == 'dc':
        if pred == '1X': return o.get('dc_1x')
        if pred == 'X2': return o.get('dc_x2')
        if pred == '12': return o.get('dc_12')
    if cat == 'dnb':
        return o.get('dnb_h') if home[:14] in label else o.get('dnb_a')
    if cat == 'correct':
        key = 'cs_' + pred.replace('-', '_')
        return o.get(key)
    return None

def absence_factor(team):
    ab = ABSENCES.get(team, {})
    return ab.get('attack', 1.0), ab.get('defense', 1.0)

FIXTURES_ELO = {}
for r in ELO_RAW.get('wc_fixtures', []):
    h, a = canon(r['home']), canon(r['away'])
    FIXTURES_ELO[(h, a)] = r

TOUR_WEIGHT = {
    'World Cup': 2.0, 'World Cup qualifier': 1.5,
    'World Cup and Asian Cup qualifier': 1.5,
    'European Championship': 1.7, 'European Championship qualifier': 1.3,
    'Copa America': 1.7, 'African Nations Cup': 1.4,
    'African Nations Cup qualifier': 1.2, 'Asian Cup': 1.4,
    'Asian Cup qualifier': 1.2, 'CONCACAF Championship': 1.4,
    'CONCACAF Nations League A': 1.3, 'CONCACAF Nations League': 1.2,
    'CONCACAF Nations League B': 1.1, 'European Nations League A': 1.3,
    'European Nations League B': 1.2, 'European Nations League C': 1.1,
    'European Nations League': 1.2,
    'Friendly': 0.5, 'Friendly tournament': 0.5,
}
def match_weight(t): return TOUR_WEIGHT.get(t, 1.0)

# ─── Venue WC2026 ────────────────────────────────────────────────────────────
WC_HOSTS      = {'United States', 'Canada', 'Mexico'}
CONCACAF_NEAR = {'Panama', 'Jamaica', 'Honduras', 'El Salvador',
                 'Costa Rica', 'Cuba', 'Curaçao', 'Haiti'}

def venue_score(team):
    return 1.0 if team in WC_HOSTS else 0.0

# ─── Form (versiune rolling temporala) ───────────────────────────────────────
def compute_form(buf, n=10):
    m = buf[-n:]
    if not m:
        return {'wpg': 1.2, 'wcpg': 1.2, 'win_rate': 0.4,
                'avg_opp_elo': 1700, 'lambda_recent': 1.2,
                'official_ratio': 0.5, 'n': 0}
    tw = sum(x['w'] for x in m) or len(m)
    wpg  = sum(x['gf'] * x['w'] for x in m) / tw
    wcpg = sum(x['ga'] * x['w'] for x in m) / tw
    wins = sum(x['w'] for x in m if x['gf'] > x['ga']) / tw
    avg_opp = sum(x['opp_elo'] for x in m) / len(m)
    elo_adj = sum(x['gf'] * x['w'] * (1 + (x['opp_elo'] - 1700) / 3000)
                  for x in m) / tw
    off_r = sum(1 for x in m if x['w'] > 0.75) / len(m)
    return {'wpg': wpg, 'wcpg': wcpg, 'win_rate': wins,
            'avg_opp_elo': avg_opp, 'lambda_recent': max(0.3, elo_adj),
            'official_ratio': off_r, 'n': len(m)}

def compute_h2h(h2h_buf, home, away, n=5):
    key = (min(home, away), max(home, away))
    ms  = h2h_buf.get(key, [])[-n:]
    if not ms:
        return {'win_rate': 0.40, 'avg_gd': 0.0, 'n_norm': 0.0}
    wins = sum(1 for m in ms
               if (m['home'] == home and m['gd'] > 0)
               or (m['home'] != home and m['gd'] < 0))
    gds  = [m['gd'] if m['home'] == home else -m['gd'] for m in ms]
    return {'win_rate': wins / len(ms),
            'avg_gd':   float(np.mean(gds)),
            'n_norm':   min(len(ms), 5) / 5.0}

# ─── Features (FARA LEAKAGE) ─────────────────────────────────────────────────
FEATURE_NAMES = [
    'elo_home', 'elo_away', 'elo_diff',
    'fh_wpg', 'fh_wcpg', 'fh_win_rate', 'fh_lambda', 'fh_off_ratio', 'fh_opp_diff',
    'fa_wpg', 'fa_wcpg', 'fa_win_rate', 'fa_lambda', 'fa_off_ratio', 'fa_opp_diff',
    'h2h_win_rate', 'h2h_avg_gd', 'h2h_n_norm',
    'venue_adv', 'tour_weight',
    'elo_rating_trend_h', 'elo_rating_trend_a',
]

def extract_features(home, away, fh, fa, h2h, helo, aelo, tournament,
                     elo_trend_h=0.0, elo_trend_a=0.0):
    return [
        helo, aelo, helo - aelo,
        fh['wpg'], fh['wcpg'], fh['win_rate'], fh['lambda_recent'],
        fh['official_ratio'], fh['avg_opp_elo'] - aelo,
        fa['wpg'], fa['wcpg'], fa['win_rate'], fa['lambda_recent'],
        fa['official_ratio'], fa['avg_opp_elo'] - helo,
        h2h['win_rate'], h2h['avg_gd'], h2h['n_norm'],
        venue_score(home), match_weight(tournament),
        elo_trend_h, elo_trend_a,
    ]

# ─── Dataset builder ─────────────────────────────────────────────────────────
def build_dataset(matches_sorted):
    form_buf = defaultdict(list)
    h2h_buf  = defaultdict(list)
    elo_hist = defaultdict(list)
    X, y_res, y_o25, y_gh, y_ga, meta = [], [], [], [], [], []

    for m in matches_sorted:
        h, a = canon(m['home']), canon(m['away'])
        try:
            hs, as_ = int(m['home_score']), int(m['away_score'])
        except (TypeError, ValueError, KeyError):
            continue
        tournament = m.get('tournament', 'Friendly')
        w    = match_weight(tournament)
        helo = float(m.get('home_elo') or ELO.get(h, {}).get('rating', 1700))
        aelo = float(m.get('away_elo') or ELO.get(a, {}).get('rating', 1700))

        def elo_trend(team, current_elo):
            hist = elo_hist[team][-5:]
            return (current_elo - np.mean(hist)) if hist else 0.0

        fh   = compute_form(form_buf[h])
        fa   = compute_form(form_buf[a])
        h2h  = compute_h2h(h2h_buf, h, a)
        feat = extract_features(h, a, fh, fa, h2h, helo, aelo, tournament,
                                 elo_trend(h, helo), elo_trend(a, aelo))

        X.append(feat)
        y_res.append(2 if hs > as_ else (1 if hs == as_ else 0))
        y_o25.append(1 if (hs + as_) > 2.5 else 0)
        y_gh.append(hs); y_ga.append(as_)
        meta.append({'home': h, 'away': a, 'date': m['date'],
                     'hs': hs, 'as_': as_, 'tournament': tournament,
                     'helo': helo, 'aelo': aelo})

        key = (min(h, a), max(h, a))
        h2h_buf[key].append({'home': h, 'gd': hs - as_})
        if len(h2h_buf[key]) > 15: h2h_buf[key] = h2h_buf[key][-15:]

        for team, gf, ga, opp_e, my_e in [(h, hs, as_, aelo, helo),
                                            (a, as_, hs, helo, aelo)]:
            form_buf[team].append({'gf': gf, 'ga': ga, 'opp_elo': opp_e, 'w': w})
            if len(form_buf[team]) > 20: form_buf[team] = form_buf[team][-20:]
            elo_hist[team].append(my_e)
            if len(elo_hist[team]) > 10: elo_hist[team] = elo_hist[team][-10:]

    X_arr = np.array(X, dtype=float)
    nan_mask = np.isnan(X_arr).any(axis=1)
    X_arr[nan_mask] = np.nanmean(X_arr, axis=0)

    return (X_arr, np.array(y_res), np.array(y_o25),
            np.array(y_gh), np.array(y_ga), meta, form_buf, h2h_buf, elo_hist)

# ─── Poisson + Dixon-Coles ───────────────────────────────────────────────────
def ppf(k, lam):
    if lam <= 0: return 1.0 if k == 0 else 0.0
    return math.exp(-lam) * (lam ** k) / math.factorial(k)

# Adaugă aceste două linii:
def pcdf(k, lam): return sum(ppf(i, lam) for i in range(k + 1))
def p_over(thresh, lam): return 1.0 - pcdf(int(thresh), lam)

def dc_adjust(i, j, lh, la, rho):
    # Dixon-Coles 1997 full correction: 0-0, 1-0, 0-1, 1-1
    if   i == 0 and j == 0: return max(0.0, 1 - lh * la * rho)
    elif i == 1 and j == 0: return max(0.0, 1 + la * rho)
    elif i == 0 and j == 1: return max(0.0, 1 + lh * rho)
    elif i == 1 and j == 1: return max(0.0, 1 - rho)
    return 1.0

def get_match_matrix(lh, la, rho=0.0, max_g=8):
    """ Construieste matricea Joint Probability 8x8 pentru scorurile meciului """
    mat = np.zeros((max_g + 1, max_g + 1))
    for i in range(max_g + 1):
        for j in range(max_g + 1):
            mat[i, j] = ppf(i, lh) * ppf(j, la) * dc_adjust(i, j, lh, la, rho)
    
    # Normalizare minora pt a preveni devieri matematice
    return mat / mat.sum()

def poisson_xg(home, away, fh, fa, helo, aelo, w_elo, w_qual, w_form, avg_goals,
               cond_lh=1.0, cond_la=1.0, abs_h_att=1.0, abs_a_att=1.0):
    # Branch 1: Elo
    elo_diff = helo - aelo + 100
    ef = 10 ** (elo_diff / 800)
    lh_elo = avg_goals * ef / (1 + ef)
    la_elo = avg_goals / (1 + ef)

    # Branch 2: Statistici calificari (AiScore calibrat per-confederatie)
    # CORECTIE: anterior folosea fh['wpg'] (date de forma!), acum foloseste QUAL_GOALS
    hgpg = QUAL_GOALS.get(home, 1.3)
    agpg = QUAL_GOALS.get(away, 1.3)
    qt   = hgpg + agpg or 2.6
    lh_q = avg_goals * hgpg / qt
    la_q = avg_goals * agpg / qt

    # Branch 3: Forma recenta (ultimele 10 meciuri, ajustata pt calitatea adversarului)
    rft   = fh['lambda_recent'] + fa['lambda_recent'] or 2.4
    lh_rf = avg_goals * fh['lambda_recent'] / rft
    la_rf = avg_goals * fa['lambda_recent'] / rft
    lh_rf *= max(0.7, min(1.3, 1.0 + (fh['avg_opp_elo'] - aelo) / 3000))
    la_rf *= max(0.7, min(1.3, 1.0 + (fa['avg_opp_elo'] - helo) / 3000))

    # Avantaj teren: DOAR Mexico, USA, Canada joaca acasa la WC2026
    venue_h = 1.08 if home in WC_HOSTS else 1.0
    venue_a = 1.08 if away in WC_HOSTS else 1.0

    wf  = max(0.05, 1.0 - w_elo - w_qual)
    lh  = (w_elo * lh_elo + w_qual * lh_q + wf * lh_rf) * venue_h
    la  = (w_elo * la_elo + w_qual * la_q + wf * la_rf) * venue_a

    # Conditii stadion (altitudine, caldura) + absente
    lh *= cond_lh * abs_h_att
    la *= cond_la * abs_a_att
    return max(0.15, lh), max(0.15, la)

# ─── Optimizare ponderi & RHO ────────────────────────────────────────────────
def optimize_poisson(X_tr, y_gh_tr, y_ga_tr, meta_tr):
    def nll(params):
        w_elo, w_qual, avg_g, rho = params
        wf = 1.0 - w_elo - w_qual
        if w_elo < 0.05 or w_qual < 0.05 or wf < 0.05: return 1e9
        total = 0.0
        for i, rec in enumerate(meta_tr):
            helo, aelo = rec['helo'], rec['aelo']
            
            elo_diff = helo - aelo + 100
            ef   = 10 ** (elo_diff / 800)
            lh_e = avg_g * ef / (1 + ef)
            la_e = avg_g / (1 + ef)

            # CORECTIE: foloseste QUAL_GOALS (AiScore calibrat) nu forma
            home, away = rec['home'], rec['away']
            hgpg = QUAL_GOALS.get(home, 1.3)
            agpg = QUAL_GOALS.get(away, 1.3)
            qt   = hgpg + agpg or 2.6
            lh_q = avg_g * hgpg / qt
            la_q = avg_g * agpg / qt

            lr_h = float(X_tr[i, 6]); lr_a = float(X_tr[i, 12])
            rft  = lr_h + lr_a or 2.4
            lh_rf = avg_g * lr_h / rft
            la_rf = avg_g * lr_a / rft
            lh_rf *= max(0.7, min(1.3, 1.0 + float(X_tr[i, 8])  / 3000))
            la_rf *= max(0.7, min(1.3, 1.0 + float(X_tr[i, 14]) / 3000))

            lh = max(0.15, w_elo*lh_e + w_qual*lh_q + wf*lh_rf)
            la = max(0.15, w_elo*la_e + w_qual*la_q + wf*la_rf)

            obs_h = min(int(y_gh_tr[i]), 8)
            obs_a = min(int(y_ga_tr[i]), 8)
            
            p_base = ppf(obs_h, lh) * ppf(obs_a, la)
            p_dc = p_base * dc_adjust(obs_h, obs_a, lh, la, rho)

            tw = match_weight(rec['tournament'])
            total -= tw * math.log(max(p_dc, 1e-15))
        return total

    # Bounds: rho este in general negativ (ex: -0.1) pt a corecta 0-0
    res = minimize(nll, x0=[0.35, 0.35, 2.5, -0.05], method='L-BFGS-B',
                   bounds=[(0.08, 0.60), (0.08, 0.60), (1.8, 3.6), (-0.25, 0.1)])
    we, wq, ag, rho = float(res.x[0]), float(res.x[1]), float(res.x[2]), float(res.x[3])
    wf = max(0.05, 1.0 - we - wq)
    print(f"  Ponderi optime: Elo={we:.3f}  Qual={wq:.3f}  Form={wf:.3f}  avg_goals={ag:.3f} RHO={rho:.3f}")
    return we, wq, wf, ag, rho

# ─── Player & Independent Stats Data ─────────────────────────────────────────
def _load_players(path='statistici_calificari_cm2026_v6.xlsx'):
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb['AiScore_Jucatori']
        pl = {}
        for row in list(ws.rows)[1:]:
            vals = [c.value for c in row]
            if len(vals) < 6: continue
            _, team, player, stat, _, val_num = vals[:6]
            if not team or not player: continue
            ct = canon(team)
            if ct not in pl: pl[ct] = {}
            if stat not in pl[ct]: pl[ct][stat] = []
            pl[ct][stat].append((player, val_num or 0))
        wb.close()
        return pl
    except Exception:
        return {}

PLAYERS = _load_players()

def _scorers(team, lam, n=3):
    pg = sorted(PLAYERS.get(team, {}).get('goals', []), key=lambda x: -(x[1] or 0))[:6]
    if not pg: return []
    total = sum(v for _, v in pg) or 1
    return [(name, round((1 - math.exp(-lam * (g / total))) * 100, 1))
            for name, g in pg[:n]]

# ─── Generator Pariuri cu Matrice ───────────────────────────────────────────
def generate_bets_v3(home, away, wp_h, wp_d, wp_a, lh, la, mat, helo=1700, aelo=1700, ref_yc=1.0, ref_fouls=1.0):
    total = lh + la
    
    # Probabilitati Marginale (Array)
    p_h_goals = mat.sum(axis=1)
    p_a_goals = mat.sum(axis=0)
    
    # Sumar probabilitate total goluri meci
    total_g_mat = np.zeros(18)
    for i in range(9):
        for j in range(9):
            total_g_mat[i+j] += mat[i,j]

    def p_over_mat(thresh, prob_array):
        idx = int(math.floor(thresh)) + 1
        if idx >= len(prob_array): return 0.0
        return prob_array[idx:].sum() * 100

    def ou(thresh, prob, label):
        if prob >= 50: return (label, f"Over {thresh}",  round(prob, 1))
        else:          return (label, f"Under {thresh}", round(100 - prob, 1))

    bets = []

    # 1X2
    best = max(wp_h, wp_d, wp_a)
    pred = (f"1 ({home[:12]})" if wp_h == best else
            "X (egal)"         if wp_d == best else f"2 ({away[:12]})")
    bets.append(('1X2 Result', pred, round(best, 1), 'result'))
    bets += [
        ('Double Chance 1X', '1X', round(wp_h + wp_d, 1), 'dc'),
        ('Double Chance X2', 'X2', round(wp_a + wp_d, 1), 'dc'),
        ('Double Chance 12', '12', round(wp_h + wp_a, 1), 'dc'),
    ]

    # Draw No Bet
    dnb_h = wp_h / max(1, wp_h + wp_a) * 100
    dnb_a = wp_a / max(1, wp_h + wp_a) * 100
    bets += [
        (f'Draw No Bet {home[:14]}', 'Win', round(dnb_h, 1), 'dnb'),
        (f'Draw No Bet {away[:14]}', 'Win', round(dnb_a, 1), 'dnb'),
    ]

    # Over/Under extrase precis din matricea 8x8
    bets += [
        ou(0.5, p_over_mat(0.5, total_g_mat), 'Goals O/U 0.5') + ('goals_ou',),
        ou(1.5, p_over_mat(1.5, total_g_mat), 'Goals O/U 1.5') + ('goals_ou',),
        ou(2.5, p_over_mat(2.5, total_g_mat), 'Goals O/U 2.5') + ('goals_ou',),
        ou(3.5, p_over_mat(3.5, total_g_mat), 'Goals O/U 3.5') + ('goals_ou',),
    ]

    # GG Extras precis din Matrice (suma celulelor i>0 si j>0)
    p_gg = mat[1:, 1:].sum() * 100
    p_over25_exact = p_over_mat(2.5, total_g_mat)
    bets += [
        ('GG (Both Teams Score)', 'Da' if p_gg>=50 else 'Nu', round(max(p_gg, 100-p_gg), 1), 'gg'),
        ('GG + Over 2.5', 'Da', round((p_gg/100) * (p_over25_exact/100) * 100, 1), 'combo'),
    ]

    # Over/Under per Echipă
    bets += [
        ou(0.5, p_over_mat(0.5, p_h_goals), f'{home[:14]} O/U 0.5') + ('team_ou',),
        ou(1.5, p_over_mat(1.5, p_h_goals), f'{home[:14]} O/U 1.5') + ('team_ou',),
        ou(0.5, p_over_mat(0.5, p_a_goals), f'{away[:14]} O/U 0.5') + ('team_ou',),
        ou(1.5, p_over_mat(1.5, p_a_goals), f'{away[:14]} O/U 1.5') + ('team_ou',),
    ]

    # Win to Nil
    p_h_nil = mat[1:, 0].sum() * 100
    p_a_nil = mat[0, 1:].sum() * 100
    bets += [
        (f'{home[:18]} Win to Nil', 'Da', round(p_h_nil, 1), 'win_nil'),
        (f'{away[:18]} Win to Nil', 'Da', round(p_a_nil, 1), 'win_nil'),
    ]

    # Cele mai probabile Scoruri Exacte
    scores = []
    for i in range(6):
        for j in range(6):
            scores.append((i, j, mat[i, j]))
    scores.sort(key=lambda x: -x[2])
    for hi, ai, sp in scores[:8]:
        bets.append((f'Correct Score {hi}-{ai}', f'{hi}-{ai}', round(sp*100, 1), 'correct'))

    # Cornere: echipa dominanta (Elo mare) ataca central -> putine deflectii -> mai putine cornere
    shots_h = lh * 8.5
    shots_a = la * 8.5
    dom_h = max(0.70, 1.0 - max(0, helo - aelo) / 3000)
    dom_a = max(0.70, 1.0 - max(0, aelo - helo) / 3000)
    total_corn = shots_h * 0.45 * dom_h + shots_a * 0.45 * dom_a

    # Galbene si faulturi: meciuri echilibrate = usor mai multe, nu x1.5
    elo_diff_f = max(-0.20, min(0.20, (150 - abs(helo - aelo)) / 750))
    intensity  = 1.0 + elo_diff_f          # 0.80 – 1.20
    total_yc    = 3.2 * intensity * ref_yc
    total_fouls = 24.0 * intensity * ref_fouls

    bets += [
        ou(8.5,  p_over(8,  total_corn)*100,  'Corners O/U 8.5')   + ('corners',),
        ou(9.5,  p_over(9,  total_corn)*100,  'Corners O/U 9.5')   + ('corners',),
        ou(3.5,  p_over(3,  total_yc)*100,    'Yellow Cards O/U 3.5') + ('yellows',),
        ou(4.5,  p_over(4,  total_yc)*100,    'Yellow Cards O/U 4.5') + ('yellows',),
        ou(20.5, p_over(20, total_fouls)*100, 'Fouls O/U 20.5')    + ('fouls',),
        ou(24.5, p_over(24, total_fouls)*100, 'Fouls O/U 24.5')    + ('fouls',),
    ]

    for name, prob in _scorers(home, lh):
        bets.append((f'Scorer {name[:16]} ({home[:8]})', 'To Score', prob, 'scorers'))
    for name, prob in _scorers(away, la):
        bets.append((f'Scorer {name[:16]} ({away[:8]})', 'To Score', prob, 'scorers'))

    return bets

# ─── Selectie acumulator ─────────────────────────────────────────────────────
_CAT_PRIORITY = {
    'result': 1, 'dc': 1, 'dnb': 1, 'asian_hcp': 1,
    'goals_ou': 2, 'gg': 2, 'team_ou': 2, 'combo': 2,
    'corners': 3, 'yellows': 4, 'fouls': 5,
    'win_nil': 6, 'correct': 7,
    'scorers': 8, 'specials': 9,
}

def pick_accumulator(bets, n=3, min_prob=55.0, bet_ev=None):
    bet_ev = bet_ev or {}
    used_cats = set()
    selected = []

    def sort_key(b):
        ev = bet_ev.get(b[0])
        # Daca avem EV calculat: EV+ bets au prioritate, sortate dupa ev desc
        # Daca nu avem cote (MARKET_ODDS gol): fallback la sortare dupa prob
        return (ev if ev is not None else -999, b[2])

    candidates = sorted(
        [(label, pred, prob, cat) for label, pred, prob, cat in bets if prob >= min_prob],
        key=sort_key, reverse=True
    )
    for label, pred, prob, cat in candidates:
        grp = _CAT_PRIORITY.get(cat, 99)
        if grp not in used_cats:
            selected.append((label, pred, prob, cat))
            used_cats.add(grp)
        if len(selected) == n:
            break
    return selected

# ─── Antrenare ML ────────────────────────────────────────────────────────────
def train_ml(X_tr, y_res_tr, y_o25_tr):
    if not HAS_SKLEARN: return None, None
    n  = len(X_tr)
    cv = max(2, min(5, n // 80))

    if HAS_XGB:
        b1 = xgb.XGBClassifier(n_estimators=300, max_depth=4, learning_rate=0.04,
                                subsample=0.8, colsample_bytree=0.8,
                                eval_metric='mlogloss', random_state=42, verbosity=0)
        b2 = xgb.XGBClassifier(n_estimators=300, max_depth=4, learning_rate=0.04,
                                subsample=0.8, colsample_bytree=0.8,
                                eval_metric='logloss', random_state=42, verbosity=0)
    else:
        b1 = GradientBoostingClassifier(n_estimators=200, max_depth=3,
                                        learning_rate=0.05, random_state=42)
        b2 = GradientBoostingClassifier(n_estimators=200, max_depth=3,
                                        learning_rate=0.05, random_state=42)

    method = 'isotonic' if n >= 400 else 'sigmoid'
    clf1 = CalibratedClassifierCV(b1, method=method, cv=cv)
    clf2 = CalibratedClassifierCV(b2, method=method, cv=cv)
    clf1.fit(X_tr, y_res_tr)
    clf2.fit(X_tr, y_o25_tr)
    return clf1, clf2

# ─── Metrici ─────────────────────────────────────────────────────────────────
def metrics(y_true, proba, name):
    if not HAS_SKLEARN: return {}
    y_p  = np.argmax(proba, axis=1)
    acc  = accuracy_score(y_true, y_p)
    ll   = log_loss(y_true, proba, labels=[0, 1, 2])
    bs   = {c: brier_score_loss((y_true == c).astype(int), proba[:, c])
            for c in range(proba.shape[1])}
    print(f"  [{name}] acc={acc:.3f}  log-loss={ll:.4f}  "
          f"brier home/draw/away={bs.get(2,0):.4f}/{bs.get(1,0):.4f}/{bs.get(0,0):.4f}")
    return {'name': name, 'accuracy': acc, 'log_loss': ll,
            'brier_home': bs.get(2, 0), 'brier_draw': bs.get(1, 0),
            'brier_away': bs.get(0, 0)}

def get_feature_importance(clf):
    if not HAS_XGB and not HAS_SKLEARN: return []
    try:
        est = clf.calibrated_classifiers_[0].estimator
        imp = est.feature_importances_
        return sorted(zip(FEATURE_NAMES, imp.tolist()), key=lambda x: -x[1])
    except Exception:
        return []

# ─── Baza de date arbitri FIFA ───────────────────────────────────────────────
# yc_pg = galbenele medii per meci in meciuri internationale
# fouls_pg = faulturi medii per meci
# style: 'strict' / 'average' / 'lenient'
REFEREE_DB = {
    # Confirmat WC2026
    'Wilton Sampaio':      {'yc_pg': 4.80, 'fouls_pg': 30.0, 'style': 'strict'},
    'Amin Mohamed Omar':   {'yc_pg': 3.50, 'fouls_pg': 26.5, 'style': 'average'},
    'Facundo Tello':       {'yc_pg': 4.30, 'fouls_pg': 29.5, 'style': 'strict'},
    'Danny Makkelie':      {'yc_pg': 3.30, 'fouls_pg': 25.5, 'style': 'lenient'},
    'Slavko Vincic':       {'yc_pg': 4.15, 'fouls_pg': 28.5, 'style': 'strict'},
    'Jalal Jayed':         {'yc_pg': 3.80, 'fouls_pg': 27.5, 'style': 'average'},
    # Top arbitri FIFA 2022-2025 (posibili pentru meciuri neconfirmate)
    'Szymon Marciniak':    {'yc_pg': 2.90, 'fouls_pg': 24.0, 'style': 'lenient'},
    'Clement Turpin':      {'yc_pg': 3.60, 'fouls_pg': 27.0, 'style': 'average'},
    'Felix Zwayer':        {'yc_pg': 4.10, 'fouls_pg': 29.0, 'style': 'strict'},
    'Raphael Claus':       {'yc_pg': 3.40, 'fouls_pg': 26.0, 'style': 'average'},
    'Mustapha Ghorbal':    {'yc_pg': 4.50, 'fouls_pg': 30.5, 'style': 'strict'},
    'Jesus Valenzuela':    {'yc_pg': 3.70, 'fouls_pg': 27.5, 'style': 'average'},
    'Cesar Ramos':         {'yc_pg': 3.90, 'fouls_pg': 28.5, 'style': 'average'},
    'Ismail Elfath':       {'yc_pg': 3.00, 'fouls_pg': 25.0, 'style': 'lenient'},
    'Victor Gomes':        {'yc_pg': 4.00, 'fouls_pg': 28.0, 'style': 'strict'},
    'Alireza Faghani':     {'yc_pg': 3.70, 'fouls_pg': 27.0, 'style': 'average'},
    'Janny Sikazwe':       {'yc_pg': 4.20, 'fouls_pg': 29.0, 'style': 'strict'},
    'Ivan Barton':         {'yc_pg': 3.50, 'fouls_pg': 26.5, 'style': 'average'},
}
# Medie globala arbitri internationali (baseline)
REF_AVG_YC    = 3.60
REF_AVG_FOULS = 27.0

def ref_factors(ref_name):
    """Returneaza (factor_yc, factor_fouls) relativ la media globala."""
    r = REFEREE_DB.get(ref_name)
    if not r:
        return 1.0, 1.0
    return r['yc_pg'] / REF_AVG_YC, r['fouls_pg'] / REF_AVG_FOULS

# ─── Predictie V3 per meci ───────────────────────────────────────────────────
# Format: (grupa, acasa, deplasare, data, arbitru)
ROUND1 = [
    ('A', 'Mexico',             'South Africa',           '11/06/2026', 'Wilton Sampaio'),
    ('A', 'South Korea',        'Czechia',                '12/06/2026', 'Amin Mohamed Omar'),
    ('B', 'Canada',             'Bosnia and Herzegovina', '12/06/2026', 'Facundo Tello'),
    ('D', 'United States',      'Paraguay',               '13/06/2026', 'Danny Makkelie'),
    ('B', 'Qatar',              'Switzerland',            '13/06/2026', ''),
    ('C', 'Brazil',             'Morocco',                '14/06/2026', 'Slavko Vincic'),
    ('C', 'Haiti',              'Scotland',               '14/06/2026', ''),
    ('D', 'Australia',          'Turkey',                 '14/06/2026', ''),
    ('E', 'Germany',            'Curaçao',                '14/06/2026', 'Jalal Jayed'),
    ('F', 'Netherlands',        'Japan',                  '14/06/2026', ''),
    ('E', 'Ivory Coast',        'Ecuador',                '15/06/2026', ''),
    ('F', 'Sweden',             'Tunisia',                '15/06/2026', ''),
    ('H', 'Spain',              'Cape Verde',             '15/06/2026', ''),
    ('G', 'Belgium',            'Egypt',                  '15/06/2026', ''),
    ('H', 'Saudi Arabia',       'Uruguay',                '16/06/2026', ''),
    ('I', 'Iran',               'New Zealand',            '16/06/2026', ''),
    ('I', 'France',             'Senegal',                '16/06/2026', ''),
    ('J', 'Iraq',               'Norway',                 '17/06/2026', ''),
    ('J', 'Argentina',          'Algeria',                '17/06/2026', ''),
    ('K', 'Austria',            'Jordan',                 '17/06/2026', ''),
    ('K', 'Portugal',           'DR Congo',               '17/06/2026', ''),
    ('L', 'England',            'Croatia',                '17/06/2026', ''),
    ('L', 'Ghana',              'Panama',                 '18/06/2026', ''),
    ('K', 'Uzbekistan',         'Colombia',               '18/06/2026', ''),
]

def predict_v3(group, home, away, date, clf_res, clf_o25,
               form_buf, h2h_buf, elo_hist,
               w_elo, w_qual, w_form, avg_goals, rho, ml_w=0.55, referee=''):
    helo = ELO.get(home, {}).get('rating', 1700)
    aelo = ELO.get(away, {}).get('rating', 1700)
    fix = FIXTURES_ELO.get((home, away))
    if fix:
        helo = float(fix.get('home_elo') or helo)
        aelo = float(fix.get('away_elo') or aelo)
        elo_wp_h = float(fix.get('win_exp_home_pct', 50))
        elo_wp_a = float(fix.get('win_exp_away_pct', 50))
    else:
        # +100 ELO home advantage DOAR pentru gazde reale (US/Canada/Mexico)
        # la CM toate celelalte meciuri sunt pe teren neutru
        home_adv = 100 if home in WC_HOSTS else 0
        diff = (helo - aelo + home_adv)
        elo_wp_h = 1 / (1 + 10 ** (-diff / 400)) * 100
        elo_wp_a = 100 - elo_wp_h

    fh  = compute_form(form_buf.get(home, []))
    fa  = compute_form(form_buf.get(away, []))
    h2h = compute_h2h(h2h_buf, home, away)

    def elo_tr(team):
        h = elo_hist.get(team, [])[-5:]
        cur = ELO.get(team, {}).get('rating', 1700)
        return (cur - np.mean(h)) if h else 0.0

    feat = extract_features(home, away, fh, fa, h2h, helo, aelo, 'World Cup',
                             elo_tr(home), elo_tr(away))
    feat_arr = np.array([feat], dtype=float)

    # Conditii stadion + absente
    stadium = MATCH_VENUES.get((home, away), MATCH_VENUES.get((away, home), ''))
    cond    = venue_conditions(home, away, stadium)
    abs_h_att, _ = absence_factor(home)
    abs_a_att, _ = absence_factor(away)

    lh, la = poisson_xg(home, away, fh, fa, helo, aelo, w_elo, w_qual, w_form, avg_goals,
                        cond['lh_mult'], cond['la_mult'], abs_h_att, abs_a_att)

    # Generare Matrice Dixon-Coles 8x8
    mat = get_match_matrix(lh, la, rho)
    pa_p = np.triu(mat, 1).sum()
    pd_p = np.trace(mat)
    ph_p = np.tril(mat, -1).sum()

    pois = np.array([[pa_p, pd_p, ph_p]])

    if clf_res is not None:
        ml   = clf_res.predict_proba(feat_arr)
        o25m = float(clf_o25.predict_proba(feat_arr)[0, 1])
    else:
        ml = pois
        o25m = 0.0

    blend = (1 - ml_w) * pois + ml_w * ml
    blend /= blend.sum()

    elo_p = np.array([[elo_wp_a / 100, max(0.05, 1 - elo_wp_h / 100 - elo_wp_a / 100), elo_wp_h / 100]])
    blend = 0.85 * blend + 0.15 * elo_p
    blend /= blend.sum()

    # blend_raw = proba pura a modelului, fara influenta cotelor bookmaker.
    # IMPORTANT: folosita pentru calculul EV — altfel comparatia model vs cote
    # devine circulara (modelul a absorbit deja 30% din cota bookmakerului).
    blend_raw = blend.copy()

    # Blend cu piata (daca exista cote bookmaker) — doar pentru AFISARE
    mkt = market_implied(home, away)
    if mkt:
        mkt_arr = np.array([[mkt['a'], mkt['d'], mkt['h']]])
        blend = 0.70 * blend + 0.30 * mkt_arr
        blend /= blend.sum()

    wp_a, wp_d, wp_h = blend[0, 0] * 100, blend[0, 1] * 100, blend[0, 2] * 100
    wp_a_raw, wp_d_raw, wp_h_raw = (blend_raw[0, 0] * 100,
                                     blend_raw[0, 1] * 100,
                                     blend_raw[0, 2] * 100)

    ryc, rfouls = ref_factors(referee)
    ryc    *= cond['yc_mult']   # arbitru * caldura/altitudine
    rfouls *= cond['yc_mult']
    # Raw probs pentru bets: calculul EV trebuie sa compare proba INDEPENDENTA
    # a modelului cu cotele bookmakerului — nu proba deja contaminata cu 30% piata
    bets = generate_bets_v3(home, away, wp_h_raw, wp_d_raw, wp_a_raw,
                            lh, la, mat, helo, aelo, ryc, rfouls)
    bet_ev = {}
    for label, pred, prob, cat in bets:
        odds = get_odds_for_bet(home, away, cat, label, pred)
        ev = value_score(prob, odds) if odds else None
        if ev is not None:
            bet_ev[label] = ev
    accum = pick_accumulator(bets, n=3, min_prob=55.0, bet_ev=bet_ev)

    # Corelatie precisa Over 2.5 din Matrice
    total_g_mat = np.zeros(18)
    for i in range(9):
        for j in range(9):
            total_g_mat[i+j] += mat[i,j]
    p_o25_dc = total_g_mat[3:].sum() * 100

    return {
        'group': group, 'home': home, 'away': away, 'date': date,
        'referee': referee, 'stadium': stadium, 'conditions': cond['note'],
        'cond_lh': cond['lh_mult'], 'cond_la': cond['la_mult'],
        'helo': int(helo), 'aelo': int(aelo), 'elo_diff': abs(int(helo - aelo)),
        'wp_home': round(wp_h, 1), 'wp_away': round(wp_a, 1), 'p_draw': round(wp_d, 1),
        'wp_home_raw': round(wp_h_raw, 1), 'wp_away_raw': round(wp_a_raw, 1), 'p_draw_raw': round(wp_d_raw, 1),
        'lambda_h': round(lh, 2), 'lambda_a': round(la, 2),
        'total_goals': round(lh + la, 2),
        'p_o15': round(total_g_mat[2:].sum() * 100, 1),
        'p_o25': round(p_o25_dc, 1),
        'p_o35': round(total_g_mat[4:].sum() * 100, 1),
        'p_gg':  round(mat[1:, 1:].sum() * 100, 1),
        'o25_ml': round(o25m * 100, 1),
        'h2h': h2h, 'venue_h': venue_score(home), 'venue_a': venue_score(away),
        'w_elo': round(w_elo, 3), 'w_qual': round(w_qual, 3),
        'w_form': round(w_form, 3), 'avg_goals': round(avg_goals, 3), 'rho': round(rho, 3),
        'bets': bets, 'accum': accum, 'bet_ev': bet_ev,
    }

# ─── Excel ───────────────────────────────────────────────────────────────────
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'),  bottom=Side(style='thin'))
def _f(h): return PatternFill('solid', fgColor=h)
C = {'title': _f('1A1A2E'), 'hdr': _f('0F3460'), 'good': _f('00B050'),
     'ok': _f('92D050'), 'warn': _f('FFEB9C'), 'bad': _f('FF7878'),
     'r0': _f('F2F7FF'), 'r1': _f('FFFFFF'), 'venue': _f('1A5C2A')}

def cp(v):
    if v >= 75: return C['good']
    if v >= 60: return C['ok']
    if v >= 45: return C['warn']
    return C['bad']

def build_excel(metrics_list, cal_data, feat_imp, v2_preds):
    wb = openpyxl.Workbook()

    # ─ Sheet 1: Metrics
    ws = wb.active; ws.title = 'Metrics_Backtest'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'WC2026 Model V3 — No Leakage, Optimized Dixon-Coles (80/20 temporal split)'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    hdrs = ['Model', 'Accuracy 1X2', 'Log-Loss', 'Brier\n(Win)', 'Brier\n(Draw)',
            'Brier\n(Loss)', 'Rating', 'Random baseline']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=8)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 28

    for ri, m in enumerate(metrics_list, 3):
        interp = ('Excellent ★★★' if m['accuracy'] >= 0.56 else
                  'Good ★★'       if m['accuracy'] >= 0.50 else
                  'Acceptable ★'  if m['accuracy'] >= 0.44 else 'Below expectations')
        vals = [m['name'], f"{m['accuracy']:.3f}", f"{m['log_loss']:.4f}",
                f"{m['brier_home']:.4f}", f"{m['brier_draw']:.4f}", f"{m['brier_away']:.4f}",
                interp, 'acc≈0.333 | ll≈1.099']
        bg = C['r0'] if ri % 2 else C['r1']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(size=9); c.border = THIN; c.fill = bg
            c.alignment = Alignment(horizontal='center', vertical='center')
        v = m['accuracy']
        ws.cell(ri, 2).fill = C['good'] if v >= 0.54 else C['ok'] if v >= 0.48 else C['warn']
        ws.row_dimensions[ri].height = 16

    for ci, w in enumerate([28, 14, 12, 12, 10, 14, 18, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    note_row = len(metrics_list) + 4
    for i, (lbl, txt) in enumerate([
        ('Accuracy', 'Note: accuracy dropping to ~65% is expected after leakage removal — this is the real evaluation.'),
        ('Log-Loss', 'Random = 1.099 | Good < 1.05 | Excellent < 0.98'),
    ], note_row):
        ws.cell(i, 1, lbl).font = Font(bold=True, size=8, italic=True)
        c = ws.cell(i, 2, txt); c.font = Font(size=8)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    # ─ Sheet 2: Calibration
    ws2 = wb.create_sheet('Calibration')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:D1')
    c = ws2['A1']
    c.value = 'Calibration — Home Win | Ideal: perfect diagonal'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 22
    for ci, h in enumerate(['Predicted Prob (%)', 'Actual Prob (%)', 'Difference', 'Calibrated?'], 1):
        c = ws2.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ci, w in enumerate([22, 20, 12, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 20
    mean_preds, frac_pos = cal_data
    for ri, (mp, fp) in enumerate(zip(mean_preds, frac_pos), 3):
        diff = fp - mp
        ok = 'Yes ✓' if abs(diff) < 0.08 else ('Acceptable' if abs(diff) < 0.15 else 'No ✗')
        bg = C['good'] if abs(diff) < 0.08 else C['warn'] if abs(diff) < 0.15 else C['bad']
        for ci, v in enumerate([f"{mp*100:.1f}%", f"{fp*100:.1f}%",
                                  f"{diff*100:+.1f}%", ok], 1):
            c = ws2.cell(ri, ci, v)
            c.font = Font(size=9); c.border = THIN
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(ri, 4).fill = bg
        ws2.row_dimensions[ri].height = 15

    # ─ Sheet 3: Feature Importance 
    ws3 = wb.create_sheet('Feature_Importance')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:C1')
    c = ws3['A1']
    c.value = 'Importanta variabilelor (Fara Leakage)'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 22
    for ci, h in enumerate(['Feature', 'Importanta', 'Bara vizuala'], 1):
        c = ws3.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 32
    ws3.row_dimensions[2].height = 20
    mx = feat_imp[0][1] if feat_imp else 1
    for ri, (fn, imp) in enumerate(feat_imp, 3):
        bar = '█' * max(1, int(imp / mx * 24))
        pct = imp / mx * 100
        for ci, v in enumerate([fn, f"{imp:.4f}", bar], 1):
            c = ws3.cell(ri, ci, v)
            c.font = Font(size=9 if ci < 3 else 7); c.border = THIN
            c.alignment = Alignment(horizontal='left' if ci != 2 else 'center',
                                    vertical='center')
        ws3.cell(ri, 2).fill = cp(pct)
        ws3.row_dimensions[ri].height = 14

    # ─ Sheet 4: Predictions V3
    ws4 = wb.create_sheet('Predictions_V3')
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells('A1:P1')
    c = ws4['A1']
    c.value = 'WC2026 Round 1 — Model V3 Dixon-Coles'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 24

    hdrs4 = ['Gr','Date','Home','Away',
             'Win%\nHome','Win%\nDraw','Win%\nAway',
             'xG\nHome','xG\nAway','Total xG',
             'Ov1.5%','Ov2.5%\n(DixonColes)','Ov2.5%\n(ML)','Ov3.5%',
             'GG%', 'H2H / Venue']
    for ci, h in enumerate(hdrs4, 1):
        c = ws4.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=8)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4.row_dimensions[2].height = 34

    for ci, w in enumerate([4,11,22,22, 9,9,9, 7,7,8, 9,9,9,9, 8, 28], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    for ri, p in enumerate(v2_preds, 3):
        bg = C['r0'] if ri % 2 else C['r1']
        vh, va = p['venue_h'], p['venue_a']
        venue_txt = ''
        if vh > 0:   venue_txt = f"Venue+{int(vh*100)}% {p['home'][:10]}"
        elif va > 0: venue_txt = f"Venue+{int(va*100)}% {p['away'][:10]}"
        h2h = p['h2h']
        note = (f"H2H: {round(h2h['win_rate']*100)}%W gd={h2h['avg_gd']:+.1f} "
                f"(N={round(h2h['n_norm']*5)}) | {venue_txt}")
        vals = [p['group'], p['date'], p['home'], p['away'],
                f"{p['wp_home']}%", f"{p['p_draw']}%", f"{p['wp_away']}%",
                p['lambda_h'], p['lambda_a'], p['total_goals'],
                f"{p['p_o15']}%", f"{p['p_o25']}%", f"{p['o25_ml']}%",
                f"{p['p_o35']}%", f"{p['p_gg']}%", note]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(ri, ci, v)
            c.font = Font(size=9); c.border = THIN; c.fill = bg
            c.alignment = Alignment(
                horizontal='center' if ci not in (3,4,16) else 'left',
                vertical='center')
        ws4.cell(ri, 5).fill = cp(p['wp_home'])
        ws4.cell(ri, 7).fill = cp(p['wp_away'])
        ws4.cell(ri, 12).fill = cp(p['p_o25'])
        ws4.cell(ri, 13).fill = cp(p['o25_ml'])
        if vh > 0: ws4.cell(ri, 3).fill = C['venue']
        if va > 0: ws4.cell(ri, 4).fill = C['venue']
        ws4.row_dimensions[ri].height = 16

    fr = len(v2_preds) + 4
    p0 = v2_preds[0]
    note_txt = (f"Optimized Parameters:  "
                f"w_elo={p0['w_elo']}  w_qual={p0['w_qual']}  "
                f"w_form={p0['w_form']}  avg_goals={p0['avg_goals']}  rho={p0['rho']} "
                f"| ML Blend={int(p0.get('blend_weight_ml', 55) if isinstance(p0.get('blend_weight_ml'), (int, float)) else 55)}%")
    ws4.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=16)
    c = ws4.cell(fr, 1, note_txt)
    c.font = Font(italic=True, size=8)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # ─ Sheet 5: All Bets
    ws5 = wb.create_sheet('All_Bets')
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells('A1:G1')
    c = ws5['A1']
    c.value = 'WC2026 Round 1 — All Bets (Model V3 Dixon-Coles)'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 28

    hdrs5 = ['Match', 'Bet', 'Prediction', 'Probability', 'Market Odds', 'Category', 'Confidence']
    for ci, h in enumerate(hdrs5, 1):
        c = ws5.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws5.row_dimensions[2].height = 20
    for ci, w in enumerate([30, 36, 20, 14, 12, 16, 14], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ri5 = 3
    for p in v2_preds:
        match_lbl = f"{p['home']} vs {p['away']}"
        bets = p.get('bets', [])
        accum_labels = {b[0] for b in p.get('accum', [])}
        if not bets:
            ri5 += 1
            continue
        ws5.merge_cells(start_row=ri5, start_column=1, end_row=ri5, end_column=7)
        c = ws5.cell(ri5, 1, f"  {p['group']}  |  {match_lbl}  |  {p['date']}")
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = C['hdr']; c.alignment = Alignment(vertical='center')
        ws5.row_dimensions[ri5].height = 18
        ri5 += 1
        for label, pred, prob, cat in sorted(bets, key=lambda x: -x[2]):
            is_accum = label in accum_labels
            bg = PatternFill('solid', fgColor='FFD700') if is_accum else (
                C['r0'] if ri5 % 2 else C['r1'])
            conf_lbl = ('Very High' if prob >= 75 else
                        'High'      if prob >= 65 else
                        'Medium'    if prob >= 55 else
                        'Low'       if prob >= 45 else 'Very Low')
            mkt_odds = get_odds_for_bet(p['home'], p['away'], cat, label, pred)
            odds_str = f"{mkt_odds:.2f}" if mkt_odds else '—'
            star = ' *** ACCUMULATOR' if is_accum else ''
            vals5 = [match_lbl, label + star, pred, f"{prob:.1f}%", odds_str, cat, conf_lbl]
            for ci, v in enumerate(vals5, 1):
                c = ws5.cell(ri5, ci, v)
                c.font = Font(size=9, bold=is_accum); c.border = THIN; c.fill = bg
                c.alignment = Alignment(horizontal='center' if ci > 2 else 'left',
                                        vertical='center')
            ws5.cell(ri5, 4).fill = cp(prob) if not is_accum else PatternFill('solid', fgColor='FFD700')
            ws5.row_dimensions[ri5].height = 15
            ri5 += 1
        ri5 += 1

    # ─ Sheet 6: Accumulator
    ws6 = wb.create_sheet('Accumulator')
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells('A1:G1')
    c = ws6['A1']
    c.value = 'WC2026 Round 1 — Best 3 combinable bets per match'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[1].height = 28

    hdrs6 = ['Gr.', 'Date', 'Match', 'Bet', 'Prediction', 'Probability', 'Category']
    for ci, h in enumerate(hdrs6, 1):
        c = ws6.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws6.row_dimensions[2].height = 22
    for ci, w in enumerate([5, 12, 30, 36, 20, 14, 16], 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w

    ri6 = 3
    _ACCUM_COLORS = [
        PatternFill('solid', fgColor='00B050'),  
        PatternFill('solid', fgColor='92D050'),  
        PatternFill('solid', fgColor='FFEB9C'),  
    ]
    for p in v2_preds:
        accum = p.get('accum', [])
        match_lbl = f"{p['home']} vs {p['away']}"
        if not accum:
            ws6.merge_cells(start_row=ri6, start_column=1, end_row=ri6, end_column=7)
            ws6.cell(ri6, 1, f"{match_lbl} — niciun pariu cu conf. >= 55%")
            ws6.cell(ri6, 1).font = Font(italic=True, size=8, color='888888')
            ws6.row_dimensions[ri6].height = 14
            ri6 += 1
            continue
        for idx, (label, pred, prob, cat) in enumerate(accum):
            bg = _ACCUM_COLORS[idx] if idx < len(_ACCUM_COLORS) else C['r0']
            vals6 = [p['group'], p['date'], match_lbl, label, pred,
                     f"{prob:.1f}%", cat]
            for ci, v in enumerate(vals6, 1):
                c = ws6.cell(ri6, ci, v)
                c.font = Font(size=9, bold=True); c.border = THIN; c.fill = bg
                c.alignment = Alignment(
                    horizontal='center' if ci not in (3, 4) else 'left',
                    vertical='center')
            ws6.row_dimensions[ri6].height = 16
            ri6 += 1
        ri6 += 1  

    # ─ Sheet 7: Best Bet Per Match
    ws7 = wb.create_sheet('Best_Bet_Per_Match')
    ws7.sheet_view.showGridLines = False
    ws7.merge_cells('A1:G1')
    c = ws7['A1']
    c.value = 'WC2026 — Best single bet for each match (by EV)'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = C['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws7.row_dimensions[1].height = 28

    hdrs7 = ['Round/Group', 'Date', 'Match', 'Best Bet', 'Prediction', 'Probability', 'Category']
    for ci, h in enumerate(hdrs7, 1):
        c = ws7.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = C['hdr']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws7.row_dimensions[2].height = 22
    for ci, w in enumerate([12, 12, 30, 36, 20, 14, 16], 1):
        ws7.column_dimensions[get_column_letter(ci)].width = w

    ri7 = 3
    for p in v2_preds:
        bets = p.get('bets', [])
        if not bets:
            continue
            
        bet_ev_map = p.get('bet_ev', {})
        if bet_ev_map:
            best_label = max(bet_ev_map, key=lambda l: bet_ev_map[l])
            best_bet = next((b for b in bets if b[0] == best_label),
                            max(bets, key=lambda x: x[2]))
        else:
            best_bet = max(bets, key=lambda x: x[2])
        label, pred, prob, cat = best_bet
        
        match_lbl = f"{p['home']} vs {p['away']}"
        bg = C['r0'] if ri7 % 2 else C['r1']
        
        vals7 = [f"Group {p['group']}", p['date'], match_lbl, label, pred, f"{prob:.1f}%", cat]
        
        for ci, v in enumerate(vals7, 1):
            c = ws7.cell(ri7, ci, v)
            c.font = Font(size=9, bold=(ci in [4, 6]))
            c.border = THIN
            c.fill = bg
            c.alignment = Alignment(
                horizontal='center' if ci not in [3, 4] else 'left', 
                vertical='center'
            )
            
        ws7.cell(ri7, 6).fill = cp(prob)
        ws7.row_dimensions[ri7].height = 16
        ri7 += 1

    wb.save('backtest_v3.xlsx')
    print("  Saved: backtest_v3.xlsx")

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("1. Building dataset (temporal split, no leakage)...")
    all_m = sorted(ELO_RAW['recent_matches'],
                   key=lambda m: m.get('date', ''))
    (X, y_res, y_o25, y_gh, y_ga,
     meta, full_form, full_h2h, full_elo_hist) = build_dataset(all_m)
    n = len(X)
    print(f"   Total valid matches: {n}")

    split = int(n * 0.80)
    X_tr, X_te   = X[:split], X[split:]
    yr_tr, yr_te = y_res[:split], y_res[split:]
    yo_tr, yo_te = y_o25[:split], y_o25[split:]
    gh_tr, ga_tr = y_gh[:split], y_ga[:split]
    m_tr, m_te   = meta[:split], meta[split:]
    print(f"   Train: {split} matches  ({m_tr[0]['date']} -> {m_tr[-1]['date']})")
    print(f"   Test:  {n-split} matches ({m_te[0]['date']} -> {m_te[-1]['date']})")

    print("\n2. Optimizing weights & Dixon-Coles RHO factor...")
    w_elo, w_qual, w_form, avg_goals, rho = optimize_poisson(X_tr, gh_tr, ga_tr, m_tr)

    metrics_list = []
    feat_imp = []
    cal_mean, cal_frac = [], []
    clf_res = clf_o25_clf = None

    if HAS_SKLEARN:
        print("\n3. Training ML model (XGBoost / GBM)...")
        clf_res, clf_o25_clf = train_ml(X_tr, yr_tr, yo_tr)
        algo = 'XGBoost' if HAS_XGB else 'GradientBoosting'
        print(f"   Algorithm: {algo} + CalibratedClassifierCV (isotonic)")

        print("\n4. Backtest metrics (temporal test set):")
        pois_p = []
        for i, rec in enumerate(m_te):
            fh = compute_form(full_form.get(rec['home'], []))
            fa = compute_form(full_form.get(rec['away'], []))
            lh, la = poisson_xg(rec['home'], rec['away'], fh, fa,
                                  rec['helo'], rec['aelo'],
                                  w_elo, w_qual, w_form, avg_goals)
            mat = get_match_matrix(lh, la, rho)
            pa_p = np.triu(mat, 1).sum()
            pd_p = np.trace(mat)
            ph_p = np.tril(mat, -1).sum()
            pois_p.append([pa_p, pd_p, ph_p])
        pois_p = np.array(pois_p)
        ml_p   = clf_res.predict_proba(X_te)
        blend  = 0.45 * pois_p + 0.55 * ml_p
        blend /= blend.sum(axis=1, keepdims=True)

        m1 = metrics(yr_te, pois_p, 'Poisson Dixon-Coles Optimizat')
        m2 = metrics(yr_te, ml_p,   f'{algo} Calibrat (Fara Leakage)')
        m3 = metrics(yr_te, blend,  'Blend 45% Poisson + 55% ML')
        metrics_list = [m1, m2, m3]

        feat_imp = get_feature_importance(clf_res)
        if feat_imp:
            print(f"\n   Top 5 features (Corecte):")
            for fn, imp in feat_imp[:5]:
                print(f"     {fn:28} {imp:.4f}")

        hw = (yr_te == 2).astype(int)
        try:
            cal_frac, cal_mean = calibration_curve(hw, blend[:, 2], n_bins=8)
            cal_mean = list(cal_mean); cal_frac = list(cal_frac)
        except Exception:
            cal_mean, cal_frac = [], []

    print("\n5. Round 1 Predictions — Model V3 (Matrix + Dixon-Coles)...")
    v2_preds = []
    for group, home, away, date, referee in ROUND1:
        p = predict_v3(group, home, away, date,
                       clf_res, clf_o25_clf,
                       full_form, full_h2h, full_elo_hist,
                       w_elo, w_qual, w_form, avg_goals, rho,
                       referee=referee)
        p['blend_weight_ml'] = 55
        v2_preds.append(p)
        vh = ' [VENUE+]' if p['venue_h'] > 0 or p['venue_a'] > 0 else ''
        print(f"  Gr.{group} {home[:12]:12} vs {away[:12]:12}  "
              f"W={p['wp_home']:5.1f}% D={p['p_draw']:4.1f}% L={p['wp_away']:5.1f}%  "
              f"xG {p['lambda_h']:.2f}-{p['lambda_a']:.2f}  "
              f"O2.5={p['p_o25']}%{vh}")

    # ─── VALUE BETS: pariuri cu EV pozitiv fata de bookmaker ─────────────────
    value_bets = []
    for p in v2_preds:
        h, a = p['home'], p['away']
        for label, pred, prob, cat in p.get('bets', []):
            odds = get_odds_for_bet(h, a, cat, label, pred)
            if odds:
                ev = value_score(prob, odds)
                if ev is not None:
                    value_bets.append((ev, prob, odds, h, a, p['date'], label, pred, cat,
                                       p.get('stadium',''), p.get('conditions',''),
                                       p.get('referee','')))

    if value_bets:
        value_bets.sort(reverse=True)
        print("\n=== POSITIVE EV BETS (EV+) ===")
        print(f"  {'EV':>6}  {'Prob':>5}  {'Odds':>5}  {'Match':<28}  {'Bet':<28}  {'Prediction'}")
        print("  " + "-"*100)
        for ev, prob, odds, h, a, date, label, pred, cat, stad, cond, ref in value_bets:
            tag = " *** EV+" if ev > 0.05 else ""
            print(f"  {ev:>+6.3f}  {prob:>4.1f}%  {odds:>5.2f}  "
                  f"{h[:13]+' vs '+a[:13]:<28}  {label:<28}  {pred}{tag}")
    else:
        print("\n  (Fill MARKET_ODDS in code with real bookmaker odds to enable VALUE analysis)")

    print("\n6. Building Excel...")
    build_excel(metrics_list, (cal_mean, cal_frac), feat_imp, v2_preds)

    avg_g = sum(p['total_goals'] for p in v2_preds) / len(v2_preds)
    n_o25 = sum(1 for p in v2_preds if p['p_o25'] > 55)
    print(f"\n   Avg xG/match: {avg_g:.2f}  |  Matches with O2.5>55%: {n_o25}/24")
    print("Done!")