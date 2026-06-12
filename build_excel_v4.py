"""Build Excel v4 with AiScore team and player stats for all confederations."""
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from curl_cffi import requests as cffi_requests

# ─── AiScore protobuf decoder ─────────────────────────────────────────────────
def _rv(data, pos):
    r = 0; s = 0
    while pos < len(data):
        b = data[pos]; pos += 1
        r |= (b & 0x7F) << s
        if not (b & 0x80): break
        s += 7
    return r, pos


def _pp(data, pos=0, end=None):
    if end is None: end = len(data)
    fields = []
    while pos < end:
        try: tv, pos = _rv(data, pos)
        except Exception: break
        fn = tv >> 3; wt = tv & 0x7
        try:
            if wt == 0: v, pos = _rv(data, pos); fields.append((fn, 'v', v))
            elif wt == 2:
                ln, pos = _rv(data, pos); v = data[pos:pos+ln]; pos += ln
                fields.append((fn, 'b', v))
            elif wt == 1: pos += 8
            elif wt == 5: pos += 4
            else: break
        except Exception: break
    return fields


def decode_team_stats(data):
    """Parse kind=1 team totals response. Returns {team_name: stat_value}."""
    outer = _pp(data)
    if not outer: return {}
    main = _pp(outer[0][2])
    teams = {}; entries = []
    for rec in main:
        if rec[1] != 'b': continue
        sub = _pp(rec[2])
        if rec[0] == 1:
            tid = ''; name = ''
            for sf in sub:
                if sf[0] == 1 and sf[1] == 'b': tid = sf[2].decode('utf-8', 'replace').strip('\x00')
                elif sf[0] == 6 and sf[1] == 'b': name = sf[2].decode('utf-8', 'replace')
            if tid: teams[tid] = name
        elif rec[0] == 3:
            for sf in sub:
                if sf[0] != 3 or sf[1] != 'b': continue
                entry = _pp(sf[2])
                tid = ''; val = ''
                for ef in entry:
                    if ef[0] == 1 and ef[1] == 'b':
                        sub2 = _pp(ef[2])
                        for sf2 in sub2:
                            if sf2[0] == 1 and sf2[1] == 'b': tid = sf2[2].decode('utf-8', 'replace').strip('\x00')
                    elif ef[0] == 3 and ef[1] == 'b': val = ef[2].decode('utf-8', 'replace')
                if tid: entries.append((tid, val))
    result = {}
    for tid, val in entries:
        name = teams.get(tid)
        if name: result[name] = val
    return result


def decode_player_stats(data):
    """Parse kind=0 player rankings. Returns list of (team_name, player_name, value)."""
    outer = _pp(data)
    if not outer: return [], {}
    main = _pp(outer[0][2])
    teams = {}; players = {}; entries = []
    for rec in main:
        if rec[1] != 'b': continue
        sub = _pp(rec[2])
        if rec[0] == 1:
            tid = ''; name = ''
            for sf in sub:
                if sf[0] == 1 and sf[1] == 'b': tid = sf[2].decode('utf-8', 'replace').strip('\x00')
                elif sf[0] == 6 and sf[1] == 'b': name = sf[2].decode('utf-8', 'replace')
            if tid: teams[tid] = name
        elif rec[0] == 2:
            pid = ''; name = ''
            for sf in sub:
                if sf[0] == 1 and sf[1] == 'b': pid = sf[2].decode('utf-8', 'replace').strip('\x00')
                elif sf[0] == 4 and sf[1] == 'b': name = sf[2].decode('utf-8', 'replace')
            if pid: players[pid] = name
        elif rec[0] == 3:
            for sf in sub:
                if sf[0] != 3 or sf[1] != 'b': continue
                entry = _pp(sf[2])
                tid = ''; pid = ''; val = ''
                for ef in entry:
                    if ef[0] == 1 and ef[1] == 'b':
                        sub2 = _pp(ef[2])
                        for sf2 in sub2:
                            if sf2[0] == 1 and sf2[1] == 'b': tid = sf2[2].decode('utf-8', 'replace').strip('\x00')
                    elif ef[0] == 2 and ef[1] == 'b':
                        sub2 = _pp(ef[2])
                        for sf2 in sub2:
                            if sf2[0] == 1 and sf2[1] == 'b': pid = sf2[2].decode('utf-8', 'replace').strip('\x00')
                    elif ef[0] == 3 and ef[1] == 'b': val = ef[2].decode('utf-8', 'replace')
                entries.append((teams.get(tid, tid[:8]), players.get(pid, pid[:8]), val))
    return entries, players


# ─── HTTP session ─────────────────────────────────────────────────────────────
API_URL = 'https://api.aiscore.com/v1/web/api/football/comp/stats'

TOURNAMENTS = {
    'CONMEBOL': '8vrqwnid45hvqn2',
    'UEFA':     'yzrkn6iz39tgqle',
    'CAF':      '2j374oigx8a2qo6',
    'AFC':      '5wv78xi4g1iekrj',
    'OFC':      'w34kgmirnduzko9',
}

TEAM_STAT_TYPES = {
    'goals': 2, 'assists': 3, 'red_cards': 4, 'yellow_cards': 5,
    'total_shots': 6, 'shots_on_target': 7, 'clearances': 8,
    'key_passes': 10, 'crosses': 11, 'crosses_acc': 12, 'fouls': 13, 'was_fouled': 14,
}

# Player stat types from web_fb_StatsType (kind=0 player rankings)
# Different from team stat types!
PLAYER_STAT_TYPES = {
    'goals': 3,           # player type 3 = Goals
    'assists': 4,         # player type 4 = Assists
    'yellow_cards': 7,    # player type 7 = Yellow cards
    'total_shots': 8,     # player type 8 = Total Shots
    'shots_on_target': 9, # player type 9 = Shots on Target
    'key_passes': 16,     # player type 16 = Key passes
}


def make_session():
    s = cffi_requests.Session(impersonate='chrome124')
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
    })
    try:
        s.get('https://www.aiscore.com/', timeout=15)
        s.get('https://www.aiscore.com/tournament-fifa-world-cup-qualification-conmebol/r1edq09i0fyqxgo/stats', timeout=15)
        time.sleep(0.5)
    except Exception:
        pass
    s.headers['Accept'] = 'application/json, text/plain, */*'
    s.headers['Referer'] = 'https://www.aiscore.com/'
    try:
        s.get(API_URL, params={'lang': 2, 'season_id': '8vrqwnid45hvqn2', 'type': 2, 'n': 1, 'kind': 0}, timeout=10)
    except Exception:
        pass
    return s


def api_get(session, season_id, stat_type, kind, n=1, retry=True):
    params = {'lang': 2, 'season_id': season_id, 'type': stat_type, 'n': n, 'kind': kind}
    try:
        r = session.get(API_URL, params=params, timeout=15)
        if r.status_code != 200 or r.content[:1] != b'\x7a':
            if retry:
                time.sleep(2)
                session2 = make_session()
                session.cookies.update(session2.cookies)
                r = session.get(API_URL, params=params, timeout=15)
        if r.content[:1] == b'\x7a':
            return r.content
    except Exception as e:
        print(f'    API error: {e}')
    return b''


def parse_value(val_str):
    """Extract numeric part from '31', '28(2)', '37(1)'."""
    if not val_str:
        return None
    idx = val_str.find('(')
    main = val_str[:idx].strip() if idx > 0 else val_str.strip()
    try:
        return int(main)
    except ValueError:
        try:
            return float(main)
        except ValueError:
            return None


def parse_pens(val_str):
    """Extract penalty count from '28(2)' → 2."""
    if not val_str:
        return None
    idx = val_str.find('(')
    if idx < 0:
        return None
    end = val_str.find(')', idx)
    try:
        return int(val_str[idx+1:end])
    except ValueError:
        return None


# ─── Name normalization ────────────────────────────────────────────────────────
AISCORE_TO_STD = {
    "Cote d'Ivoire":                  'Ivory Coast',
    'Turkiye':                         'Türkiye',
    'IR Iran':                         'Iran',
    'Democratic Republic of the Congo':'DR Congo',
    'USA':                             'United States',
    'United States':                   'United States',
}


def normalize(name):
    return AISCORE_TO_STD.get(name, name)


# ─── Excel helpers ────────────────────────────────────────────────────────────
CONF_COLORS = {
    'CONMEBOL': 'FFF2CC',
    'UEFA':     'DEEAF1',
    'CAF':      'E2EFDA',
    'AFC':      'FCE4D6',
    'OFC':      'EAD1DC',
}

HDR_FILL = PatternFill(fgColor='2F4F8F', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(size=10)

thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER


def style_row(ws, row_idx, conf=None):
    fill_color = CONF_COLORS.get(conf, 'FFFFFF') if conf else 'FFFFFF'
    fill = PatternFill(fgColor=fill_color, fill_type='solid')
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = BODY_FONT
            cell.border = BORDER


def auto_width(ws, min_w=8, max_w=25):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        col_max = max_w
        for cell in col:
            if cell.value:
                col_max = max(col_max, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[col_letter].width = max(min_w, col_max)


# ─── Main build ───────────────────────────────────────────────────────────────
def main():
    print('Setting up session...')
    session = make_session()

    # ── Collect team stats ──────────────────────────────────────────────────
    print('\nCollecting team stats...')
    team_data = {}  # {conf: {team_name: {stat: value}}}

    for conf, season_id in TOURNAMENTS.items():
        team_data[conf] = {}
        print(f'  {conf}:')
        for stat_key, stat_type in TEAM_STAT_TYPES.items():
            raw = api_get(session, season_id, stat_type, kind=1)
            if not raw:
                print(f'    {stat_key}: no data')
                continue
            stat_map = decode_team_stats(raw)
            print(f'    {stat_key}: {len(stat_map)} teams')
            for ai_name, val in stat_map.items():
                std_name = normalize(ai_name)
                if std_name not in team_data[conf]:
                    team_data[conf][std_name] = {}
                team_data[conf][std_name][stat_key] = val
            time.sleep(0.3)

    # ── Collect player stats ─────────────────────────────────────────────────
    print('\nCollecting player stats...')
    player_rows = []  # list of dicts

    for conf, season_id in TOURNAMENTS.items():
        print(f'  {conf}:')
        for stat_key, stat_type in PLAYER_STAT_TYPES.items():
            raw = api_get(session, season_id, stat_type, kind=0)
            if not raw:
                continue
            entries, _ = decode_player_stats(raw)
            print(f'    {stat_key}: {len(entries)} players')
            for team_ai, player, val in entries:
                team_std = normalize(team_ai)
                player_rows.append({
                    'confederation': conf,
                    'team': team_std,
                    'player': player,
                    'stat': stat_key,
                    'value': val,
                    'value_num': parse_value(val) or 0,
                })
            time.sleep(0.3)

    # ── Build Excel ──────────────────────────────────────────────────────────
    print('\nBuilding Excel...')
    wb = openpyxl.load_workbook('statistici_calificari_cm2026_v3.xlsx')

    # Remove existing AiScore sheets if present
    for sh_name in ['AiScore_Echipe', 'AiScore_Jucatori']:
        if sh_name in wb.sheetnames:
            del wb[sh_name]

    # ── Sheet 1: Team stats ───────────────────────────────────────────────────
    ws_teams = wb.create_sheet('AiScore_Echipe')

    stat_cols = ['goals', 'goals_pk', 'assists', 'red_cards', 'yellow_cards',
                 'total_shots', 'shots_on_target', 'clearances',
                 'key_passes', 'crosses', 'crosses_acc', 'fouls', 'was_fouled']
    headers = ['Confederatie', 'Echipa',
               'Goluri', 'Goluri(PK)', 'Pase decisive', 'Crt Rosii', 'Crt Galbene',
               'Suturi Total', 'Suturi pe Poarta', 'Degajari',
               'Pase Cheie', 'Centrari', 'Centrari Precise', 'Faulturi Comise', 'Faulturi Primite']

    ws_teams.append(headers)
    style_header(ws_teams, 1)

    row_num = 2
    for conf, teams in team_data.items():
        for team_name in sorted(teams.keys()):
            stats = teams[team_name]
            goals_raw = stats.get('goals', '')
            row = [
                conf, team_name,
                parse_value(stats.get('goals', '')),
                parse_pens(stats.get('goals', '')),
                parse_value(stats.get('assists', '')),
                parse_value(stats.get('red_cards', '')),
                parse_value(stats.get('yellow_cards', '')),
                parse_value(stats.get('total_shots', '')),
                parse_value(stats.get('shots_on_target', '')),
                parse_value(stats.get('clearances', '')),
                parse_value(stats.get('key_passes', '')),
                parse_value(stats.get('crosses', '')),
                parse_value(stats.get('crosses_acc', '')),
                parse_value(stats.get('fouls', '')),
                parse_value(stats.get('was_fouled', '')),
            ]
            ws_teams.append(row)
            style_row(ws_teams, row_num, conf)
            row_num += 1

    ws_teams.freeze_panes = 'C2'
    auto_width(ws_teams)
    ws_teams.row_dimensions[1].height = 40

    # ── Sheet 2: Player stats ─────────────────────────────────────────────────
    ws_players = wb.create_sheet('AiScore_Jucatori')

    p_headers = ['Confederatie', 'Echipa', 'Jucator', 'Statistica', 'Valoare', 'Valoare Nr']
    ws_players.append(p_headers)
    style_header(ws_players, 1)

    # Sort by stat then by value descending
    player_rows.sort(key=lambda x: (x['stat'], -x['value_num']))
    for i, row_d in enumerate(player_rows, start=2):
        ws_players.append([
            row_d['confederation'], row_d['team'], row_d['player'],
            row_d['stat'], row_d['value'], row_d['value_num']
        ])
        style_row(ws_players, i, row_d['confederation'])

    ws_players.freeze_panes = 'D2'
    auto_width(ws_players)
    ws_players.row_dimensions[1].height = 30

    # ── Save ──────────────────────────────────────────────────────────────────
    out_file = 'statistici_calificari_cm2026_v4.xlsx'
    wb.save(out_file)
    print(f'\nSaved: {out_file}')

    # Summary
    total_teams = sum(len(t) for t in team_data.values())
    print(f'Teams: {total_teams}')
    print(f'Player rows: {len(player_rows)}')


if __name__ == '__main__':
    main()
