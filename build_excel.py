"""Scrie rezultatele în Excel.

Foaia "Meciuri": date brute, un rând per meci.
Foaia "Sumar_Echipe": agregate pe echipă — calculate cu FORMULE Excel
(SUMIF/AVERAGEIF/COUNTIF) ce referă foaia Meciuri, deci se recalculează
dacă editezi datele brute. "Cel mai bun jucător" e valoare (rating mediu).
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from collect import best_player

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")

# Ordinea coloanelor în foaia Meciuri (litera contează pentru formule)
RAW_COLS = [
    "echipa", "adversar", "data", "competitie", "teren",
    # Atac
    "xg", "suturi_total", "suturi_pe_poarta", "suturi_pe_langa",
    "suturi_blocate", "suturi_in_careu", "suturi_afara_careu",
    "sanse_mari", "sanse_mari_ratate", "bara",
    # Posesie
    "posesie", "pase_total", "pase_precise", "centrari",
    "mingi_lungi", "dribling", "intrari_treime_finala", "mingi_prin_aparare",
    # Set piece-uri
    "cornere", "aruncari_de_la_margine", "lovituri_libere", "lovituri_de_poarta",
    # Apărare
    "fault_comise", "fault_suferite_zona", "ofsaid",
    "galbene", "rosii",
    "recuperari", "interceptari", "degajari",
    "dueluri_aeriene", "dueluri_sol",
    "tackle_reusit", "tackle_total", "save_uri",
]
RAW_HEADERS = [
    "Echipă", "Adversar", "Dată", "Competiție", "Teren",
    "xG", "Șuturi total", "Șuturi pe poartă", "Șuturi pe lângă",
    "Șuturi blocate", "Șuturi în careu", "Șuturi afara careului",
    "Șanse mari", "Șanse mari ratate", "Bara",
    "Posesie (%)", "Pase total", "Pase precise", "Centrări",
    "Mingi lungi", "Dribling", "Intrări treime finală", "Mingi prin apărare",
    "Cornere", "Aruncări margine", "Lovituri libere", "Lovituri de poartă",
    "Fault comise", "Fault suferite (treime fin.)", "Ofsaid",
    "Galbene", "Roșii",
    "Recuperări", "Interceptări", "Degajări",
    "Dueluri aeriene", "Dueluri sol",
    "Tackle reușit", "Tackle total", "Save-uri",
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 32)


def write_matches(wb, matches):
    ws = wb.active
    ws.title = "Meciuri"
    ws.append(RAW_HEADERS)
    for m in matches:
        ws.append([m.get(c) for c in RAW_COLS])
    _style_header(ws, len(RAW_HEADERS))
    ws.freeze_panes = "A2"
    _autowidth(ws)
    return ws


def write_summary(wb, matches, players):
    ws = wb.create_sheet("Sumar_Echipe")

    # Coloane sumar: (header, col_raw, tip)  tip: sum / avg
    SUMMARY_COLS = [
        ("Meciuri",                  None,                    "count"),
        ("xG (tot.)",                "xg",                    "sum"),
        ("xG mediu/meci",            "xg",                    "avg"),
        ("Șuturi total",             "suturi_total",          "sum"),
        ("Șuturi/meci",              "suturi_total",          "avg"),
        ("Șuturi pe poartă",         "suturi_pe_poarta",      "sum"),
        ("Șuturi/poartă per meci",   "suturi_pe_poarta",      "avg"),
        ("Șuturi în careu",          "suturi_in_careu",       "sum"),
        ("Șanse mari",               "sanse_mari",            "sum"),
        ("Cornere (tot.)",           "cornere",               "sum"),
        ("Cornere/meci",             "cornere",               "avg"),
        ("Posesie medie (%)",        "posesie",               "avg"),
        ("Pase precise (tot.)",      "pase_precise",          "sum"),
        ("Centrări (tot.)",          "centrari",              "sum"),
        ("Fault comise (tot.)",      "fault_comise",          "sum"),
        ("Galbene (tot.)",           "galbene",               "sum"),
        ("Roșii (tot.)",             "rosii",                 "sum"),
        ("Ofsaid (tot.)",            "ofsaid",                "sum"),
        ("Recuperări (tot.)",        "recuperari",            "sum"),
        ("Interceptări (tot.)",      "interceptari",          "sum"),
        ("Degajări (tot.)",          "degajari",              "sum"),
        ("Save-uri (tot.)",          "save_uri",              "sum"),
        ("Cel mai bun jucător",      None,                    "player"),
        ("Rating mediu",             None,                    "rating"),
    ]

    headers = ["Echipă"] + [c[0] for c in SUMMARY_COLS]
    ws.append(headers)

    L = {c: get_column_letter(i + 1) for i, c in enumerate(RAW_COLS)}
    key = "Meciuri!$" + L["echipa"] + ":$" + L["echipa"]

    def sumif(col, tc):
        rng = "Meciuri!$" + L[col] + ":$" + L[col]
        return f"=SUMIF({key},{tc},{rng})"

    def avgif(col, tc):
        rng = "Meciuri!$" + L[col] + ":$" + L[col]
        return f"=IFERROR(ROUND(AVERAGEIF({key},{tc},{rng}),2),\"\")"

    seen = []
    for t in matches:
        if t["echipa"] not in seen:
            seen.append(t["echipa"])

    for r, team in enumerate(seen, start=2):
        tc = f"$A{r}"
        bp_name, bp_rating = best_player(players.get(team, {}))
        ws.cell(row=r, column=1, value=team)
        for col_i, (_, raw_col, tip) in enumerate(SUMMARY_COLS, start=2):
            if tip == "count":
                ws.cell(row=r, column=col_i, value=f"=COUNTIF({key},{tc})")
            elif tip == "sum" and raw_col:
                ws.cell(row=r, column=col_i, value=sumif(raw_col, tc))
            elif tip == "avg" and raw_col:
                ws.cell(row=r, column=col_i, value=avgif(raw_col, tc))
            elif tip == "player":
                ws.cell(row=r, column=col_i, value=bp_name or "—")
            elif tip == "rating":
                ws.cell(row=r, column=col_i, value=bp_rating)

    _style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=1):
        for c in row:
            if c.row > 1 and c.font.name != FONT:
                c.font = Font(name=FONT)
    _autowidth(ws)
    return ws


def export(matches, players, path=None):
    path = path or config.OUTPUT_XLSX
    wb = Workbook()
    write_matches(wb, matches)
    write_summary(wb, matches, players)
    wb.save(path)
    return path
