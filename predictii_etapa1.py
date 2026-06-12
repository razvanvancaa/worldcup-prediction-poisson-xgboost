"""WC2026 Round 1 prediction model — Poisson + Elo + Recent Form.

Data sources used:
  1. Elo ratings (eloratings.net via elo_data.json) — strength + win probability
  2. AiScore qualifying stats (aiscore_team_stats.json) — shots, cards, fouls
  3. Recent form: last 10 matches per team from elo_data.json recent_matches
       - Official matches weighted 1.5-2.0x vs friendlies 0.5x
       - Opponent Elo incorporated to adjust quality of recent results
  4. wc_fixtures: pre-computed Elo win probabilities for each WC2026 match

Outputs: predictii_etapa1.xlsx
  - Sheet 1: Sumar_Etapa1       — overview table, all 24 matches
  - Sheet 2: Predictii_Detaliate — per-match detailed markets
  - Sheet 3: Best_Bets          — ALL bets sorted by confidence descending
  - Sheet 4: Forma_Recenta      — last 10 matches used per team (with weights)
"""
import json, math, re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Load data ────────────────────────────────────────────────────────────────
with open('aiscore_team_stats.json', encoding='utf-8') as f:
    AISCORE_RAW = json.load(f)
with open('elo_data.json', encoding='utf-8') as f:
    ELO_RAW = json.load(f)

# ─── Name normalization ───────────────────────────────────────────────────────
ALIAS = {
    "Turkiye": "Turkey", "Türkiye": "Turkey",
    "Cote d'Ivoire": "Ivory Coast", "Côte d'Ivoire": "Ivory Coast",
    "IR Iran": "Iran",
    "Democratic Republic of the Congo": "DR Congo",
    "USA": "United States",
    "Curacao": "Curaçao",
    "Cabo Verde": "Cape Verde",
    "Bosnia and Herzegovina": "Bosnia and Herzegovina",
    "Bosnia & Herzegovina": "Bosnia and Herzegovina",
    "Korea Republic": "South Korea",
    "Republic of Korea": "South Korea",
}

def canon(name):
    return ALIAS.get(name, name)

# Flatten AiScore: canonical_name -> stats dict
AISCORE = {}
for conf, teams in AISCORE_RAW.items():
    for name, stats in teams.items():
        c = canon(name)
        if c not in AISCORE:
            AISCORE[c] = {'_conf': conf}
        for k, v in stats.items():
            if k not in AISCORE[c]:
                AISCORE[c][k] = v

# Elo dict: canonical_name -> elo_row
ELO = {canon(r['team']): r for r in ELO_RAW['wc_ratings']}

# WC fixtures elo predictions: (home_name, away_name) -> fixture_row
FIXTURES_ELO = {}
for r in ELO_RAW['wc_fixtures']:
    hname = canon(r['home']); aname = canon(r['away'])
    FIXTURES_ELO[(hname, aname)] = r
    FIXTURES_ELO[(aname, hname)] = {
        'home': r['away'], 'away': r['home'],
        'home_elo': r.get('away_elo', 1700), 'away_elo': r.get('home_elo', 1700),
        'home_rank': r.get('away_rank', 100), 'away_rank': r.get('home_rank', 100),
        'win_exp_home_pct': r.get('win_exp_away_pct', 50),
        'win_exp_away_pct': r.get('win_exp_home_pct', 50),
    }

# ─── Tournament official-match weighting ─────────────────────────────────────
TOUR_WEIGHT = {
    'World Cup': 2.0,
    'World Cup qualifier': 1.5,
    'World Cup and Asian Cup qualifier': 1.5,
    'European Championship': 1.7,
    'European Championship qualifier': 1.3,
    'Copa America': 1.7,
    'African Nations Cup': 1.4,
    'African Nations Cup qualifier': 1.2,
    'Asian Cup': 1.4,
    'Asian Cup qualifier': 1.2,
    'CONCACAF Championship': 1.4,
    'CONCACAF Nations League A': 1.3,
    'CONCACAF Nations League': 1.2,
    'CONCACAF Nations League B': 1.1,
    'European Nations League A': 1.3,
    'European Nations League B': 1.2,
    'European Nations League C': 1.1,
    'European Nations League': 1.2,
    'European Nations League A/B Play-off': 1.2,
    'Arab Cup': 1.0,
    'Gulf Cup': 0.8,
    'East Asian Championship': 1.0,
    'Central Asian Nations Cup': 0.8,
    'COSAFA Cup': 0.7,
    'Friendly': 0.5,
    'Friendly tournament': 0.5,
    'FIFA Series': 0.5,
    'FIFA Series and Capital Cup': 0.5,
    "King's Cup": 0.6,
    'Kirin Cup': 0.6,
    'Kirin Challenge Cup': 0.6,
}

def match_weight(tournament):
    return TOUR_WEIGHT.get(tournament, 1.0)

def is_official(tournament):
    return match_weight(tournament) > 0.75


# ─── Recent form extraction ───────────────────────────────────────────────────
RECENT_BY_TEAM = defaultdict(list)
for m in ELO_RAW['recent_matches']:
    h, a = canon(m['home']), canon(m['away'])
    w = match_weight(m['tournament'])
    RECENT_BY_TEAM[h].append({
        'date': m['date'], 'role': 'home',
        'goals_for': m['home_score'], 'goals_against': m['away_score'],
        'opp': a, 'opp_elo': m.get('away_elo', 1700),
        'my_elo': m.get('home_elo', 1700),
        'tournament': m['tournament'], 'weight': w,
    })
    RECENT_BY_TEAM[a].append({
        'date': m['date'], 'role': 'away',
        'goals_for': m['away_score'], 'goals_against': m['home_score'],
        'opp': h, 'opp_elo': m.get('home_elo', 1700),
        'my_elo': m.get('away_elo', 1700),
        'tournament': m['tournament'], 'weight': w,
    })

for team in RECENT_BY_TEAM:
    RECENT_BY_TEAM[team].sort(key=lambda x: x['date'])

def recent_form(team, n=10):
    matches = RECENT_BY_TEAM.get(team, [])[-n:]
    if not matches:
        return {'wpg': 1.2, 'wcpg': 1.2, 'win_rate': 0.4, 'avg_opp_elo': 1700,
                'n_official': 0, 'n_total': 0, 'official_ratio': 0.5,
                'lambda_recent': 1.2, 'matches_summary': []}

    total_w = sum(m['weight'] for m in matches) or len(matches)
    wpg  = sum(m['goals_for']     * m['weight'] for m in matches) / total_w
    wcpg = sum(m['goals_against'] * m['weight'] for m in matches) / total_w
    wins = sum(m['weight'] for m in matches if m['goals_for'] > m['goals_against'])
    win_rate = wins / total_w
    opp_elos = [m['opp_elo'] for m in matches]
    avg_opp  = sum(opp_elos) / len(opp_elos)
    # Quality-adjusted goals per game
    elo_adj_goals = sum(m['goals_for'] * m['weight'] * (1 + (m['opp_elo'] - 1700) / 3000)
                        for m in matches) / total_w
    n_official = sum(1 for m in matches if is_official(m['tournament']))

    return {
        'wpg': wpg, 'wcpg': wcpg, 'win_rate': win_rate,
        'avg_opp_elo': avg_opp, 'n_official': n_official,
        'n_total': len(matches), 'official_ratio': n_official / len(matches),
        'lambda_recent': max(0.3, elo_adj_goals),
        'matches_summary': [(m['date'], m['opp'], m['goals_for'], m['goals_against'],
                              m['tournament'], round(m['weight'], 1)) for m in matches[-5:]],
    }


# ─── AiScore helpers ─────────────────────────────────────────────────────────
CONF_GAMES = {
    'CONMEBOL': 18, 'UEFA': 10, 'CAF': 10,
    'AFC': 18, 'CONCACAF': 18, 'OFC': 4,
}

def _num(s):
    if s is None: return None
    m = re.match(r'^(-?\d+(?:\.\d+)?)', str(s).strip().replace('−', '-'))
    return float(m.group(1)) if m else None

def get_conf(team):
    return AISCORE.get(team, {}).get('_conf', 'UEFA')

def est_games(team):
    conf = get_conf(team)
    base = CONF_GAMES.get(conf, 10)
    shots = _num(AISCORE.get(team, {}).get('total_shots'))
    if shots and shots > 0:
        return max(4, min(base, round(shots / 13.0)))
    return base

def per_game(team, stat):
    val = _num(AISCORE.get(team, {}).get(stat))
    if val is None: return None
    g = est_games(team)
    return val / g if g > 0 else None


# ─── Player data ─────────────────────────────────────────────────────────────
def load_players(wb_path):
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    ws = wb['AiScore_Jucatori']
    players = {}
    for row in list(ws.rows)[1:]:
        vals = [c.value for c in row]
        if len(vals) < 6: continue
        _, team, player, stat, _, val_num = vals[:6]
        if not team or not player: continue
        ct = canon(team)
        if ct not in players: players[ct] = {}
        if stat not in players[ct]: players[ct][stat] = []
        players[ct][stat].append((player, val_num or 0))
    wb.close()
    return players

PLAYERS = load_players('statistici_calificari_cm2026_v6.xlsx')


# ─── WC calibration ──────────────────────────────────────────────────────────
WC_AVG_GOALS      = 2.5
WC_AVG_SHOTS_TEAM = 13.0

# Factori per-confederatie: statisticile din calificari nu sunt comparabile
# intre confederatii (CAF/OFC joaca impotriva adversarilor mult mai slabi
# decat UEFA/CONMEBOL, deci cifrele din calificari sunt inflate diferit)
CONF_QUAL_TO_WC = {
    'CONMEBOL': {'goals': 0.90, 'shots': 0.92, 'cards': 0.90, 'fouls': 0.88},
    'UEFA':     {'goals': 0.85, 'shots': 0.90, 'cards': 0.92, 'fouls': 0.90},
    'CONCACAF': {'goals': 0.78, 'shots': 0.78, 'cards': 0.88, 'fouls': 0.87},
    'CAF':      {'goals': 0.70, 'shots': 0.68, 'cards': 0.85, 'fouls': 0.85},
    'AFC':      {'goals': 0.73, 'shots': 0.72, 'cards': 0.86, 'fouls': 0.86},
    'OFC':      {'goals': 0.62, 'shots': 0.60, 'cards': 0.82, 'fouls': 0.83},
}

def qual_to_wc(team, stat):
    conf = get_conf(team)
    return CONF_QUAL_TO_WC.get(conf, CONF_QUAL_TO_WC['UEFA']).get(stat, 0.85)


# ─── Poisson helpers ─────────────────────────────────────────────────────────
def ppf(k, lam):
    if lam <= 0: return 1.0 if k == 0 else 0.0
    return math.exp(-lam) * (lam ** k) / math.factorial(k)

def pcdf(k, lam):
    return sum(ppf(i, lam) for i in range(k + 1))

def p_over_half(threshold, lam):
    return 1.0 - pcdf(int(threshold), lam)


# ─── 24 Fixtures ─────────────────────────────────────────────────────────────
ROUND1 = [
    ('A', 'Mexico',             'South Africa',          '11/06/2026'),
    ('A', 'South Korea',        'Czechia',               '12/06/2026'),
    ('B', 'Canada',             'Bosnia and Herzegovina', '12/06/2026'),
    ('D', 'United States',      'Paraguay',              '13/06/2026'),
    ('B', 'Qatar',              'Switzerland',           '13/06/2026'),
    ('C', 'Brazil',             'Morocco',               '14/06/2026'),
    ('C', 'Haiti',              'Scotland',              '14/06/2026'),
    ('D', 'Australia',          'Turkey',                '14/06/2026'),
    ('E', 'Germany',            'Curaçao',               '14/06/2026'),
    ('F', 'Netherlands',        'Japan',                 '14/06/2026'),
    ('E', 'Ivory Coast',        'Ecuador',               '15/06/2026'),
    ('F', 'Sweden',             'Tunisia',               '15/06/2026'),
    ('H', 'Spain',              'Cape Verde',            '15/06/2026'),
    ('G', 'Belgium',            'Egypt',                 '15/06/2026'),
    ('H', 'Saudi Arabia',       'Uruguay',               '16/06/2026'),
    ('I', 'Iran',               'New Zealand',           '16/06/2026'),
    ('I', 'France',             'Senegal',               '16/06/2026'),
    ('J', 'Iraq',               'Norway',                '17/06/2026'),
    ('J', 'Argentina',          'Algeria',               '17/06/2026'),
    ('K', 'Austria',            'Jordan',                '17/06/2026'),
    ('K', 'Portugal',           'DR Congo',              '17/06/2026'),
    ('L', 'England',            'Croatia',               '17/06/2026'),
    ('L', 'Ghana',              'Panama',                '18/06/2026'),
    ('K', 'Uzbekistan',         'Colombia',              '18/06/2026'),
]


# ─── Core prediction ─────────────────────────────────────────────────────────
def get_elo_match(home, away):
    row = FIXTURES_ELO.get((home, away))
    if row:
        helo = row.get('home_elo') or ELO.get(home, {}).get('rating', 1700)
        aelo = row.get('away_elo') or ELO.get(away, {}).get('rating', 1700)
        wp_h = row.get('win_exp_home_pct', 50)
        wp_a = row.get('win_exp_away_pct', 50)
    else:
        helo = ELO.get(home, {}).get('rating', 1700)
        aelo = ELO.get(away, {}).get('rating', 1700)
        diff = helo - aelo + 100
        wp_h = 1 / (1 + 10 ** (-diff / 400)) * 100
        wp_a = 100 - wp_h
    return helo, aelo, float(wp_h), float(wp_a)


def expected_goals(home, away, form_h, form_a):
    """Hybrid: 35% Elo + 35% qualifying stats + 30% recent form (official-weighted)."""
    helo, aelo, _, _ = get_elo_match(home, away)

    # Branch 1: Elo
    elo_diff = helo - aelo + 100
    ef = 10 ** (elo_diff / 800)
    lh_elo = WC_AVG_GOALS * ef / (1 + ef)
    la_elo = WC_AVG_GOALS / (1 + ef)

    # Branch 2: Qualifying stats (calibrate per confederatie)
    hgpg = (per_game(home, 'goals') or 1.4) * qual_to_wc(home, 'goals')
    agpg = (per_game(away, 'goals') or 1.2) * qual_to_wc(away, 'goals')
    qt = hgpg + agpg or 2.6
    lh_qual = WC_AVG_GOALS * hgpg / qt
    la_qual = WC_AVG_GOALS * agpg / qt

    # Branch 3: Recent form
    rfh = form_h['lambda_recent']
    rfa = form_a['lambda_recent']
    rft = rfh + rfa or 2.4
    lh_rf = WC_AVG_GOALS * rfh / rft
    la_rf = WC_AVG_GOALS * rfa / rft

    # Opponent Elo adjustment: if recent opponents much weaker than current opponent, scale down
    opp_adj_h = max(0.7, min(1.3, 1.0 + (form_h['avg_opp_elo'] - aelo) / 3000))
    opp_adj_a = max(0.7, min(1.3, 1.0 + (form_a['avg_opp_elo'] - helo) / 3000))
    lh_rf *= opp_adj_h
    la_rf *= opp_adj_a

    lh = 0.35 * lh_elo + 0.35 * lh_qual + 0.30 * lh_rf
    la = 0.35 * la_elo + 0.35 * la_qual + 0.30 * la_rf
    return max(0.15, lh), max(0.15, la)


def expected_shots(team, opp_elo, is_home, form):
    q = qual_to_wc(team, 'shots')
    base = per_game(team, 'total_shots')
    base = (base or WC_AVG_SHOTS_TEAM / q) * q
    my_elo = ELO.get(team, {}).get('rating', 1700)
    elo_adj = 1 + (my_elo - opp_elo) / 3000
    ha_adj = 1.05 if is_home else 0.95
    form_adj = 1.0 + (form.get('win_rate', 0.4) - 0.4) * 0.15
    return max(4.0, min(22.0, base * elo_adj * ha_adj * form_adj))


def expected_sot(team, total_shots):
    total = _num(AISCORE.get(team, {}).get('total_shots'))
    sot   = _num(AISCORE.get(team, {}).get('shots_on_target'))
    if total and sot and total > 0:
        return total_shots * (sot / total)
    return total_shots * 0.35


def expected_corners(team, shots):
    # ~0.37 cornere/sut calibrat pe media WC2022 (~10 cornere/meci cu ~13 suturi/echipa)
    return max(3.0, min(7.5, shots * 0.37))


def expected_yellows(team, opp_team, form):
    my_yc = per_game(team, 'yellow_cards')
    my_yc = (my_yc or 1.6) * qual_to_wc(team, 'cards')
    my_elo  = ELO.get(team, {}).get('rating', 1700)
    opp_elo = ELO.get(opp_team, {}).get('rating', 1700)
    rivalry = 1 + max(0, (300 - abs(my_elo - opp_elo))) / 600
    form_adj = 1.0 + (form.get('wcpg', 1.2) - 1.2) * 0.05
    return my_yc * rivalry * max(0.85, min(1.2, form_adj))


def expected_fouls(team, opp_team):
    fouls = per_game(team, 'fouls')
    fouls = (fouls or 12) * qual_to_wc(team, 'fouls')
    my_elo  = ELO.get(team, {}).get('rating', 1700)
    opp_elo = ELO.get(opp_team, {}).get('rating', 1700)
    underdog = 1 + max(0, opp_elo - my_elo) / 2000
    return fouls * underdog


def expected_offsides(team):
    kp = per_game(team, 'key_passes') or 12
    return max(0.5, kp / 12 * 2.0)


def top_scorers(team, lambda_goals, n=3):
    pg = PLAYERS.get(team, {}).get('goals', [])
    if not pg: return []
    pg = sorted(pg, key=lambda x: -(x[1] or 0))[:6]
    total = sum(v for _, v in pg) or 1
    result = []
    for pname, pgoals in pg[:n]:
        share = (pgoals or 0) / total
        p_score = 1 - math.exp(-lambda_goals * share)
        result.append((pname, round(p_score * 100, 1)))
    return result


def player_yc_probs(team, team_yc_pg):
    ycp = sorted(PLAYERS.get(team, {}).get('yellow_cards', []), key=lambda x: -(x[1] or 0))
    if not ycp: return []
    total = sum(v for _, v in ycp) or 1
    result = []
    for pname, pyc in ycp[:2]:
        p = (pyc / total) * team_yc_pg * 0.9
        result.append((pname, round(min(45, p * 100), 1)))
    return result


def confidence_score(home, away, market, elo_diff):
    h_data = bool(AISCORE.get(home))
    a_data = bool(AISCORE.get(away))
    h_form = len(RECENT_BY_TEAM.get(home, []))
    a_form = len(RECENT_BY_TEAM.get(away, []))
    helo = ELO.get(home, {}).get('rating', 1700)
    aelo = ELO.get(away, {}).get('rating', 1700)
    # Official match ratio
    hm = RECENT_BY_TEAM.get(home, [])[-10:]
    am = RECENT_BY_TEAM.get(away, [])[-10:]
    off_ratio_h = sum(1 for m in hm if is_official(m['tournament'])) / max(1, len(hm))
    off_ratio_a = sum(1 for m in am if is_official(m['tournament'])) / max(1, len(am))
    avg_off_ratio = (off_ratio_h + off_ratio_a) / 2

    base = {
        'result':   78, 'goals_ou': 63, 'gg':        58,
        'shots':    58, 'sot':      56, 'corners':   50,
        'yellows':  53, 'fouls':    52, 'offsides':  46,
        'scorers':  36, 'specials': 28,
        'dc':       72, 'dnb':      70, 'asian_hcp': 65,
        'ht':       52, 'win_nil':  48, 'correct':   22,
        'team_ou':  58, 'combo':    50,
    }.get(market, 50)

    if elo_diff > 300: base += 10
    elif elo_diff > 200: base += 7
    elif elo_diff > 100: base += 4
    elif elo_diff < 30: base -= 6

    if not (h_data and a_data): base -= 10
    elif not h_data or not a_data: base -= 5

    if h_form >= 8 and a_form >= 8:
        base += 4
    elif h_form < 3 or a_form < 3:
        base -= 5

    # More official matches in recent form = higher confidence
    if avg_off_ratio >= 0.7: base += 3
    elif avg_off_ratio < 0.4: base -= 3

    if min(helo, aelo) < 1500: base -= 6
    elif min(helo, aelo) < 1600: base -= 3

    return max(18, min(88, base))


# ─── Predict ─────────────────────────────────────────────────────────────────
def predict_match(group, home, away, date):
    helo, aelo, wp_h, wp_a = get_elo_match(home, away)
    elo_diff = abs(helo - aelo)

    form_h = recent_form(home, 10)
    form_a = recent_form(away, 10)

    lambda_h, lambda_a = expected_goals(home, away, form_h, form_a)

    shots_h = expected_shots(home, aelo, True,  form_h)
    shots_a = expected_shots(away, helo, False, form_a)
    sot_h   = expected_sot(home, shots_h)
    sot_a   = expected_sot(away, shots_a)
    corn_h  = expected_corners(home, shots_h)
    corn_a  = expected_corners(away, shots_a)
    yc_h    = expected_yellows(home, away, form_h)
    yc_a    = expected_yellows(away, home, form_a)
    fouls_h = expected_fouls(home, away)
    fouls_a = expected_fouls(away, home)
    offs_h  = expected_offsides(home)
    offs_a  = expected_offsides(away)

    total_goals = lambda_h + lambda_a
    total_sot   = sot_h + sot_a
    total_corn  = corn_h + corn_a
    total_yc    = yc_h + yc_a
    total_fouls = fouls_h + fouls_a
    total_offs  = offs_h + offs_a

    p_o15 = p_over_half(1, total_goals)
    p_o25 = p_over_half(2, total_goals)
    p_o35 = p_over_half(3, total_goals)
    p_gg  = (1 - math.exp(-lambda_h)) * (1 - math.exp(-lambda_a))
    p_draw = max(8.0, min(35.0, round(sum(ppf(n, lambda_h) * ppf(n, lambda_a) for n in range(6)) * 100, 1)))

    top_scores = sorted(
        {(h, a): ppf(h, lambda_h) * ppf(a, lambda_a) for h in range(6) for a in range(6)}.items(),
        key=lambda x: -x[1])[:4]

    def conf(m): return confidence_score(home, away, m, elo_diff)

    # ── Bets list ────────────────────────────────────────────────────────────
    def ou(thresh, prob_pct, label, conf_key):
        if prob_pct >= 50:
            return (label, f"Over {thresh}", round(prob_pct, 1), conf(conf_key))
        else:
            return (label, f"Under {thresh}", round(100 - prob_pct, 1), conf(conf_key))

    bets = []

    # ── 1. Rezultat 1X2 ──────────────────────────────────────────────────────
    if wp_h >= p_draw and wp_h >= wp_a:   res, rp = f"1 ({home})", wp_h
    elif wp_a >= p_draw and wp_a >= wp_h: res, rp = f"2 ({away})", wp_a
    else:                                  res, rp = "X (egal)", p_draw
    bets.append(('Rezultat 1/X/2', res, round(rp, 1), conf('result')))

    # ── 2. Sansa dubla (Double Chance) ───────────────────────────────────────
    p_1x = round(wp_h + p_draw, 1)
    p_x2 = round(wp_a + p_draw, 1)
    p_12 = round(wp_h + wp_a, 1)
    bets += [
        ('Sansa Dubla 1X', '1X (Acasa sau Egal)', p_1x, conf('result')),
        ('Sansa Dubla X2', 'X2 (Egal sau Deplasare)', p_x2, conf('result')),
        ('Sansa Dubla 12', '12 (Orice victorie)', p_12, conf('result')),
    ]

    # ── 3. Draw No Bet ───────────────────────────────────────────────────────
    p_dnb_h = round(wp_h / max(1, wp_h + wp_a) * 100, 1)
    p_dnb_a = round(wp_a / max(1, wp_h + wp_a) * 100, 1)
    bets += [
        (f'Draw No Bet — {home[:15]}', 'Victorie (fara egal)', p_dnb_h, conf('result')),
        (f'Draw No Bet — {away[:15]}', 'Victorie (fara egal)', p_dnb_a, conf('result')),
    ]

    # ── 4. Goluri Over/Under ─────────────────────────────────────────────────
    bets += [
        ou(0.5, (1 - ppf(0, total_goals))*100, 'Goluri Over/Under 0.5', 'goals_ou'),
        ou(1.5, p_o15*100, 'Goluri Over/Under 1.5', 'goals_ou'),
        ou(2.5, p_o25*100, 'Goluri Over/Under 2.5', 'goals_ou'),
        ou(3.5, p_o35*100, 'Goluri Over/Under 3.5', 'goals_ou'),
        ou(4.5, p_over_half(4, total_goals)*100, 'Goluri Over/Under 4.5', 'goals_ou'),
        ou(5.5, p_over_half(5, total_goals)*100, 'Goluri Over/Under 5.5', 'goals_ou'),
    ]

    # ── 5. Ambele marcheaza (GG / BTTS) ──────────────────────────────────────
    gg_pred = 'Da (GG)' if p_gg*100 >= 50 else 'Nu (NU GG)'
    gg_prob = p_gg*100 if p_gg*100 >= 50 else (1-p_gg)*100
    bets.append(('Ambele marcheaza (GG)', gg_pred, round(gg_prob, 1), conf('gg')))

    # ── 6. GG + Over/Under (piete combinate) ────────────────────────────────
    p_gg_o25 = p_gg * p_o25  # aproximare (corelate, dar simplu)
    p_gg_u25 = p_gg * (1 - p_o25)
    p_ng_o25 = (1 - p_gg) * p_o25
    bets += [
        ('GG + Over 2.5',   'Da',  round(p_gg_o25 * 100, 1), conf('gg')),
        ('GG + Under 2.5',  'Da',  round(p_gg_u25 * 100, 1), conf('gg')),
        ('NU GG + Over 2.5','Da',  round(p_ng_o25 * 100, 1), conf('gg')),
    ]

    # ── 7. Goluri echipe individuale ─────────────────────────────────────────
    p_h_o05 = round((1 - math.exp(-lambda_h)) * 100, 1)
    p_h_o15 = round(p_over_half(1, lambda_h) * 100, 1)
    p_h_o25 = round(p_over_half(2, lambda_h) * 100, 1)
    p_a_o05 = round((1 - math.exp(-lambda_a)) * 100, 1)
    p_a_o15 = round(p_over_half(1, lambda_a) * 100, 1)
    p_a_o25 = round(p_over_half(2, lambda_a) * 100, 1)
    bets += [
        (f'{home[:16]} Over/Under 0.5', f"{'Over' if p_h_o05>=50 else 'Under'} 0.5",
         max(p_h_o05, 100-p_h_o05), conf('goals_ou')),
        (f'{home[:16]} Over/Under 1.5', f"{'Over' if p_h_o15>=50 else 'Under'} 1.5",
         max(p_h_o15, 100-p_h_o15), conf('goals_ou')),
        (f'{home[:16]} Over/Under 2.5', f"{'Over' if p_h_o25>=50 else 'Under'} 2.5",
         max(p_h_o25, 100-p_h_o25), conf('goals_ou')),
        (f'{away[:16]} Over/Under 0.5', f"{'Over' if p_a_o05>=50 else 'Under'} 0.5",
         max(p_a_o05, 100-p_a_o05), conf('goals_ou')),
        (f'{away[:16]} Over/Under 1.5', f"{'Over' if p_a_o15>=50 else 'Under'} 1.5",
         max(p_a_o15, 100-p_a_o15), conf('goals_ou')),
        (f'{away[:16]} Over/Under 2.5', f"{'Over' if p_a_o25>=50 else 'Under'} 2.5",
         max(p_a_o25, 100-p_a_o25), conf('goals_ou')),
    ]

    # ── 8. Handicap Asiatic ───────────────────────────────────────────────────
    # AH -0.5 (echipa acasa castiga orice) = prob victorie acasa
    # AH -1   (echipa acasa castiga cu 2+) = P(home goals - away goals >= 2)
    # AH +0.5 (echipa deplasare nu pierde)
    p_ah_h05 = round(wp_h, 1)
    p_ah_h1  = round(sum(ppf(i, lambda_h) * ppf(j, lambda_a)
                         for i in range(9) for j in range(9)
                         if i - j >= 2) * 100, 1)
    p_ah_h15 = round(sum(ppf(i, lambda_h) * ppf(j, lambda_a)
                         for i in range(9) for j in range(9)
                         if i - j >= 3) * 100, 1)
    p_ah_a05 = round(100 - wp_h, 1)
    bets += [
        (f'Handicap Asiatic {home[:14]} -0.5', 'Victorie', p_ah_h05, conf('result')),
        (f'Handicap Asiatic {home[:14]} -1',   'Victorie cu 2+', p_ah_h1, conf('result')),
        (f'Handicap Asiatic {home[:14]} -1.5', 'Victorie cu 3+', p_ah_h15, conf('result')),
        (f'Handicap Asiatic {away[:14]} +0.5', 'Fara infrangere', p_ah_a05, conf('result')),
    ]

    # ── 9. Prima repriza ─────────────────────────────────────────────────────
    # Aprox: ~42% din goluri se marcheaza in prima repriza
    lh1, la1 = lambda_h * 0.42, lambda_a * 0.42
    total_1h  = lh1 + la1
    p_1h_o05  = round((1 - pcdf(0, total_1h)) * 100, 1)
    p_1h_o15  = round(p_over_half(1, total_1h) * 100, 1)
    p_1h_gg   = round((1 - math.exp(-lh1)) * (1 - math.exp(-la1)) * 100, 1)
    ph1, pd1, pa1 = (sum(ppf(i, lh1) * ppf(j, la1)
                         for i in range(6) for j in range(6)
                         if (i>j if k==0 else i==j if k==1 else i<j))
                     for k in range(3))
    s1 = ph1 + pd1 + pa1
    ph1, pd1, pa1 = ph1/s1*100, pd1/s1*100, pa1/s1*100
    r1h = max(ph1, pd1, pa1)
    r1_pred = (f"1 ({home[:10]})" if ph1==r1h else
               "X (egal)" if pd1==r1h else f"2 ({away[:10]})")
    bets += [
        ('Rezultat Repriza 1', r1_pred, round(r1h, 1), conf('result') - 8),
        ('Goluri R1 Over/Under 0.5', f"{'Over' if p_1h_o05>=50 else 'Under'} 0.5",
         max(p_1h_o05, 100-p_1h_o05), conf('goals_ou') - 5),
        ('Goluri R1 Over/Under 1.5', f"{'Over' if p_1h_o15>=50 else 'Under'} 1.5",
         max(p_1h_o15, 100-p_1h_o15), conf('goals_ou') - 5),
        ('GG Repriza 1', f"{'Da' if p_1h_gg>=50 else 'Nu'} GG",
         max(p_1h_gg, 100-p_1h_gg), conf('gg') - 8),
    ]

    # ── 10. Win to Nil ────────────────────────────────────────────────────────
    p_h_win_nil = round(sum(ppf(i, lambda_h) * ppf(0, lambda_a)
                            for i in range(1, 9)) * 100, 1)
    p_a_win_nil = round(sum(ppf(0, lambda_h) * ppf(j, lambda_a)
                            for j in range(1, 9)) * 100, 1)
    bets += [
        (f'{home[:16]} Win to Nil', 'Da', p_h_win_nil, conf('result') - 5),
        (f'{away[:16]} Win to Nil', 'Da', p_a_win_nil, conf('result') - 5),
    ]

    # ── 11. Scor exact (Correct Score) ───────────────────────────────────────
    all_scores = sorted(
        {(i, j): ppf(i, lambda_h) * ppf(j, lambda_a)
         for i in range(6) for j in range(6)}.items(),
        key=lambda x: -x[1])
    for (hi, ai), sp in all_scores[:8]:
        bets.append((f'Scor Exact {hi}-{ai}', f"{hi}-{ai}",
                     round(sp * 100, 1), conf('result') - 12))

    # ── 12. SOT ──────────────────────────────────────────────────────────────
    bets += [
        ou(7.5, p_over_half(7, total_sot)*100, 'SOT Over/Under 7.5', 'sot'),
        ou(8.5, p_over_half(8, total_sot)*100, 'SOT Over/Under 8.5', 'sot'),
        ou(9.5, p_over_half(9, total_sot)*100, 'SOT Over/Under 9.5', 'sot'),
    ]

    # ── 13. Cornere ──────────────────────────────────────────────────────────
    bets += [
        ou(8.5,  p_over_half(8,  total_corn)*100, 'Cornere Over/Under 8.5',  'corners'),
        ou(9.5,  p_over_half(9,  total_corn)*100, 'Cornere Over/Under 9.5',  'corners'),
        ou(10.5, p_over_half(10, total_corn)*100, 'Cornere Over/Under 10.5', 'corners'),
        ou(11.5, p_over_half(11, total_corn)*100, 'Cornere Over/Under 11.5', 'corners'),
        ou(12.5, p_over_half(12, total_corn)*100, 'Cornere Over/Under 12.5', 'corners'),
    ]

    # ── 14. Cartonase galbene ────────────────────────────────────────────────
    bets += [
        ou(1.5, p_over_half(1, total_yc)*100, 'G. Galbene Over/Under 1.5', 'yellows'),
        ou(2.5, p_over_half(2, total_yc)*100, 'G. Galbene Over/Under 2.5', 'yellows'),
        ou(3.5, p_over_half(3, total_yc)*100, 'G. Galbene Over/Under 3.5', 'yellows'),
        ou(4.5, p_over_half(4, total_yc)*100, 'G. Galbene Over/Under 4.5', 'yellows'),
        ou(5.5, p_over_half(5, total_yc)*100, 'G. Galbene Over/Under 5.5', 'yellows'),
    ]

    # ── 15. Faulturi ─────────────────────────────────────────────────────────
    bets += [
        ou(20.5, p_over_half(20, total_fouls)*100, 'Faulturi Over/Under 20.5', 'fouls'),
        ou(22.5, p_over_half(22, total_fouls)*100, 'Faulturi Over/Under 22.5', 'fouls'),
        ou(25.5, p_over_half(25, total_fouls)*100, 'Faulturi Over/Under 25.5', 'fouls'),
        ou(28.5, p_over_half(28, total_fouls)*100, 'Faulturi Over/Under 28.5', 'fouls'),
        ou(30.5, p_over_half(30, total_fouls)*100, 'Faulturi Over/Under 30.5', 'fouls'),
    ]

    # ── 16. Ofsaiduri ────────────────────────────────────────────────────────
    bets += [
        ou(1.5, p_over_half(1, total_offs)*100, 'Ofsaid Over/Under 1.5', 'offsides'),
        ou(2.5, p_over_half(2, total_offs)*100, 'Ofsaid Over/Under 2.5', 'offsides'),
        ou(3.5, p_over_half(3, total_offs)*100, 'Ofsaid Over/Under 3.5', 'offsides'),
        ou(4.5, p_over_half(4, total_offs)*100, 'Ofsaid Over/Under 4.5', 'offsides'),
    ]

    # ── 17. Speciale ─────────────────────────────────────────────────────────
    p_penalty = 0.20 + 0.05 * (1 - abs(wp_h - 50) / 50)
    p_red     = 0.08 + 0.03 * (1 - abs(wp_h - 50) / 50)
    p_header  = 1 - math.exp(-(lambda_h + lambda_a) * 0.20)

    # Penalty + rosu per echipa (aprox)
    p_pen_h = p_penalty * wp_a / max(1, wp_h + wp_a) * 100
    p_pen_a = p_penalty * wp_h / max(1, wp_h + wp_a) * 100

    bets += [
        ('Penalty in meci', 'Da' if p_penalty >= 0.5 else 'Nu',
         round(max(p_penalty, 1-p_penalty)*100, 1), conf('specials')),
        (f'Penalty pentru {home[:14]}', 'Da', round(p_pen_h, 1), conf('specials') - 5),
        (f'Penalty pentru {away[:14]}', 'Da', round(p_pen_a, 1), conf('specials') - 5),
        ('Gol cu capul', 'Da' if p_header >= 0.5 else 'Nu',
         round(max(p_header, 1-p_header)*100, 1), conf('specials')),
        ('Rosu in meci', 'Nu' if p_red < 0.5 else 'Da',
         round((1-p_red)*100 if p_red < 0.5 else p_red*100, 1), conf('specials')),
        ('Meci se termina la egalitate (X)', f"{'Da' if p_draw>=30 else 'Nu'} ({round(p_draw,1)}%)",
         max(p_draw, 100-p_draw), conf('result') - 3),
    ]

    # ── 18. Marcatori & cartonase jucatori ───────────────────────────────────
    scorers_h = top_scorers(home, lambda_h, 3)
    scorers_a = top_scorers(away, lambda_a, 3)
    yc_pl_h   = player_yc_probs(home, yc_h)
    yc_pl_a   = player_yc_probs(away, yc_a)

    for pname, pprob in scorers_h:
        bets.append((f'Marcator: {pname} ({home[:12]})', 'Marchează', pprob, conf('scorers')))
    for pname, pprob in scorers_a:
        bets.append((f'Marcator: {pname} ({away[:12]})', 'Marchează', pprob, conf('scorers')))
    for pname, pprob in yc_pl_h:
        bets.append((f'G.Galben: {pname} ({home[:10]})', 'Galben Da', pprob, conf('specials')))
    for pname, pprob in yc_pl_a:
        bets.append((f'G.Galben: {pname} ({away[:10]})', 'Galben Da', pprob, conf('specials')))

    bets_sorted = sorted(bets, key=lambda x: (-x[3], -(x[2] or 0)))

    return {
        'group': group, 'home': home, 'away': away, 'date': date,
        'helo': helo, 'aelo': aelo, 'elo_diff': elo_diff,
        'wp_home': round(wp_h, 1), 'wp_away': round(wp_a, 1), 'p_draw': p_draw,
        'lambda_h': round(lambda_h, 2), 'lambda_a': round(lambda_a, 2),
        'total_goals': round(total_goals, 2),
        'p_o15': round(p_o15*100,1), 'p_o25': round(p_o25*100,1), 'p_o35': round(p_o35*100,1),
        'p_gg': round(p_gg*100,1), 'p_btts_no': round((1-p_gg)*100,1),
        'top_scores': [f"{h}-{a} ({round(p*100,1)}%)" for (h,a),p in top_scores],
        'shots_h': round(shots_h,1), 'shots_a': round(shots_a,1),
        'total_shots': round(shots_h+shots_a,1),
        'sot_h': round(sot_h,1), 'sot_a': round(sot_a,1), 'total_sot': round(total_sot,1),
        'corn_h': round(corn_h,1), 'corn_a': round(corn_a,1), 'total_corn': round(total_corn,1),
        'yc_h': round(yc_h,2), 'yc_a': round(yc_a,2), 'total_yc': round(total_yc,2),
        'fouls_h': round(fouls_h,1), 'fouls_a': round(fouls_a,1), 'total_fouls': round(total_fouls,1),
        'offs_h': round(offs_h,1), 'offs_a': round(offs_a,1), 'total_offs': round(total_offs,1),
        'scorers_h': scorers_h, 'scorers_a': scorers_a,
        'yc_pl_h': yc_pl_h, 'yc_pl_a': yc_pl_a,
        'p_penalty': round(p_penalty*100,1), 'p_header': round(p_header*100,1),
        'p_red': round(p_red*100,1),
        'p_1x': round(p_1x, 1), 'p_x2': round(p_x2, 1), 'p_12': round(p_12, 1),
        'p_dnb_h': round(p_dnb_h, 1), 'p_dnb_a': round(p_dnb_a, 1),
        'p_ah_h1': round(p_ah_h1, 1), 'p_ah_h15': round(p_ah_h15, 1),
        'p_h_o15': round(p_h_o15, 1), 'p_a_o15': round(p_a_o15, 1),
        'p_h_win_nil': round(p_h_win_nil, 1), 'p_a_win_nil': round(p_a_win_nil, 1),
        'p_1h_o05': round(p_1h_o05, 1), 'p_1h_o15': round(p_1h_o15, 1),
        'form_h': form_h, 'form_a': form_a,
        'all_bets': bets_sorted,
        'conf': {k: confidence_score(home, away, k, elo_diff)
                 for k in ['result','goals_ou','gg','shots','sot','corners',
                            'yellows','fouls','offsides','scorers','specials',
                            'dc','dnb','asian_hcp','ht','win_nil','correct',
                            'team_ou','combo']},
        'max_conf': max(confidence_score(home, away, k, elo_diff)
                        for k in ['result','goals_ou','gg']),
    }


# ─── Excel builder ────────────────────────────────────────────────────────────
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'),  bottom=Side(style='thin'))

def _fill(h): return PatternFill('solid', fgColor=h)

F = {
    'title':   _fill('1A1A2E'), 'elo':     _fill('0F3460'),
    'goals':   _fill('1A5C2A'), 'shots':   _fill('5C1A1A'),
    'corners': _fill('5C4A1A'), 'cards':   _fill('4A1A5C'),
    'fouls':   _fill('1A4A5C'), 'offs':    _fill('1A5C5C'),
    'scorers': _fill('3D5C1A'), 'spec':    _fill('5C3D1A'),
    'row0':    _fill('F2F7FF'), 'row1':    _fill('FFFFFF'),
    'form':    _fill('2C3E50'),
}

def pf_c(p):
    if p >= 75: return _fill('00B050')
    elif p >= 65: return _fill('92D050')
    elif p >= 55: return _fill('FFEB9C')
    elif p >= 45: return _fill('FFC000')
    else: return _fill('FF7878')

def pf_p(p):
    p = p or 0
    if p >= 80: return _fill('00B050')
    elif p >= 65: return _fill('92D050')
    elif p >= 50: return _fill('FFEB9C')
    elif p >= 35: return _fill('FFC000')
    else: return _fill('FF7878')


def build_excel(preds):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ────────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Sumar_Etapa1'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:AE1')
    c = ws['A1']
    c.value = 'WC2026 — PREDICTII RUNDA 1  |  Model: Poisson + Elo + Forma Recenta 10 meciuri (ponderat oficial vs amical)'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = F['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    hdrs = ['Gr','Data','Acasa','Deplasare',
            'Elo\nAcasa','Elo\nDepl.','Win%\nAcasa','Win%\nEgal','Win%\nDepl.',
            'xG\nAcasa','xG\nDepl.','Total\nxG',
            'Ov1.5\n%','Ov2.5\n%','Ov3.5\n%','GG\n%',
            'Suturi\nAcasa','Suturi\nDepl.',
            'SOT\nAcasa','SOT\nDepl.',
            'Corn.\nAcasa','Corn.\nDepl.',
            'Galb.\nAcasa','Galb.\nDepl.',
            'Faulturi\nTotal',
            'Conf\nRez.','Conf\nGol.',
            'Conf\nSOT','Conf\nCorn.','Conf\nGalb.','Conf\nFaulturi']
    cfills = ([F['title']]*4 + [F['elo']]*5 + [F['goals']]*7 + [F['shots']]*4 +
              [F['corners']]*2 + [F['cards']]*2 + [F['fouls']] + [F['title']]*2 +
              [F['shots'], F['corners'], F['cards'], F['fouls']])
    for ci, (h, f) in enumerate(zip(hdrs, cfills), 1):
        c = ws.cell(2, ci, h)
        c.fill = f; c.font = Font(bold=True, color='FFFFFF', size=8)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN
    ws.row_dimensions[2].height = 34

    for ci, w in enumerate([4,11,22,22,7,7,8,8,8,7,7,7,9,9,9,9,8,8,7,7,7,7,7,7,8,9,9,9,9,9,9], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, p in enumerate(preds, 3):
        bg = F['row0'] if ri % 2 == 1 else F['row1']
        vals = [p['group'], p['date'], p['home'], p['away'],
                p['helo'], p['aelo'],
                f"{p['wp_home']}%", f"{p['p_draw']}%", f"{p['wp_away']}%",
                p['lambda_h'], p['lambda_a'], p['total_goals'],
                f"{p['p_o15']}%", f"{p['p_o25']}%", f"{p['p_o35']}%", f"{p['p_gg']}%",
                p['shots_h'], p['shots_a'], p['sot_h'], p['sot_a'],
                p['corn_h'], p['corn_a'], p['yc_h'], p['yc_a'], p['total_fouls'],
                f"{p['conf']['result']}%", f"{p['conf']['goals_ou']}%",
                f"{p['conf']['sot']}%", f"{p['conf']['corners']}%",
                f"{p['conf']['yellows']}%", f"{p['conf']['fouls']}%"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal='center' if ci not in (3,4) else 'left', vertical='center')
            c.border = THIN; c.fill = bg
        ws.cell(ri, 7).fill = pf_p(p['wp_home'])
        ws.cell(ri, 9).fill = pf_p(p['wp_away'])
        ws.cell(ri,14).fill = pf_p(p['p_o25'])
        ws.cell(ri,16).fill = pf_p(p['p_gg'])
        ws.cell(ri,26).fill = pf_c(p['conf']['result'])
        ws.cell(ri,27).fill = pf_c(p['conf']['goals_ou'])
        ws.cell(ri,28).fill = pf_c(p['conf']['sot'])
        ws.cell(ri,29).fill = pf_c(p['conf']['corners'])
        ws.cell(ri,30).fill = pf_c(p['conf']['yellows'])
        ws.cell(ri,31).fill = pf_c(p['conf']['fouls'])
        ws.row_dimensions[ri].height = 16

    # ── Sheet 2: Detailed ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Predictii_Detaliate')
    ws2.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [32, 22, 22, 12, 12, 12]):
        ws2.column_dimensions[col].width = w

    cur = [1]

    def tr(text, fk='title'):
        ws2.merge_cells(start_row=cur[0], start_column=1, end_row=cur[0], end_column=6)
        c = ws2.cell(cur[0], 1, text)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = F[fk]; c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[cur[0]].height = 20
        cur[0] += 1

    def hr(cols, fk):
        for ci, txt in enumerate(cols, 1):
            c = ws2.cell(cur[0], ci, txt)
            c.font = Font(bold=True, color='FFFFFF', size=8)
            c.fill = F[fk]; c.border = THIN
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2.row_dimensions[cur[0]].height = 24; cur[0] += 1

    def dr(vals, pcs=(), cc=None):
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(cur[0], ci, v)
            c.font = Font(size=9)
            c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left', vertical='center')
        for pc in pcs:
            try: ws2.cell(cur[0], pc).fill = pf_p(float(str(vals[pc-1]).replace('%','')))
            except: pass
        if cc is not None:
            try: ws2.cell(cur[0], cc).fill = pf_c(int(str(vals[cc-1]).replace('%','')))
            except: pass
        ws2.row_dimensions[cur[0]].height = 14; cur[0] += 1

    for p in preds:
        tr(f"Gr.{p['group']} | {p['date']} | {p['home']} vs {p['away']}  (Elo {p['helo']}-{p['aelo']}, D{p['elo_diff']})")

        fh, fa = p['form_h'], p['form_a']
        hr(['Forma Recenta (10 meciuri)', 'Acasa', 'Deplasare', 'Info', '', ''], 'form')
        def fmt_rec(form):
            ms = form.get('matches_summary', [])
            return ' | '.join(f"{m[1][:8]} {m[2]}-{m[3]}" for m in ms[-3:]) if ms else 'N/A'
        dr(['Meciuri analizate',           fh['n_total'], fa['n_total'], f"din {10} max",'',''])
        dr(['Meciuri oficiale (ponderate)', fh['n_official'], fa['n_official'],'','',''])
        dr(['Goluri marcate/meci (pond.)',  round(fh['wpg'],2), round(fa['wpg'],2),'','',''])
        dr(['Goluri primite/meci (pond.)',  round(fh['wcpg'],2), round(fa['wcpg'],2),'','',''])
        dr(['Avg Elo adversari (ult. 10)',  round(fh['avg_opp_elo']), round(fa['avg_opp_elo']),'calitate adv.','',''])
        dr(['Rata victorii (pond.)',        f"{round(fh['win_rate']*100,1)}%", f"{round(fa['win_rate']*100,1)}%",'','',''])
        dr(['Ultimele 3 meciuri',           fmt_rec(fh), fmt_rec(fa),'','',''])

        # ─ Rezultat + Sansa Dubla + DNB + Handicap ──────────────────────────
        hr(['Piata', 'Acasa', 'Egal', 'Deplasare', 'Prob.%', 'Incredere'], 'elo')
        dr(['Rezultat 1/X/2', f"{p['wp_home']}%", f"{p['p_draw']}%", f"{p['wp_away']}%",
            f"Max {max(p['wp_home'],p['p_draw'],p['wp_away'])}%", f"{p['conf']['result']}%"],
           pcs=(2,3,4,5), cc=6)
        dr(['Sansa Dubla 1X', f"{p['p_1x']}%",'','', f"{p['p_1x']}%", f"{p['conf']['dc']}%"],
           pcs=(5,), cc=6)
        dr(['Sansa Dubla X2', '','',f"{p['p_x2']}%", f"{p['p_x2']}%", f"{p['conf']['dc']}%"],
           pcs=(5,), cc=6)
        dr(['Sansa Dubla 12', f"{p['wp_home']}%",'',f"{p['wp_away']}%", f"{p['p_12']}%", f"{p['conf']['dc']}%"],
           pcs=(5,), cc=6)
        dr([f"Draw No Bet — {p['home'][:16]}", f"{p['p_dnb_h']}%",'','',
            f"{p['p_dnb_h']}%", f"{p['conf']['dnb']}%"], pcs=(5,), cc=6)
        dr([f"Draw No Bet — {p['away'][:16]}", '','',f"{p['p_dnb_a']}%",
            f"{p['p_dnb_a']}%", f"{p['conf']['dnb']}%"], pcs=(5,), cc=6)
        dr([f"Handicap Asiatic {p['home'][:12]} -1 (castig 2+)",'','','',
            f"{p['p_ah_h1']}%", f"{p['conf']['asian_hcp']}%"], pcs=(5,), cc=6)
        dr([f"Handicap Asiatic {p['home'][:12]} -1.5 (castig 3+)",'','','',
            f"{p['p_ah_h15']}%", f"{p['conf']['asian_hcp']}%"], pcs=(5,), cc=6)

        # ─ Goluri ────────────────────────────────────────────────────────────
        hr(['Goluri', 'Acasa', 'Deplasare', 'Total', 'Prob.%', 'Incredere'], 'goals')
        for row in [
            ['xG estimat', p['lambda_h'], p['lambda_a'], p['total_goals'],
             f"~{p['total_goals']}", f"{p['conf']['goals_ou']}%"],
            ['Over 0.5','','','', f"{round((1-ppf(0,p['total_goals']))*100,1)}%", f"{p['conf']['goals_ou']}%"],
            ['Over 1.5','','','', f"{p['p_o15']}%", f"{p['conf']['goals_ou']}%"],
            ['Over 2.5','','','', f"{p['p_o25']}%", f"{p['conf']['goals_ou']}%"],
            ['Over 3.5','','','', f"{p['p_o35']}%", f"{p['conf']['goals_ou']}%"],
            ['Over 4.5','','','', f"{round(p_over_half(4,p['total_goals'])*100,1)}%", f"{p['conf']['goals_ou']}%"],
            ['Over 5.5','','','', f"{round(p_over_half(5,p['total_goals'])*100,1)}%", f"{p['conf']['goals_ou']}%"],
            ['GG (ambele marcheaza)','','','', f"{p['p_gg']}%", f"{p['conf']['gg']}%"],
            ['NU GG','','','', f"{p['p_btts_no']}%", f"{p['conf']['gg']}%"],
            ['GG + Over 2.5','','','',
             f"{round(p['p_gg']/100*p['p_o25']/100*100,1)}%", f"{p['conf']['combo']}%"],
            [f"{p['home'][:14]} Over 1.5", f"{p['p_h_o15']}%",'','',
             f"{p['p_h_o15']}%", f"{p['conf']['team_ou']}%"],
            [f"{p['away'][:14]} Over 1.5",'','',f"{p['p_a_o15']}%",
             f"{p['p_a_o15']}%", f"{p['conf']['team_ou']}%"],
        ] + [['Scor Exact #'+str(i+1),'','','',s,''] for i,s in enumerate(p['top_scores'])]:
            dr(row, pcs=(5,), cc=6)

        # ─ Prima Repriza ─────────────────────────────────────────────────────
        hr(['Prima Repriza', 'Acasa', 'Egal', 'Deplasare', 'Prob.%', 'Incredere'], 'form')
        dr(['Rezultat Repriza 1 (est.)',
            f"~{round(p['wp_home']*0.85,1)}%", f"~{round(p['p_draw']*1.3,1)}%",
            f"~{round(p['wp_away']*0.85,1)}%",'', f"{p['conf']['ht']}%"])
        dr(['Over 0.5 Goluri R1','','','', f"{p['p_1h_o05']}%", f"{p['conf']['ht']}%"],
           pcs=(5,), cc=6)
        dr(['Over 1.5 Goluri R1','','','', f"{p['p_1h_o15']}%", f"{p['conf']['ht']}%"],
           pcs=(5,), cc=6)

        # ─ Win to Nil ────────────────────────────────────────────────────────
        hr(['Win to Nil / Clean Sheet', 'Prob.%', 'Incredere','','',''], 'elo')
        dr([f"{p['home'][:20]} Win to Nil", f"{p['p_h_win_nil']}%",
            f"{p['conf']['win_nil']}%",'','',''], pcs=(2,), cc=3)
        dr([f"{p['away'][:20]} Win to Nil", f"{p['p_a_win_nil']}%",
            f"{p['conf']['win_nil']}%",'','',''], pcs=(2,), cc=3)

        # ─ Suturi / SOT ──────────────────────────────────────────────────────
        hr(['Suturi / SOT', 'Acasa', 'Deplasare', 'Total', 'Prob.%', 'Incredere'], 'shots')
        for row in [
            ['Suturi total (est.)', p['shots_h'], p['shots_a'], p['total_shots'],'',
             f"{p['conf']['shots']}%"],
            ['SOT (pe poarta)', p['sot_h'], p['sot_a'], p['total_sot'],'',
             f"{p['conf']['sot']}%"],
            ['Over 7.5 SOT','','','',
             f"{round(p_over_half(7, p['total_sot'])*100,1)}%", f"{p['conf']['sot']}%"],
            ['Over 8.5 SOT','','','',
             f"{round(p_over_half(8, p['total_sot'])*100,1)}%", f"{p['conf']['sot']}%"],
            ['Over 9.5 SOT','','','',
             f"{round(p_over_half(9, p['total_sot'])*100,1)}%", f"{p['conf']['sot']}%"],
        ]: dr(row, pcs=(5,), cc=6)

        # ─ Cornere ───────────────────────────────────────────────────────────
        hr(['Cornere', 'Acasa', 'Deplasare', 'Total', 'Prob.%', 'Incredere'], 'corners')
        for row in [
            ['Cornere (est.)', p['corn_h'], p['corn_a'], p['total_corn'],'',
             f"{p['conf']['corners']}%"],
            ['Over 8.5','','','',
             f"{round(p_over_half(8,p['total_corn'])*100,1)}%", f"{p['conf']['corners']}%"],
            ['Over 9.5','','','',
             f"{round(p_over_half(9,p['total_corn'])*100,1)}%", f"{p['conf']['corners']}%"],
            ['Over 10.5','','','',
             f"{round(p_over_half(10,p['total_corn'])*100,1)}%", f"{p['conf']['corners']}%"],
            ['Over 11.5','','','',
             f"{round(p_over_half(11,p['total_corn'])*100,1)}%", f"{p['conf']['corners']}%"],
            ['Over 12.5','','','',
             f"{round(p_over_half(12,p['total_corn'])*100,1)}%", f"{p['conf']['corners']}%"],
        ]: dr(row, pcs=(5,), cc=6)

        # ─ Cartonase / Faulturi / Ofsaid ─────────────────────────────────────
        hr(['Cartonase / Faulturi / Ofsaid', 'Acasa', 'Deplasare', 'Total',
            'Prob.%', 'Incredere'], 'cards')
        for row in [
            ['G. Galbene (est.)', p['yc_h'], p['yc_a'], p['total_yc'],'',
             f"{p['conf']['yellows']}%"],
            ['Over 1.5 Galbene','','','',
             f"{round(p_over_half(1,p['total_yc'])*100,1)}%", f"{p['conf']['yellows']}%"],
            ['Over 2.5 Galbene','','','',
             f"{round(p_over_half(2,p['total_yc'])*100,1)}%", f"{p['conf']['yellows']}%"],
            ['Over 3.5 Galbene','','','',
             f"{round(p_over_half(3,p['total_yc'])*100,1)}%", f"{p['conf']['yellows']}%"],
            ['Over 4.5 Galbene','','','',
             f"{round(p_over_half(4,p['total_yc'])*100,1)}%", f"{p['conf']['yellows']}%"],
            ['Over 5.5 Galbene','','','',
             f"{round(p_over_half(5,p['total_yc'])*100,1)}%", f"{p['conf']['yellows']}%"],
            ['Faulturi (est.)', p['fouls_h'], p['fouls_a'], p['total_fouls'],'',
             f"{p['conf']['fouls']}%"],
            ['Over 20.5 Faulturi','','','',
             f"{round(p_over_half(20,p['total_fouls'])*100,1)}%", f"{p['conf']['fouls']}%"],
            ['Over 22.5 Faulturi','','','',
             f"{round(p_over_half(22,p['total_fouls'])*100,1)}%", f"{p['conf']['fouls']}%"],
            ['Over 25.5 Faulturi','','','',
             f"{round(p_over_half(25,p['total_fouls'])*100,1)}%", f"{p['conf']['fouls']}%"],
            ['Over 28.5 Faulturi','','','',
             f"{round(p_over_half(28,p['total_fouls'])*100,1)}%", f"{p['conf']['fouls']}%"],
            ['Ofsaid (est.)', p['offs_h'], p['offs_a'], p['total_offs'],'',
             f"{p['conf']['offsides']}%"],
            ['Over 1.5 Ofsaid','','','',
             f"{round(p_over_half(1,p['total_offs'])*100,1)}%", f"{p['conf']['offsides']}%"],
            ['Over 2.5 Ofsaid','','','',
             f"{round(p_over_half(2,p['total_offs'])*100,1)}%", f"{p['conf']['offsides']}%"],
            ['Over 3.5 Ofsaid','','','',
             f"{round(p_over_half(3,p['total_offs'])*100,1)}%", f"{p['conf']['offsides']}%"],
            ['Over 4.5 Ofsaid','','','',
             f"{round(p_over_half(4,p['total_offs'])*100,1)}%", f"{p['conf']['offsides']}%"],
        ]: dr(row, pcs=(5,), cc=6)

        # ─ Marcatori & Speciale ───────────────────────────────────────────────
        hr(['Marcatori & Speciale', 'Prob. %', 'Incredere','','',''], 'scorers')
        def fpl(lst): return ' | '.join(f"{n} ({pp}%)" for n, pp in lst) if lst else 'N/A'
        for row in [
            [f'Marcatori {p["home"][:18]}', fpl(p['scorers_h']), f"{p['conf']['scorers']}%",'','',''],
            [f'Marcatori {p["away"][:18]}', fpl(p['scorers_a']), f"{p['conf']['scorers']}%",'','',''],
            ['Penalty in meci', f"{p['p_penalty']}%", f"{p['conf']['specials']}%",'','',''],
            ['Gol cu capul', f"{p['p_header']}%", f"{p['conf']['specials']}%",'','',''],
            ['Rosu in meci', f"{p['p_red']}%", f"{p['conf']['specials']}%",'','',''],
            [f'Win to Nil {p["home"][:14]}', f"{p['p_h_win_nil']}%", f"{p['conf']['win_nil']}%",'','',''],
            [f'Win to Nil {p["away"][:14]}', f"{p['p_a_win_nil']}%", f"{p['conf']['win_nil']}%",'','',''],
            [f'G.Galben juc. {p["home"][:14]}', fpl(p['yc_pl_h']), f"{p['conf']['specials']}%",'','',''],
            [f'G.Galben juc. {p["away"][:14]}', fpl(p['yc_pl_a']), f"{p['conf']['specials']}%",'','',''],
        ]: dr(row, pcs=(2,), cc=3)

        # ─ Top pariuri per meci ───────────────────────────────────────────────
        hr(['TOP 10 Pariuri (Incredere desc.)', 'Predictie', 'Probabilitate',
            'Incredere','',''], 'elo')
        for mkt, pred, prob, c_pct in p['all_bets'][:10]:
            dr([mkt, pred, f"{prob}%", f"{c_pct}%",'',''], pcs=(3,), cc=4)

        ws2.row_dimensions[cur[0]].height = 8; cur[0] += 1

    # ── Sheet 3: Best Bets globally sorted ──────────────────────────────────
    ws3 = wb.create_sheet('Best_Bets')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:G1')
    c = ws3['A1']
    c.value = ('★ TOATE PARIURILE — ordonate descrescator dupa INCREDERE'
               ' | 1X2, DC, DNB, AH, O/U 0.5-5.5, GG, BTTS+OU, R1, Win-Nil,'
               ' Scor Exact, SOT, Cornere, Galbene, Faulturi, Ofsaid, Marcatori, Speciale')
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = F['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 24

    for ci, (h, w) in enumerate(zip(
        ['Meci', 'Piata', 'Predictie (model)', 'Probabilitate %', 'Incredere Model %', 'Gr.', 'Data'],
        [32, 30, 24, 15, 17, 5, 12]
    ), 1):
        c = ws3.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = F['elo']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 22

    # Collect all bets globally, sort by confidence desc
    all_bets_global = []
    for p in preds:
        mn = f"{p['home']} - {p['away']}"
        for mkt, pred, prob, cp in p['all_bets']:
            if prob is not None:
                all_bets_global.append((mn, mkt, pred, prob, cp, p['group'], p['date']))
    all_bets_global.sort(key=lambda x: (-x[4], -x[3]))

    for ri, (mn, mkt, pred, prob, cp, grp, dt) in enumerate(all_bets_global, 3):
        vals = [mn, mkt, pred, f"{prob}%", f"{cp}%", grp, dt]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(ri, ci, v)
            c.font = Font(size=9, bold=(ci == 3))
            c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci not in (1,2,3) else 'left', vertical='center')
        ws3.cell(ri, 4).fill = pf_p(prob)
        ws3.cell(ri, 5).fill = pf_c(cp)
        ws3.row_dimensions[ri].height = 13

    # ── Sheet 4: Recent form matches ─────────────────────────────────────────
    ws4 = wb.create_sheet('Forma_Recenta')
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:H1')
    c = ws4['A1']
    c.value = 'FORMA RECENTA — Meciuri folosite in model per echipa (verde=oficial, portocaliu=amical)'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = F['title']; c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 22

    for ci, (h, w) in enumerate(zip(
        ['Echipa', 'Data', 'Adversar', 'Scor', 'Turneu', 'Pondere\n(oficial=1.5x)', 'Elo Adv.', 'Rol'],
        [20, 12, 20, 8, 28, 14, 10, 6]
    ), 1):
        c = ws4.cell(2, ci, h)
        c.font = Font(bold=True, color='FFFFFF', size=8)
        c.fill = F['form']; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 28

    ri = 3
    seen_teams = set()
    for group, home, away, date in ROUND1:
        for team in [home, away]:
            if team in seen_teams: continue
            seen_teams.add(team)
            matches = RECENT_BY_TEAM.get(team, [])[-10:]
            for m in matches:
                bg = _fill('E8F8E8') if m['weight'] > 0.75 else _fill('FFF3E0')
                gf, ga = m['goals_for'], m['goals_against']
                vals = [team, m['date'], m['opp'], f"{gf}-{ga}", m['tournament'],
                        round(m['weight'],1), m['opp_elo'], m['role']]
                for ci, v in enumerate(vals, 1):
                    c = ws4.cell(ri, ci, v)
                    c.font = Font(size=8); c.border = THIN; c.fill = bg
                    c.alignment = Alignment(horizontal='center' if ci not in (1,3,5) else 'left',
                                            vertical='center')
                ws4.row_dimensions[ri].height = 13; ri += 1
            ws4.row_dimensions[ri].height = 4; ri += 1

    wb.save('predictii_etapa1.xlsx')
    return 'predictii_etapa1.xlsx'


# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Computing predictions...')
    preds = []
    for group, home, away, date in ROUND1:
        p = predict_match(group, home, away, date)
        preds.append(p)
        fh, fa = p['form_h'], p['form_a']
        print(f"  Gr.{group} {home[:12]:12} vs {away[:12]:12} "
              f"Elo {p['helo']}-{p['aelo']} D{p['elo_diff']:3d} "
              f"xG {p['lambda_h']:.2f}-{p['lambda_a']:.2f} "
              f"O2.5={p['p_o25']}% GG={p['p_gg']}% "
              f"Form {fh['n_total']}m({fh['n_official']}off)/{fa['n_total']}m({fa['n_official']}off)")

    print('\nBuilding Excel...')
    path = build_excel(preds)
    print(f'Saved: {path}')

    avg_g = sum(p['total_goals'] for p in preds) / len(preds)
    n_o25 = sum(1 for p in preds if p['p_o25'] > 55)
    n_gg  = sum(1 for p in preds if p['p_gg'] > 50)
    print(f'\nAvg xG/meci: {avg_g:.2f}')
    print(f'Meciuri Over2.5>55%: {n_o25}/24 | GG>50%: {n_gg}/24')
    print('\n=== Top 15 pariuri dupa incredere (global) ===')
    all_b = []
    for p in preds:
        for mkt, pred, prob, cp in p['all_bets']:
            if prob: all_b.append((f"{p['home'][:8]}-{p['away'][:8]}", mkt, pred, prob, cp))
    all_b.sort(key=lambda x: (-x[4], -x[3]))
    for mn, mkt, pred, prob, cp in all_b[:15]:
        print(f"  [{cp:2d}%conf, {prob:5.1f}%] {mn:20} | {mkt:30} => {pred}")
