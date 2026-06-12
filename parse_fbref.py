"""Parser pentru date FBref copiate din browser.

Cum se folosește:
  1. Mergi pe FBref la tabelul dorit (Squad Standard Stats, Squad Goalkeeping, Player Shooting).
  2. Selectează TOATĂ tabela (Ctrl+A pe tabelă sau selectează manual de la headerul primei coloane
     până la ultimul rând).
  3. Copiază (Ctrl+C) și lipește într-un fișier .txt (ex: fbref_standard.txt).
  4. Rulează: python -X utf8 parse_fbref.py

Produce: fbref_merged.json cu statistici per echipă UEFA, gata de fuzionat în Excel.
"""
import json
import os
import re
import sys


# --- Mapare nume FBref -> nume echipă din config.py ---
TEAM_MAP = {
    "England": "England",
    "France": "France",
    "Croatia": "Croatia",
    "Norway": "Norway",
    "Portugal": "Portugal",
    "Germany": "Germany",
    "Netherlands": "Netherlands",
    "Switzerland": "Switzerland",
    "Scotland": "Scotland",
    "Spain": "Spain",
    "Austria": "Austria",
    "Belgium": "Belgium",
    "Bosnia & Herzegovina": "Bosnia & Herzegovina",
    "Bosnia-Herzegovina": "Bosnia & Herzegovina",
    "Bosnia Herzeg": "Bosnia & Herzegovina",
    "Sweden": "Sweden",
    "Turkey": "Türkiye",
    "Türkiye": "Türkiye",
    "Czechia": "Czechia",
    "Czech Republic": "Czechia",
    # CONMEBOL
    "Argentina": "Argentina",
    "Brazil": "Brazil",
    "Colombia": "Colombia",
    "Ecuador": "Ecuador",
    "Paraguay": "Paraguay",
    "Uruguay": "Uruguay",
    # CAF
    "Algeria": "Algeria",
    "Cape Verde": "Cape Verde",
    "Cabo Verde": "Cape Verde",
    "Egypt": "Egypt",
    "Ghana": "Ghana",
    "Ivory Coast": "Ivory Coast",
    "Côte d'Ivoire": "Ivory Coast",
    "Cote d'Ivoire": "Ivory Coast",
    "Morocco": "Morocco",
    "Senegal": "Senegal",
    "South Africa": "South Africa",
    "Tunisia": "Tunisia",
    # CONCACAF
    "United States": "USA",
    "USA": "USA",
    "Mexico": "Mexico",
    "Canada": "Canada",
    "Panama": "Panama",
    "Curaçao": "Curacao",
    "Curacao": "Curacao",
    "Haiti": "Haiti",
    # AFC
    "Australia": "Australia",
    "Iran": "Iran",
    "Japan": "Japan",
    "Jordan": "Jordan",
    "Qatar": "Qatar",
    "Saudi Arabia": "Saudi Arabia",
    "Korea Republic": "South Korea",
    "South Korea": "South Korea",
    "Uzbekistan": "Uzbekistan",
    # OFC
    "New Zealand": "New Zealand",
    # Play-off
    "DR Congo": "DR Congo",
    "Dem. Rep. Congo": "DR Congo",
    "Congo DR": "DR Congo",
    "Iraq": "Iraq",
}


def _strip_markers(text):
    """Elimină rândurile cu note de subsol FBref (încep cu * sau sunt goale)."""
    lines = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("*") or stripped.startswith("#"):
            continue
        lines.append(stripped)
    return lines


def _find_header_row(lines):
    """Găsește indexul rândului cu header (conține 'Squad' sau 'Player')."""
    for i, line in enumerate(lines):
        cols = re.split(r"\t", line)
        first = cols[0].strip().lower() if cols else ""
        if first in ("squad", "player", "rk"):
            return i
    return None


def _squad_name_clean(raw):
    """Elimină prefixul codul-țară de 2-3 litere minuscule (ex: 'at Austria' -> 'Austria')."""
    raw = raw.strip()
    parts = raw.split(" ", 1)
    if len(parts) == 2 and re.match(r'^[a-z]{2,3}$', parts[0]):
        return parts[1].strip()
    return raw


def parse_squad_table(filepath):
    """Parsează un tabel FBref de tip Squad (Standard Stats sau Goalkeeping).
    Întoarce {echipa: {coloana: valoare}}.

    Dacă tabela conține coloane per-90 (Gls, Ast, CrdY, CrdR sub 'Performance'),
    calculează și totaluri: total = val_per90 × 90s."""
    if not os.path.exists(filepath):
        return {}
    with open(filepath, encoding="utf-8") as f:
        raw = f.read()

    lines = _strip_markers(raw)
    hi = _find_header_row(lines)
    if hi is None:
        print(f"  ! Nu am găsit header în {filepath}")
        return {}

    raw_headers = [h.strip() for h in re.split(r"\t", lines[hi])]
    # Deduplică coloane cu același nume (ex: "Gls" apare de 2 ori în Standard Stats)
    seen_counts = {}
    headers = []
    for h in raw_headers:
        cnt = seen_counts.get(h, 0)
        seen_counts[h] = cnt + 1
        headers.append(h if cnt == 0 else f"{h}_{cnt}")

    result = {}
    for line in lines[hi + 1:]:
        cols = re.split(r"\t", line)
        if len(cols) < 2:
            continue
        row = {headers[i]: cols[i].strip() for i in range(min(len(headers), len(cols)))}
        squad_raw = row.get("Squad", "").strip()
        squad_raw = _squad_name_clean(squad_raw)
        squad = TEAM_MAP.get(squad_raw, squad_raw)
        if not squad:
            continue

        # Dacă există "90s", calculăm totaluri din valorile per-90
        nineties = _num(row.get("90s"))
        if nineties and nineties > 0:
            for col in ("Gls", "Ast", "G+A", "G-PK", "CrdY", "CrdR"):
                per90_val = _num(row.get(col))
                if per90_val is not None:
                    row[f"{col}_total"] = round(per90_val * nineties)

        result[squad] = row
    return result


def parse_player_table(filepath):
    """Parsează tabelul Player Shooting FBref.
    Întoarce {echipa: [{coloana: valoare}, ...]} grupat pe Squad."""
    if not os.path.exists(filepath):
        return {}
    with open(filepath, encoding="utf-8") as f:
        raw = f.read()

    lines = _strip_markers(raw)
    hi = _find_header_row(lines)
    if hi is None:
        print(f"  ! Nu am găsit header în {filepath}")
        return {}

    headers = [h.strip() for h in re.split(r"\t", lines[hi])]
    result = {}

    for line in lines[hi + 1:]:
        cols = re.split(r"\t", line)
        if len(cols) < 3:
            continue
        row = {headers[i]: cols[i].strip() for i in range(min(len(headers), len(cols)))}
        squad_raw = row.get("Squad", "").strip()
        squad = TEAM_MAP.get(squad_raw, squad_raw)
        if squad:
            result.setdefault(squad, []).append(row)
    return result


def _num(val, default=None):
    """Convertește string la float; ignoră 'N/A', '', '-'."""
    if val is None or str(val).strip() in ("", "-", "N/A", "n/a"):
        return default
    try:
        return float(str(val).replace(",", ".").replace("%", ""))
    except ValueError:
        return default


def merge_squad_stats(standard, goalkeeping, shooting_by_squad):
    """Combină cele trei surse într-un dict per echipă cu statistici relevante.

    Preferă coloanele *_total (calculate din per90 × 90s) față de valorile brute
    per-90 când tabelul standard conține rate, nu totaluri absolute.
    """
    all_teams = set(standard) | set(goalkeeping)
    merged = {}

    for team in all_teams:
        std = standard.get(team, {})
        gk = goalkeeping.get(team, {})
        players = shooting_by_squad.get(team, [])

        # Statistici ofensive (Squad Standard Stats)
        xg_total = _num(std.get("xG"))
        npxg_total = _num(std.get("npxG"))
        poss = _num(std.get("Poss"))
        mp = _num(std.get("MP"))

        # Preferă _total (din per-90 × 90s) dacă există, altfel valoarea brută
        gls  = _num(std.get("Gls_total"))  or _num(std.get("Gls"))
        ast  = _num(std.get("Ast_total"))  or _num(std.get("Ast"))
        crd_y= _num(std.get("CrdY_total")) or _num(std.get("CrdY"))
        crd_r= _num(std.get("CrdR_total")) or _num(std.get("CrdR"))

        shots = _num(std.get("Sh"))       # coloana "Sh" în tabelul de shooting
        sot = _num(std.get("SoT"))        # "SoT" în shooting

        # Statistici portar (Squad Goalkeeping)
        ga = _num(gk.get("GA"))
        ga90 = _num(gk.get("GA90"))
        saves = _num(gk.get("Saves"))
        save_pct = _num(gk.get("Save%"))
        cs = _num(gk.get("CS"))
        sota = _num(gk.get("SoTA"))       # Shots on Target Against

        # Sumare jucători din Player Shooting
        pl_xg_sum = None
        pl_sh_sum = None
        pl_sot_sum = None
        if players:
            pl_xg_vals = [_num(p.get("xG")) for p in players]
            pl_sh_vals = [_num(p.get("Sh")) for p in players]
            pl_sot_vals = [_num(p.get("SoT")) for p in players]
            pl_xg_vals = [v for v in pl_xg_vals if v is not None]
            pl_sh_vals = [v for v in pl_sh_vals if v is not None]
            pl_sot_vals = [v for v in pl_sot_vals if v is not None]
            if pl_xg_vals:
                pl_xg_sum = round(sum(pl_xg_vals), 2)
            if pl_sh_vals:
                pl_sh_sum = int(sum(pl_sh_vals))
            if pl_sot_vals:
                pl_sot_sum = int(sum(pl_sot_vals))

        merged[team] = {
            "fbref_mp": int(mp) if mp is not None else None,
            "fbref_xg": xg_total,
            "fbref_npxg": npxg_total,
            "fbref_poss": poss,
            "fbref_gls": int(gls) if gls is not None else None,
            "fbref_ast": int(ast) if ast is not None else None,
            "fbref_yellow": int(crd_y) if crd_y is not None else None,
            "fbref_red": int(crd_r) if crd_r is not None else None,
            "fbref_shots": int(shots) if shots is not None else (pl_sh_sum),
            "fbref_sot": int(sot) if sot is not None else (pl_sot_sum),
            # Goalkeeping
            "fbref_ga": int(ga) if ga is not None else None,
            "fbref_ga90": ga90,
            "fbref_saves": int(saves) if saves is not None else None,
            "fbref_save_pct": save_pct,
            "fbref_cs": int(cs) if cs is not None else None,
            "fbref_sota": int(sota) if sota is not None else None,
            # Din sumarea jucătorilor (backup dacă lipsesc coloanele echipă)
            "fbref_pl_xg": pl_xg_sum,
            "fbref_pl_shots": pl_sh_sum,
            "fbref_pl_sot": pl_sot_sum,
        }
    return merged


def add_fbref_sheet(wb_path, fbref_data):
    """Adaugă foaia FBref_Extra la un Excel existent."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  ! openpyxl nu e instalat")
        return

    wb = load_workbook(wb_path)
    if "FBref_Extra" in wb.sheetnames:
        del wb["FBref_Extra"]

    ws = wb.create_sheet("FBref_Extra")

    HEADERS = [
        "Echipă", "Meciuri jucate",
        "xG (FBref)", "npxG (FBref)", "Posesie (%)",
        "Goluri", "Assist", "Galbene", "Roșii",
        "Șuturi", "Șuturi pe poartă",
        "GA (goluri primite)", "GA/90", "Salvări", "Salvări %", "Clean Sheets", "SoTA",
        "xG jucători (sum.)", "Șuturi jucători (sum.)", "SoT jucători (sum.)",
    ]
    KEYS = [
        None, "fbref_mp",
        "fbref_xg", "fbref_npxg", "fbref_poss",
        "fbref_gls", "fbref_ast", "fbref_yellow", "fbref_red",
        "fbref_shots", "fbref_sot",
        "fbref_ga", "fbref_ga90", "fbref_saves", "fbref_save_pct", "fbref_cs", "fbref_sota",
        "fbref_pl_xg", "fbref_pl_shots", "fbref_pl_sot",
    ]

    fill = PatternFill("solid", start_color="1F4E78")
    hfont = Font(name="Arial", bold=True, color="FFFFFF")
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hfont
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for team, stats in sorted(fbref_data.items()):
        row = [team] + [stats.get(k) for k in KEYS[1:]]
        ws.append(row)

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 28)

    ws.freeze_panes = "A2"
    wb.save(wb_path)
    print(f"  Foaia 'FBref_Extra' adăugată la {wb_path}")


def main():
    print("=== Parser FBref ===")
    print()
    print("Fișiere așteptate (tab-separate, copiate din browser):")
    print("  fbref_standard.txt   -> Squad Standard Stats (xG, Posesie, Goluri, Cartonașe)")
    print("  fbref_gk.txt         -> Squad Goalkeeping Stats (Save%, CS, GA90)")
    print("  fbref_shooting.txt   -> Player Shooting Stats (xG per jucător, Sh, SoT)")
    print()

    standard = parse_squad_table("fbref_standard.txt")
    gk = parse_squad_table("fbref_gk.txt")
    shooting = parse_player_table("fbref_shooting.txt")

    if not standard and not gk and not shooting:
        print("  ! Niciunul din fișierele fbref_*.txt nu a fost găsit.")
        print("    Creează fișierele cu datele copiate din FBref și rulează din nou.")
        return

    print(f"  Standard Stats: {len(standard)} echipe")
    print(f"  Goalkeeping:    {len(gk)} echipe")
    print(f"  Shooting:       {len(shooting)} echipe, {sum(len(v) for v in shooting.values())} jucători")

    merged = merge_squad_stats(standard, gk, shooting)
    print(f"  Total echipe după merge: {len(merged)}")

    out_json = "fbref_merged.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"  Salvat: {out_json}")

    # Caută cel mai recent Excel și adaugă foaia FBref_Extra
    import glob as _glob
    xlsx_files = sorted(_glob.glob("statistici_calificari_cm2026*.xlsx"))
    if xlsx_files:
        xlsx = xlsx_files[-1]
        print(f"  Adaug foaia la: {xlsx}")
        add_fbref_sheet(xlsx, merged)
    else:
        print("  ! Nu am găsit fișierul xlsx. Rulează main.py mai întâi.")

    # Afișează sumarul pentru echipele calificate UEFA
    print()
    print("Echipe calificate UEFA extrase din FBref:")
    print(f"  {'Echipă':<25} {'Meciuri':>7} {'Poss':>6} {'Goluri':>7} {'Ast':>5} {'Galbene':>8} {'Roșii':>6}")
    print(f"  {'-'*25} {'-'*7} {'-'*6} {'-'*7} {'-'*5} {'-'*8} {'-'*6}")
    for team in sorted(merged):
        d = merged[team]
        print(f"  {team:<25} {str(d.get('fbref_mp') or '-'):>7} "
              f"{str(d.get('fbref_poss') or '-'):>6} "
              f"{str(d.get('fbref_gls') or '-'):>7} "
              f"{str(d.get('fbref_ast') or '-'):>5} "
              f"{str(d.get('fbref_yellow') or '-'):>8} "
              f"{str(d.get('fbref_red') or '-'):>6}")

    print()
    print("Gata! Deschide Excelul și caută foaia 'FBref_Extra'.")


if __name__ == "__main__":
    main()
