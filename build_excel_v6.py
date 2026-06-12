"""Build Excel v6: combined sheet merging AiScore + Elo + FBref + Sofascore per WC team."""
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Name normalisation: all sources -> canonical (Elo) name ─────────────────
ALIAS_TO_CANON = {
    # AiScore -> Elo
    "Turkiye":                             "Turkey",
    "Cote d'Ivoire":                       "Ivory Coast",
    "IR Iran":                             "Iran",
    "Democratic Republic of the Congo":    "DR Congo",
    "USA":                                 "United States",
    "Curacao":                             "Curaçao",
    "Cabo Verde":                          "Cape Verde",
    # FBref -> Elo
    "Côte d'Ivoire":                       "Ivory Coast",
    "Türkiye":                             "Turkey",
    "Bosnia-Herzegovina":                  "Bosnia and Herzegovina",
    "Bosnia & Herzegovina":                "Bosnia and Herzegovina",
    # Sofascore -> Elo
    "Ivory Coast":                         "Ivory Coast",
    "DR Congo":                            "DR Congo",
    "Bosnia & Herzegovina":                "Bosnia and Herzegovina",
    # eloratings already correct for its own data
}

# Confederation per team (Elo canonical)
CONF_MAP = {
    "Spain": "UEFA", "France": "UEFA", "England": "UEFA", "Germany": "UEFA",
    "Norway": "UEFA", "Croatia": "UEFA", "Portugal": "UEFA", "Netherlands": "UEFA",
    "Belgium": "UEFA", "Switzerland": "UEFA", "Scotland": "UEFA", "Austria": "UEFA",
    "Bosnia and Herzegovina": "UEFA", "Sweden": "UEFA", "Turkey": "UEFA",
    "Czechia": "UEFA",
    "Argentina": "CONMEBOL", "Brazil": "CONMEBOL", "Colombia": "CONMEBOL",
    "Ecuador": "CONMEBOL", "Paraguay": "CONMEBOL", "Uruguay": "CONMEBOL",
    "Algeria": "CAF", "Cape Verde": "CAF", "Egypt": "CAF", "Ghana": "CAF",
    "Ivory Coast": "CAF", "Morocco": "CAF", "Senegal": "CAF",
    "South Africa": "CAF", "Tunisia": "CAF",
    "United States": "CONCACAF", "Mexico": "CONCACAF", "Canada": "CONCACAF",
    "Panama": "CONCACAF", "Curaçao": "CONCACAF", "Haiti": "CONCACAF",
    "Australia": "AFC", "Iran": "AFC", "Japan": "AFC", "Jordan": "AFC",
    "Qatar": "AFC", "Saudi Arabia": "AFC", "South Korea": "AFC", "Uzbekistan": "AFC",
    "New Zealand": "OFC",
    "DR Congo": "CAF", "Iraq": "AFC",
}

# ─── Styles ──────────────────────────────────────────────────────────────────
def _hdr(fill_hex):
    fill = PatternFill('solid', fgColor=fill_hex)
    font = Font(bold=True, color='FFFFFF', size=10)
    return fill, font

ELO_FILL,   ELO_FONT   = _hdr('1F4E79')
AIS_FILL,   AIS_FONT   = _hdr('375623')
FBREF_FILL, FBREF_FONT = _hdr('7B2C2C')
SFA_FILL,   SFA_FONT   = _hdr('7B5B00')
BASE_FILL,  BASE_FONT  = _hdr('404040')

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WIN_FILL  = PatternFill('solid', fgColor='C6EFCE')
DRAW_FILL = PatternFill('solid', fgColor='FFEB9C')
LOSS_FILL = PatternFill('solid', fgColor='FFC7CE')
GOLD_FILL = PatternFill('solid', fgColor='FFD700')
BLUE_FILL = PatternFill('solid', fgColor='BDD7EE')

CONF_FILLS = {
    'UEFA':     PatternFill('solid', fgColor='DDEEFF'),
    'CONMEBOL': PatternFill('solid', fgColor='FFF2CC'),
    'CAF':      PatternFill('solid', fgColor='FFE0CC'),
    'CONCACAF': PatternFill('solid', fgColor='E2F0D9'),
    'AFC':      PatternFill('solid', fgColor='F2E0FF'),
    'OFC':      PatternFill('solid', fgColor='E0F7FA'),
}


def _canon(name):
    return ALIAS_TO_CANON.get(name, name)


def _num(s):
    if s is None: return None
    s = str(s).strip().replace('−', '-').replace('−', '-')
    m = re.match(r'^(-?\d+(?:\.\d+)?)', s)
    return float(m.group(1)) if m else None


def _pk(s):
    """Extract penalty goals from '28(2)' -> 2."""
    if s is None: return None
    m = re.search(r'\((\d+)\)', str(s))
    return int(m.group(1)) if m else None


def load_aiscore():
    """Return dict: canon_name -> {stat: value}."""
    with open('aiscore_team_stats.json', encoding='utf-8') as f:
        raw = json.load(f)
    result = {}
    for _conf, teams in raw.items():
        for name, stats in teams.items():
            canon = _canon(name)
            if canon not in result:
                result[canon] = {}
            for k, v in stats.items():
                if k not in result[canon]:
                    result[canon][k] = v
    return result


def load_elo():
    """Return dict: canon_name -> elo_row."""
    with open('elo_data.json', encoding='utf-8') as f:
        data = json.load(f)
    result = {}
    for row in data['wc_ratings']:
        canon = _canon(row['team'])
        result[canon] = row
    return result


def load_fbref(wb):
    """Read FBref_Extra from workbook. Return dict: canon_name -> row dict."""
    ws = wb['FBref_Extra']
    rows = list(ws.rows)
    headers = [c.value for c in rows[0]]
    result = {}
    for row in rows[1:]:
        name = row[0].value
        if not name: continue
        canon = _canon(name)
        result[canon] = {headers[i]: row[i].value for i in range(1, len(headers))}
    return result


def load_sofascore_meciuri(wb):
    """Compute per-team stats from Meciuri sheet (skip rows with empty stats)."""
    ws = wb['Meciuri']
    rows = list(ws.rows)
    headers = [c.value for c in rows[0]]

    col = {h: i for i, h in enumerate(headers) if h}

    def _get(row, colname, default=None):
        i = col.get(colname)
        if i is None: return default
        return row[i].value

    teams = {}
    for row in rows[1:]:
        team = row[0].value
        if not team: continue
        # Skip rows with no shot data (incomplete Sofascore rows)
        if _get(row, 'Șuturi total') is None:
            continue
        canon = _canon(team)
        if canon not in teams:
            teams[canon] = {'meciuri': 0, 'xg': [], 'posesie': [], 'suturi': [], 'suturi_pm': []}
        d = teams[canon]
        d['meciuri'] += 1
        xg = _get(row, 'xG')
        if xg is not None: d['xg'].append(float(xg))
        pos = _get(row, 'Posesie (%)')
        if pos is not None: d['posesie'].append(float(pos))
        s = _get(row, 'Șuturi total')
        if s is not None: d['suturi'].append(float(s))
        spm = _get(row, 'Șuturi pe poartă')
        if spm is not None: d['suturi_pm'].append(float(spm))

    def _avg(lst): return round(sum(lst)/len(lst), 2) if lst else None
    def _tot(lst): return round(sum(lst), 1) if lst else None

    return {
        name: {
            'meciuri_sfa': d['meciuri'],
            'xg_total': _tot(d['xg']),
            'xg_per_meci': _avg(d['xg']),
            'suturi_total_sfa': _tot(d['suturi']),
            'suturi_pm_sfa': _tot(d['suturi_pm']),
            'posesie_medie': _avg(d['posesie']),
        }
        for name, d in teams.items()
    }


def _set_hdr(ws, row, col, text, fill, font):
    c = ws.cell(row, col, text)
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN


def _set_val(ws, row, col, val, align='center', bold=False, fill=None):
    c = ws.cell(row, col, val)
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = THIN
    if fill: c.fill = fill


def add_sumar_complet(wb, aiscore, elo, fbref, sofascore):
    ws = wb.create_sheet('Sumar_Complet', 0)  # insert as first sheet
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'C3'

    # ── Double-row header ────────────────────────────────────────────────────
    # Row 1: section titles
    # Row 2: column names
    sections = [
        (1, 2, 'Echipă', BASE_FILL, BASE_FONT),
        (3, 7, 'ELO RATINGS', ELO_FILL, ELO_FONT),
        (8, 12, 'ELO – STATISTICI ALL-TIME', ELO_FILL, ELO_FONT),
        (13, 24, 'AISCORE – CALIFICARI WCQ 2026', AIS_FILL, AIS_FONT),
        (25, 30, 'FBREF – CALIFICARI (UEFA)', FBREF_FILL, FBREF_FONT),
        (31, 35, 'SOFASCORE – MECIURI ANALIZATE', SFA_FILL, SFA_FONT),
    ]
    for c_start, c_end, title, fill, font in sections:
        cell = ws.cell(1, c_start, title)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
        if c_start != c_end:
            ws.merge_cells(
                start_row=1, start_column=c_start,
                end_row=1, end_column=c_end
            )

    # Row 2: actual column names
    cols2 = [
        # Base
        (1, 'Echipă', BASE_FILL, BASE_FONT),
        (2, 'Confederație', BASE_FILL, BASE_FONT),
        # Elo current
        (3, 'Rang\nGlobal', ELO_FILL, ELO_FONT),
        (4, 'Rating\nElo', ELO_FILL, ELO_FONT),
        (5, 'Trend\n1 An', ELO_FILL, ELO_FONT),
        (6, 'Trend\n2 Ani', ELO_FILL, ELO_FONT),
        (7, 'Trend\n5 Ani', ELO_FILL, ELO_FONT),
        # Elo all-time
        (8,  'Meciuri\nTotal', ELO_FILL, ELO_FONT),
        (9,  'Victorii', ELO_FILL, ELO_FONT),
        (10, 'Egaluri', ELO_FILL, ELO_FONT),
        (11, 'Infrangeri', ELO_FILL, ELO_FONT),
        (12, 'Rata\nVictorii %', ELO_FILL, ELO_FONT),
        # AiScore
        (13, 'Goluri\nMarcate', AIS_FILL, AIS_FONT),
        (14, 'Goluri\nPK', AIS_FILL, AIS_FONT),
        (15, 'Pase\nDecisive', AIS_FILL, AIS_FONT),
        (16, 'Carti\nRosii', AIS_FILL, AIS_FONT),
        (17, 'Carti\nGalbene', AIS_FILL, AIS_FONT),
        (18, 'Suturi\nTotal', AIS_FILL, AIS_FONT),
        (19, 'Suturi pe\nPoarta', AIS_FILL, AIS_FONT),
        (20, 'Degajari', AIS_FILL, AIS_FONT),
        (21, 'Pase\nCheie', AIS_FILL, AIS_FONT),
        (22, 'Centrari\nTotal', AIS_FILL, AIS_FONT),
        (23, 'Centrari\nPrecise', AIS_FILL, AIS_FONT),
        (24, 'Faulturi\nComise', AIS_FILL, AIS_FONT),
        # FBref
        (25, 'xG\n(FBref)', FBREF_FILL, FBREF_FONT),
        (26, 'npxG\n(FBref)', FBREF_FILL, FBREF_FONT),
        (27, 'Posesie %\n(FBref)', FBREF_FILL, FBREF_FONT),
        (28, 'Goluri\n(FBref)', FBREF_FILL, FBREF_FONT),
        (29, 'Goluri\nPrimite', FBREF_FILL, FBREF_FONT),
        (30, 'Salvari %\n(FBref)', FBREF_FILL, FBREF_FONT),
        # Sofascore
        (31, 'Meciuri\nAnalizate', SFA_FILL, SFA_FONT),
        (32, 'xG Total\n(Sfa)', SFA_FILL, SFA_FONT),
        (33, 'xG/Meci\n(Sfa)', SFA_FILL, SFA_FONT),
        (34, 'Suturi/Meci\n(Sfa)', SFA_FILL, SFA_FONT),
        (35, 'Posesie %\n(Sfa)', SFA_FILL, SFA_FONT),
    ]
    for col_idx, text, fill, font in cols2:
        c = ws.cell(2, col_idx, text)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32

    # ── Column widths ────────────────────────────────────────────────────────
    widths = [26, 12, 9, 9, 10, 10, 10, 10, 10, 9, 11, 12,
              10, 9, 10, 9, 11, 10, 10, 10, 10, 10, 10, 11,
              10, 10, 10, 10, 10, 11, 11, 10, 10, 11, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ────────────────────────────────────────────────────────────
    # Sort: by Elo rank, ungrouped
    all_teams = sorted(elo.keys(), key=lambda t: elo[t].get('global_rank') or 999)

    for r_idx, team in enumerate(all_teams, 3):
        elo_r = elo.get(team, {})
        ais_r = aiscore.get(team, {})
        fbr   = fbref.get(team, {})
        sfa_r = sofascore.get(team, {})
        conf  = CONF_MAP.get(team, '')

        total = elo_r.get('total_matches') or 0
        wins  = elo_r.get('wins') or 0
        wr    = round(wins / total * 100, 1) if total else None

        # AiScore values
        goals_str = ais_r.get('goals')
        goals_num = _num(goals_str)
        goals_pk  = _pk(goals_str)

        row_vals = [
            team, conf,
            # Elo
            elo_r.get('global_rank'), elo_r.get('rating'),
            elo_r.get('rating_1y_chg'), elo_r.get('rating_2y_chg'), elo_r.get('rating_5y_chg'),
            total, wins, elo_r.get('draws'), elo_r.get('losses'), wr,
            # AiScore
            goals_num, goals_pk,
            _num(ais_r.get('assists')),
            _num(ais_r.get('red_cards')),
            _num(ais_r.get('yellow_cards')),
            _num(ais_r.get('total_shots')),
            _num(ais_r.get('shots_on_target')),
            _num(ais_r.get('clearances')),
            _num(ais_r.get('key_passes')),
            _num(ais_r.get('crosses')),
            _num(ais_r.get('crosses_acc')),
            _num(ais_r.get('fouls')),
            # FBref
            fbr.get('xG (FBref)'), fbr.get('npxG (FBref)'), fbr.get('Posesie (%)'),
            fbr.get('Goluri'), fbr.get('GA (goluri primite)'), fbr.get('Salvări %'),
            # Sofascore
            sfa_r.get('meciuri_sfa'), sfa_r.get('xg_total'), sfa_r.get('xg_per_meci'),
            sfa_r.get('suturi_total_sfa'), sfa_r.get('posesie_medie'),
        ]

        conf_fill = CONF_FILLS.get(conf)
        for c_idx, val in enumerate(row_vals, 1):
            align = 'left' if c_idx == 1 else 'center'
            _set_val(ws, r_idx, c_idx, val, align=align, fill=conf_fill if c_idx <= 2 else None)

        # Color Elo rating
        elo_val = elo_r.get('rating') or 0
        if elo_val >= 2000:
            ws.cell(r_idx, 4).fill = GOLD_FILL
        elif elo_val >= 1900:
            ws.cell(r_idx, 4).fill = WIN_FILL
        elif elo_val >= 1800:
            ws.cell(r_idx, 4).fill = BLUE_FILL

        # Color Elo trends
        for col_offset, col_idx in enumerate([5, 6, 7]):
            chg = row_vals[2 + col_offset + 2]  # 1y=idx4, 2y=idx5, 5y=idx6 in row_vals
            if chg is not None:
                ws.cell(r_idx, col_idx).fill = WIN_FILL if chg > 0 else (LOSS_FILL if chg < 0 else PatternFill())

        # Color win rate
        if wr is not None:
            ws.cell(r_idx, 12).fill = (GOLD_FILL if wr >= 65 else WIN_FILL if wr >= 50 else
                                       DRAW_FILL if wr >= 40 else LOSS_FILL)

    ws.row_dimensions[r_idx + 1 if 'r_idx' in dir() else 3].height = 16

    # Set default row height for data rows
    for i in range(3, len(all_teams) + 3):
        ws.row_dimensions[i].height = 16


def build_v6():
    print('Loading data sources...')
    aiscore   = load_aiscore()
    elo_data  = load_elo()
    print(f'  AiScore: {len(aiscore)} teams')
    print(f'  Elo: {len(elo_data)} WC teams')

    print('Loading existing v5 workbook...')
    wb = openpyxl.load_workbook('statistici_calificari_cm2026_v5.xlsx')

    fbref    = load_fbref(wb)
    sofascore = load_sofascore_meciuri(wb)
    print(f'  FBref: {len(fbref)} teams')
    print(f'  Sofascore (meciuri cu date): {len(sofascore)} teams')

    # Remove if exists
    if 'Sumar_Complet' in wb.sheetnames:
        del wb['Sumar_Complet']

    print('Building Sumar_Complet...')
    add_sumar_complet(wb, aiscore, elo_data, fbref, sofascore)

    # Also add WC fixtures if not already there (copy from v5 which already has them)
    out = 'statistici_calificari_cm2026_v6.xlsx'
    wb.save(out)
    print(f'\nSaved: {out}')
    print(f'Sheets: {wb.sheetnames}')

    # Quick coverage check
    print('\n=== Coverage check (Sumar_Complet) ===')
    ws = wb['Sumar_Complet']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    no_aiscore = [r[0] for r in rows if r[12] is None]   # col 13 = goals
    no_elo     = [r[0] for r in rows if r[3] is None]    # col 4 = elo rating
    no_fbref   = [r[0] for r in rows if r[24] is None and r[1] == 'UEFA']  # col 25 = xG
    print(f'  Teams with Elo data: {len(rows) - len(no_elo)}/{len(rows)}')
    print(f'  Teams with AiScore data: {len(rows) - len(no_aiscore)}/{len(rows)}')
    if no_aiscore:
        print(f'  Missing AiScore: {no_aiscore}')
    print(f'  UEFA teams missing FBref xG: {no_fbref}')


if __name__ == '__main__':
    build_v6()
